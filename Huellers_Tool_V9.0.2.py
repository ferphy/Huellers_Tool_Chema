import tkinter
from tkinter import filedialog 
from tkinter import ttk
from tkinter import messagebox 
import pandas as pd
import numpy as np
from pandastable import Table, TableModel
import datetime
import os
import openpyxl
import shutil
import xlwt
from matplotlib import font_manager
from PIL import Image, ImageDraw, ImageFont
import re
from io import BytesIO
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#class

class Site:
    def __init__(self, name, tech, freq, nodes):
        self.name = name
        self.tech = tech
        self.freq =  freq
        self.nodes = nodes
        self.dfret = []
        self.dfretcheck = []
        self.textret = []
        self.techrets = []
        self.freqrets = []
        self.dfret = pd.DataFrame()
        self.dfretcheck = pd.DataFrame()

    def load_dfret(self,df,dfretcheck):
        self.dfret = df
        self.dfretcheck = dfretcheck
    def load_textret(self,text):
        self.textret = text
    def load_techfoundinrets(self,ok):
        self.techrets = ok
    def load_freqrets(self,f):
        self.freqrets = f

class Cluster:
    def __init__(self, name):
        self.name = name  
        self.sitesc = np.array([])
        self.sitess = np.array([])
        self.dfct = pd.DataFrame()
    def add_dfct(self,dfct):
        self.dfct = dfct
    def add_sitec(self,site):
        self.sitesc = np.append(self.sitesc,site)
    def add_sites(self,site):
        self.sitess = np.append(self.sitess,site)

def savetxt(s,dir):
    with open(dir+'.txt', 'w') as f:
        f.write(s)

def create_folder(fn):
    isExist = os.path.exists(fn)
    if not isExist:
        # Create a new directory because it does not exist
        os.makedirs(fn)

def read_clusters_txt(path):

    if (path == '') | (path[-12::] != 'Clusters.txt'):
        messagebox.showwarning(message="It is necessary to add in the same directory a Clusters.txt file", title="Warning")
        return
    try:
        with open(path,encoding="utf-8") as f:
            text = f.read()
    except IOError:
        messagebox.showwarning('Error: File does not appear to exist.')
        return 0
    c = text.split('\n')
    if '' in c:
        c.remove('')
    return c

def load_csv_cellscoring():
    global dfcellscoring
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (Thor cell scoring)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return
    dfcellscoring = pd.read_csv(filename,sep=';')
    return dfcellscoring

def load_csv_thorkeys(filename):
    thorkeys = []
    try:
        dftk = pd.read_csv(filename, sep=';')
        thorkeys = dftk.TKeys.tolist()
    except:
        messagebox.showwarning(message='No thorkeys csv file available: '+filename)
    return thorkeys

def load_balcsv():
    global dfbal
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (dfbal)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return
    dfbal = pd.read_csv(filename, sep=';')
def load_retcssv():
    global dfret
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (RETs)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return
    dfret = pd.read_csv(filename, sep=';')

def load_csv_ctable():
    global dfct
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (celltable)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return
    dfct = pd.read_csv(filename, sep=';')

def load_csv(title):
    try:
        filenames = filedialog.askopenfilenames(initialdir = os.getcwd(), 
        title = "Select a csv file " + title, 
        filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
        if len(filenames) == 0:
            return
        filenames = list(filenames)
        df = pd.read_csv(filenames[0], sep=';')
        if len(filenames) > 1:
            for fn in filenames[1::]:
                df = pd.concat([df,pd.read_csv(fn, sep=';')])
    except:
        messagebox.showwarning(message='No '+title+' csv file loaded: '+filename)
        df = pd.DataFrame()
    return df

def pasaranumero(colnum, df):
    dfn = df.copy()
    for c in colnum:
        dfn[c] = (dfn[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False).str.strip())
        dfn[c] = pd.to_numeric(dfn[c], errors='coerce')
    return dfn

def load_KPIs_3G():
    global dfkpi3g
    filenames = filedialog.askopenfilenames(initialdir = os.getcwd(), 
    title = "Select a csv file (KPIs 3G)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    filenames = list(filenames)
    if len(filenames) == 0:
        return
    #dfkpi3g = pd.read_csv(filenames[0],sep=';')
    #if len(filenames) > 1:
    #    for fn in filenames[1::]:
    #        dfkpi3g = pd.concat([dfkpi3g,pd.read_csv(fn, sep=';')])
    df3glist = []
    columnas = ['Date','RNC ID','RNC Name','NodeB ID','NodeB Name','Cell ID','Cell Name','Integrity','3G_QF_DCR_Voice(%)',
                '% CSSR CS HW(%)','% CSSR PS HW(%)','3G_QF_DCR_PS(%)','Voice SHO Success(#)','3G_QF_PS_HHO_Success_Rate(%)',
                '3G_QF_IRAT_3G_to_2G_Voice_HO (excluding preparation)(%)','QB_3G_TP (0.0 - 0.6 Km)(#)','QB_3G_TP (0.6 - 1.1 Km)(#)',
                'QB_3G_TP (1.1 - 2.2 Km)(#)','QB_3G_TP (2.2 - 3.6 Km)(#)','QB_3G_TP (3.6 - 6.0 Km)(#)','QB_3G_TP (>6 Km)(#)',
                '3G_QF_DL_Data_Traffic(kB)','3G_QF_Initiated_Calls(#)','3G_QF_IRAT_3G_to_2G_Voice_HO (excluding preparation)_Attempts(#)',
                '3G_QF_RAB_Drop_PS(#)','3G_QF_RAB_Drop_Voice(#)','3G_QF_RAB_PS_Release(#)','3G_QF_RAB_Released(#)','3G_QF_RSSI_UL(dBm)',
                '3G_QF_Speech (min)(s)','3G_QF_Terminated_Voice_Calls_excluding_IRAT_HO(#)','3G_QF_UL_Data_Traffic(kB)','3G_QF_Calls ending in 2G(%)',
                '3G_QF_Cell_Availability_Daily(%)','3G_QF_Cell_Availability_Hourly(%)','3G_QF_CSSR_CS(%)','3G_QF_CSSR_CS_Attempts(#)',
                '3G_QF_CSSR_CS_Control_Plane(%)','3G_QF_CSSR_CS_Control_Plane_Attempts(#)','3G_QF_CSSR_CS_Control_Plane_Success(#)',
                '3G_QF_CSSR_CS_User_Plane(%)','3G_QF_CSSR_CS_User_Plane_Attempts(#)','3G_QF_CSSR_CS_User_Plane_Success(#)','3G_QF_CSSR_PS(%)',
                '3G_QF_CSSR_PS_Attempts(#)','3G_QF_CSSR_PS_SP(%)','3G_QF_CSSR_PS_SP_Attempts(#)','3G_QF_CSSR_PS_SP_Success(#)',
                '3G_QF_CSSR_PS_UP(%)','3G_QF_CSSR_PS_UP_Attempts(#)','3G_QF_CSSR_PS_UP_Success(#)','3G_QF_Intensidad_Paging(Times/s)',
                '3G_QF_Voice_HHO_Attempts(#)','3G_QF_Voice_HHO_Success_Rate(%)','3G_QF_Voice_SHO_Attempts(#)','3G_QF_Voice_SHO_Success_Rate(%)',
                '3G_QF_Voice_Traffic(Erl)','VS.TP.UE.0','VS.TP.UE.1','VS.TP.UE.2','VS.TP.UE.3','VS.TP.UE.4','VS.TP.UE.5','VS.TP.UE.6.9',
                'VS.TP.UE.10.15','VS.TP.UE.16.25','VS.TP.UE.26.35','VS.TP.UE.36.55','VS.TP.UE.More55','3G_MeanDistance(#)','TP1 (0.0 - 0.3 Km)',
                'TP2 (0.3 - 0.7 Km)','TP3 (0.7 - 1.1 Km)','TP4 (1.1 - 2.2 Km)','TP5 (2.2 - 3.7 Km)','TP6 (3.7 - 6.2 Km)','TP7 (6.2 - 14.0 Km)','TP8 (>14.0 Km)','Site','Sector','Band','3G_MeanDistance']
    colnum = ['3G_QF_DCR_Voice(%)','% CSSR CS HW(%)', '% CSSR PS HW(%)', '3G_QF_DCR_PS(%)', 'Voice SHO Success(#)',
                '3G_QF_PS_HHO_Success_Rate(%)','3G_QF_IRAT_3G_to_2G_Voice_HO (excluding preparation)(%)','3G_QF_DL_Data_Traffic(kB)',
                '3G_QF_RSSI_UL(dBm)','3G_QF_Speech (min)(s)', '3G_QF_Terminated_Voice_Calls_excluding_IRAT_HO(#)',
                '3G_QF_UL_Data_Traffic(kB)', '3G_QF_Calls ending in 2G(%)',
                '3G_QF_Cell_Availability_Daily(%)', '3G_QF_Cell_Availability_Hourly(%)', '3G_QF_CSSR_CS(%)',
                '3G_QF_CSSR_CS_Attempts(#)','3G_QF_CSSR_CS_Control_Plane(%)', '3G_QF_CSSR_CS_Control_Plane_Attempts(#)',
                '3G_QF_CSSR_CS_Control_Plane_Success(#)','3G_QF_CSSR_CS_User_Plane(%)', '3G_QF_CSSR_CS_User_Plane_Attempts(#)',
                '3G_QF_CSSR_CS_User_Plane_Success(#)', '3G_QF_CSSR_PS(%)','3G_QF_CSSR_PS_Attempts(#)', '3G_QF_CSSR_PS_SP(%)', '3G_QF_CSSR_PS_SP_Attempts(#)',
                '3G_QF_CSSR_PS_SP_Success(#)','3G_QF_CSSR_PS_UP(%)', '3G_QF_CSSR_PS_UP_Attempts(#)', '3G_QF_CSSR_PS_UP_Success(#)',
                '3G_QF_Intensidad_Paging(Times/s)','3G_QF_Voice_HHO_Attempts(#)', '3G_QF_Voice_HHO_Success_Rate(%)', '3G_QF_Voice_SHO_Attempts(#)',
                '3G_QF_Voice_SHO_Success_Rate(%)','3G_QF_Voice_Traffic(Erl)', 'VS.TP.UE.0', 'VS.TP.UE.1', 'VS.TP.UE.2', 'VS.TP.UE.3', 'VS.TP.UE.4',
                'VS.TP.UE.5', 'VS.TP.UE.6.9','VS.TP.UE.10.15', 'VS.TP.UE.16.25', 'VS.TP.UE.26.35', 'VS.TP.UE.36.55', 'VS.TP.UE.More55', '3G_MeanDistance(#)',
                'TP1 (0.0 - 0.3 Km)','TP2 (0.3 - 0.7 Km)', 'TP3 (0.7 - 1.1 Km)', 'TP4 (1.1 - 2.2 Km)', 'TP5 (2.2 - 3.7 Km)','TP6 (3.7 - 6.2 Km)', 'TP7 (6.2 - 14.0 Km)', 'TP8 (>14.0 Km)','Sector']

    dictprov = {'V': 'VAL', 'B': 'BAL', 'U': 'MUR', 'A': 'AND', 'E': 'EXT', 'R': 'ARA', 'K': 'CYM', 'X': 'CLM', 'M': 'MAD', 'W': 'CAN'}
    for fn in filenames:
        df = pd.read_csv(fn, sep=';')
        if 'DIA' in df.columns:
            df['HORA'] = df['HORA'].astype(str).str.zfill(2)
            df['Date'] = pd.to_datetime(df['DIA'].astype(str) + df['HORA'], format='%Y%m%d%H')
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df.rename(columns={'SITE': 'Site'}, inplace=True)
            df.loc[:, 'Sector'] = [x[-2] for x in df['3G_UTRANCELL'].tolist()]
            df.loc[:, 'Band'] = [x[-3] for x in df['3G_UTRANCELL'].tolist()]
            df.rename(columns={'3G_UTRANCELL': 'Cell Name'}, inplace=True)
            df.rename(columns={'3G_MeanDistance(#)':'3G_MeanDistance'}, inplace=True)
            df['3G_MeanDistance'] = (df['3G_MeanDistance'].astype(str).str.replace(',', '.', regex=False).str.strip())
            df['3G_MeanDistance'] = pd.to_numeric(df['3G_MeanDistance'], errors='coerce')
        else:
            df.Date = pd.to_datetime(df.Date)
            df.loc[:, 'Site'] = [dictprov[x[0]] + x[1:5] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Sector'] = [x[-1] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Band'] = [x[-2] for x in df['Cell Name'].tolist()]
        for col in columnas:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[columnas]
        df3glist.append(df)
    dfkpi3g = pd.concat(df3glist, ignore_index=True)
    dfkpi3g = pasaranumero(colnum, dfkpi3g)
    dfkpi3g['3G_MeanDistance(#)'] = dfkpi3g[['3G_MeanDistance', '3G_MeanDistance(#)']].sum(axis=1, skipna=True)
    dfkpi3g.drop(columns=['3G_MeanDistance'], inplace=True)

    return dfkpi3g

def load_KPIs_4G():
    global dfkpi4g, dfct
    filenames = filedialog.askopenfilenames(initialdir = os.getcwd(), 
    title = "Select a csv file (KPIs 4G)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    filenames = list(filenames)
    if len(filenames) == 0:
        return
    #dfkpi4g = pd.read_csv(filenames[0],sep=';')
    #if len(filenames) > 1:
    #    for fn in filenames[1::]:
    #        dfkpi4g = pd.concat([dfkpi4g,pd.read_csv(fn, sep=';')])
    columnas = ['Date', 'eNodeB ID', 'eNodeB Name', 'Cell ID', 'Cell Name', 'Integrity', '4G_QF_VoLTE_DCR(%)', '4G_QF_VoLTE_CSSR(%)', '4G_QF_CSSR_PS_ERAB(%)', '4G_QF_DCR_PS(%)', '4G_QF_Exito HO Preparation 4G/3G SRVCC(%)', '4G_QF_Exito HO Execution 4G/3G SRVCC(%)', '4G_QF_CSFB_Success_Rate(%)', '4G_QF_IntraLTE HO SuccRate(%)', '4G_User_DL_Throughput(Mbps)(Mbps)', '4G_User_UL_Throughput(Mbps)(Mbps)', '%MIMO 4G(%)', 'PRB.DL.Usage.RATE(%)', 'PDCCH.Usage.RATE(%)', 'TA (0 - 0.6 Km)(n)', 'TA (0.6 - 1.1 Km) (#)', 'TA (1.1 - 2 Km)(#)', 'TA (2 - 3.5 Km)(n)', 'TA (3.5 - 6.7 Km)(n)', 'TA (>6.7 Km)(n)', 'MIMO_Rank2(%)', 'MIMO_Rank4(%)', '4G_QF_Availability(%)', '4G_QF_CA_Primary_Cell(%)', '4G_QF_CA_Secondary_Cell(%)', 'Downtime_Hua_Availability_Rate_Daily(%)', 'Downtime_Hua_Availability_Rate_Hourly(%)', 'CSFB attempts to WCDMA(Times)', '% ERAB Accessibility(ALL QCI) HW(%)', '% ERAB Accessibility(QCI6-9)(%)', '4G_QF_CSSR_PS_RRC(%)', '4G_QF_VoLTE_CSSR_Attempts(#)', '_Att_DCR_PS_Total_AllQCI_Hua(#)', '_Att_DCR_PS_Total_Hua(#)', '4G_QF_Downlink_Traffic_Volume(MB)', '4G_QF_Downlink_Traffic_Volume_GB(GB)', '4G_QF_Intensidad_Paging(%)', '4G_QF_IntraLTE HOSR (including preparation)(%)', '% VOLTE_Intrafreq_HOSR_HW(%)', '4G_QF_IRAT_HO_4G_to_3G_success (including preparation)(%)', 'IRAT HO 4G to 3G success (excluding preparation)(%)', 'IRAT HO 4G to 3G success (including preparation)(%)', '4G_QF_UL_PUSCH_Interference(dBm)', '4G_HW_Avg_RRC_Connected_Users(Number)', 'L.TA.UE.Index0', 'L.TA.UE.Index1', 'L.TA.UE.Index2', 'L.TA.UE.Index3', 'L.TA.UE.Index4', 'L.TA.UE.Index5', 'L.TA.UE.Index6', 'L.ChMeas.PRB.DL.Used.Avg', 'L.ChMeas.PRB.DL.Avail', 'Average coverage distance(m)', 'TA1 (0.0 - 0.3 Km)', 'TA2 (0.3 - 0.6 Km)', 'TA3 (0.6 - 1.0 Km)', 'TA4 (1.0 - 2.2 Km)', 'TA5 (2.2 - 3.6 Km)', 'TA6 (3.6 - 6.7 Km)', 'TA7 (6.7 - 14.5 Km)', 'TA8 (>14.5 Km)','Site','Sector','Site-Sector','Band','Average coverage distance']
    colnum = ['4G_QF_VoLTE_DCR(%)', '4G_QF_VoLTE_CSSR(%)', '4G_QF_CSSR_PS_ERAB(%)', '4G_QF_DCR_PS(%)',
               '4G_QF_Exito HO Preparation 4G/3G SRVCC(%)', '4G_QF_Exito HO Execution 4G/3G SRVCC(%)',
               '4G_QF_CSFB_Success_Rate(%)', '4G_QF_IntraLTE HO SuccRate(%)', '4G_User_DL_Throughput(Mbps)(Mbps)',
               '4G_User_UL_Throughput(Mbps)(Mbps)', '%MIMO 4G(%)', 'PRB.DL.Usage.RATE(%)', 'PDCCH.Usage.RATE(%)',
               'TA (0 - 0.6 Km)(n)', 'TA (0.6 - 1.1 Km) (#)', 'TA (1.1 - 2 Km)(#)', 'TA (2 - 3.5 Km)(n)',
               'TA (3.5 - 6.7 Km)(n)', 'TA (>6.7 Km)(n)', 'MIMO_Rank2(%)', 'MIMO_Rank4(%)', '4G_QF_Availability(%)',
               '4G_QF_CA_Primary_Cell(%)', '4G_QF_CA_Secondary_Cell(%)', 'Downtime_Hua_Availability_Rate_Daily(%)',
               'Downtime_Hua_Availability_Rate_Hourly(%)', 'CSFB attempts to WCDMA(Times)',
               '% ERAB Accessibility(ALL QCI) HW(%)', '% ERAB Accessibility(QCI6-9)(%)', '4G_QF_CSSR_PS_RRC(%)',
               '4G_QF_VoLTE_CSSR_Attempts(#)', '_Att_DCR_PS_Total_AllQCI_Hua(#)', '_Att_DCR_PS_Total_Hua(#)',
               '4G_QF_Downlink_Traffic_Volume(MB)', '4G_QF_Downlink_Traffic_Volume_GB(GB)',
               '4G_QF_Intensidad_Paging(%)', '4G_QF_IntraLTE HOSR (including preparation)(%)',
               '% VOLTE_Intrafreq_HOSR_HW(%)', '4G_QF_IRAT_HO_4G_to_3G_success (including preparation)(%)',
               'IRAT HO 4G to 3G success (excluding preparation)(%)',
               'IRAT HO 4G to 3G success (including preparation)(%)', '4G_QF_UL_PUSCH_Interference(dBm)',
               '4G_HW_Avg_RRC_Connected_Users(Number)', 'L.TA.UE.Index0', 'L.TA.UE.Index1', 'L.TA.UE.Index2',
               'L.TA.UE.Index3', 'L.TA.UE.Index4', 'L.TA.UE.Index5', 'L.TA.UE.Index6', 'L.ChMeas.PRB.DL.Used.Avg',
               'L.ChMeas.PRB.DL.Avail', 'Average coverage distance(m)','TA1 (0.0 - 0.3 Km)', 'TA2 (0.3 - 0.6 Km)',
               'TA3 (0.6 - 1.0 Km)', 'TA4 (1.0 - 2.2 Km)', 'TA5 (2.2 - 3.6 Km)', 'TA6 (3.6 - 6.7 Km)',
               'TA7 (6.7 - 14.5 Km)', 'TA8 (>14.5 Km)','Sector']

    dicv = {
        0: 39, 1: 195, 2: 429, 3: 819, 4: 1521, 5: 2769,
        6: 5109, 7: 10569, 8: 22269, 9: 41769, 10: 65169
    }
    df4glist = []
    for fn in filenames:
        df = pd.read_csv(fn, sep=';')
        if 'HORA' in df.columns:
            
            df.FECHA = pd.to_datetime(df.FECHA, dayfirst=True)
            df['HORA'] = df['HORA'].str.split(':').str[0].str.zfill(2)
            df['Date'] = pd.to_datetime(df['FECHA'].dt.strftime('%Y%m%d') + df['HORA'], format='%Y%m%d%H')
            df = df.merge(dfct[['CELLNAME', 'NODE']], on='CELLNAME', how='left')
            df.rename(columns={'NODE': 'eNodeB Name'}, inplace=True)
            df.rename(columns={'SITE': 'Site'}, inplace=True)
            df.rename(columns={'4G PRB USAGE (E4GPU001)': 'PRB.DL.Usage.RATE(%)'}, inplace=True)
            df.rename(columns={'PDCCH USAGE (E4GPD001)': 'PDCCH.Usage.RATE(%)'}, inplace=True)
            df.rename(columns={'4G Avg PDCP SDU DL/UL Throughput (Mbps) (E4GTDL001)': '4G_User_DL_Throughput(Mbps)(Mbps)'}, inplace=True)
            df.loc[:, 'Sector'] = [x[-2] for x in df['CELLNAME'].tolist()]
            df.loc[:, 'Site-Sector'] = [x[0:3] + x[4:8] + '-' + x[9] for x in df['CELLNAME'].tolist()]
            df.loc[:, 'Band'] = [x[-3] for x in df['CELLNAME'].tolist()]
            df.rename(columns={'CELLNAME': 'Cell Name'}, inplace=True)
            for i in range(30):
                col = f'Pmtainit2distr {i}'
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.').str.strip(), errors='coerce')
            numerador = sum(df[f'Pmtainit2distr {i}'].fillna(0) * valor for i, valor in dicv.items())
            numerador += df[[f'Pmtainit2distr {i}' for i in range(11, 30)]].fillna(0).sum(axis=1) * 76830
            denominador = df[[f'Pmtainit2distr {i}' for i in range(30)]].fillna(0).sum(axis=1)
            df['Average coverage distance'] = numerador / denominador
        else:
            df.Date = pd.to_datetime(df.Date, dayfirst=True)
            df.loc[:, 'Site'] = [x[0:3] + x[4:8] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Sector'] = [x[9] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Site-Sector'] = [x[0:3] + x[4:8] + '-' + x[9] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Band'] = [x[8] for x in df['Cell Name'].tolist()]
        for col in columnas:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[columnas]
        df4glist.append(df)
    dfkpi4g = pd.concat(df4glist, ignore_index=True)
    dfkpi4g = pasaranumero(colnum, dfkpi4g)
    dfkpi4g['Average coverage distance(m)'] = dfkpi4g[['Average coverage distance', 'Average coverage distance(m)']].sum(axis=1, skipna=True)
    dfkpi4g.drop(columns=['Average coverage distance'], inplace=True)

    return dfkpi4g

def load_KPIs_5G():
    global dfkpi5g
    filenames = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (KPIs 5G)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    filenames = list(filenames)
    if len(filenames) == 0:
        return
    dfkpi5g = pd.read_csv(filenames[0],sep=';')
    if len(filenames) > 1:
        for fn in filenames[1::]:
            dfkpi5g = pd.concat([dfkpi5g,pd.read_csv(fn, sep=';')])
    return dfkpi5g

def load_KPIs_Huella_5G():
    global dfkpih5g
    filenames = filedialog.askopenfilenames(initialdir = os.getcwd(), 
    title = "Select a csv file (KPIs 5G Huella)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filenames) == 0:
        return
    '''
    dfkpih5g = pd.read_csv(filenames[0],sep=';')
    if len(filenames) > 1:
        for fn in filenames[1::]:
            dfkpih5g = pd.concat([dfkpih5g,pd.read_csv(fn, sep=';')])
    '''
    columnas = ['Date','gNodeB ID','gNodeB Name','Cell ID','Cell Name','Integrity','N.PRB.DL.Used.Avg','AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp','eN.PRB.DL.Used.Avg','5G_Hua_PRB_Use_DL(%)','5G_QF MIMO (Rank2)(%)','5G_QF MIMO (Rank4)(%)','Average PDCCH Load(%)','UL Average Interference Noise(dBm)','N.PRB.DL.Avail.Avg','N.RA.TA.UE.Index0','N.RA.TA.UE.Index1','N.RA.TA.UE.Index2','N.RA.TA.UE.Index3','N.RA.TA.UE.Index4','N.RA.TA.UE.Index5','N.RA.TA.UE.Index6','N.RA.TA.UE.Index7','N.RA.TA.UE.Index8','N.RA.TA.UE.Index9','N.RA.TA.UE.Index10','N.RA.TA.UE.Index11','N.RA.TA.UE.Index12','5G_QF DL Traffic Volume(GB)','H5GTDL001: Average DL User Throughput (Mbps) v138','H5GTUL001: Average UL User Throughput (Mbps) v138','H5GD001: % 5G/NR Drop Call Rate NSA v138','H5GSAD001: % 5G/NR SA Drop Call Rate v138','H5GCAVAIL001: % Cell Availability V138','H5GA001: % 5G/NR Setup Success Rate NSA v138','H5GD004: % 5G/NR Drop Call Rate (gNB releases) NSA v138','TP1 (0.0 - 0.3 Km)','TP2 (0.3 - 0.7 Km)','TP3 (0.7 - 1.1 Km)','TP4 (1.1 - 2.2 Km)','TP5 (2.2 - 3.6 Km)','TP6 (3.6 - 6.2 Km)','TP7 (6.2 - 14.5 Km)','TP8 (>14.5 Km)','PRB.DL','PRB.DL(%)','Site','Sector','Site-Sector','Band']
    colnum = ['5G_Hua_PRB_Use_DL(%)', 'Average PDCCH Load(%)',
                'UL Average Interference Noise(dBm)', 'AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp', 'N.RA.TA.UE.Index0', 'N.RA.TA.UE.Index1',
                'N.RA.TA.UE.Index2', 'N.RA.TA.UE.Index3', 'N.RA.TA.UE.Index4', 'N.RA.TA.UE.Index5', 'N.RA.TA.UE.Index6',
                'N.RA.TA.UE.Index7', 'N.RA.TA.UE.Index8', 'N.RA.TA.UE.Index9', 'N.RA.TA.UE.Index10',
                'N.RA.TA.UE.Index11', 'N.RA.TA.UE.Index12', '5G_QF DL Traffic Volume(GB)',
                'H5GTDL001: Average DL User Throughput (Mbps) v138',
                'H5GTUL001: Average UL User Throughput (Mbps) v138', 'H5GD001: % 5G/NR Drop Call Rate NSA v138',
                'H5GSAD001: % 5G/NR SA Drop Call Rate v138', 'H5GCAVAIL001: % Cell Availability V138',
                'H5GA001: % 5G/NR Setup Success Rate NSA v138', 'eN.PRB.DL.Used.Avg',
                'H5GD004: % 5G/NR Drop Call Rate (gNB releases) NSA v138', 'TP1 (0.0 - 0.3 Km)', 'TP2 (0.3 - 0.7 Km)',
                'TP3 (0.7 - 1.1 Km)', 'TP4 (1.1 - 2.2 Km)', 'TP5 (2.2 - 3.6 Km)', 'TP6 (3.6 - 6.2 Km)',
                'TP7 (6.2 - 14.5 Km)', 'TP8 (>14.5 Km)', 'TP1 (0.0 - 0.3 Km)', 'TP2 (0.3 - 0.7 Km)',
               'TP3 (0.7 - 1.1 Km)', 'TP4 (1.1 - 2.2 Km)', 'TP5 (2.2 - 3.6 Km)', 'TP6 (3.6 - 6.2 Km)',
               'TP7 (6.2 - 14.5 Km)', 'TP8 (>14.5 Km)', 'PRB.DL(%)','Sector']

    df5glist = []
    for fn in filenames:
        df = pd.read_csv(fn, sep=';')
        if 'Dia' in df.columns:
            #df.Dia = pd.to_datetime(df.Dia, yearfirst=True)
            df['HORA.'] = df['HORA.'].str.split(':').str[0].str.zfill(2)
            df['Date'] = pd.to_datetime(df['Dia'].astype(str) + df['HORA.'], format='%Y%m%d%H')
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df.rename(columns={'SITE': 'Site'}, inplace=True)
            df['PRB.DL(%)'] = df['5G PRB Use']
            df.rename(columns={'5G PRB Use': '5G_Hua_PRB_Use_DL(%)'}, inplace=True)
            df.loc[:, 'Sector'] = [x[-2] for x in df['5G_GCELDA'].tolist()]
            df.loc[:, 'Site-Sector'] = [x[0:3] + x[4:8] + '-' + x[9] for x in df['5G_GCELDA'].tolist()]
            df.loc[:, 'Band'] = [x[-3] for x in df['5G_GCELDA'].tolist()]
            df.rename(columns={'5G_GCELDA': 'Cell Name'}, inplace=True)
            df.rename(columns={'N.PRB.DL.Used.Avg': 'eN.PRB.DL.Used.Avg'}, inplace=True)
        else:
            df.Date = pd.to_datetime(df.Date)
            df.loc[:, 'Site'] = [x[0:3] + x[4:8] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Sector'] = [x[9] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Site-Sector'] = [x[0:3] + x[4:8] + '-' + x[9] for x in df['Cell Name'].tolist()]
            df.loc[:, 'Band'] = [x[8] for x in df['Cell Name'].tolist()]
            df['N.PRB.DL.Used.Avg'] = df['N.PRB.DL.Used.Avg'].str.replace(',', '.')
            df['N.PRB.DL.Avail.Avg'] = df['N.PRB.DL.Avail.Avg'].str.replace(',', '.')
            df['N.PRB.DL.Used.Avg'] = pd.to_numeric(df['N.PRB.DL.Used.Avg'], errors='coerce')
            df['N.PRB.DL.Avail.Avg'] = pd.to_numeric(df['N.PRB.DL.Avail.Avg'], errors='coerce')
            df['PRB.DL'] = 100 * df['N.PRB.DL.Used.Avg'] / df['N.PRB.DL.Avail.Avg']
        for col in columnas:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[columnas]
        df5glist.append(df)
    dfkpih5g = pd.concat(df5glist, ignore_index=True)
    dfkpih5g = pasaranumero(colnum, dfkpih5g)
    dfkpih5g['PRB.DL(%)'] = dfkpih5g[['PRB.DL(%)', 'PRB.DL']].sum(axis=1, skipna=True)
    dfkpih5g.drop(columns=['PRB.DL'], inplace=True)
    dfkpih5g['N.PRB.DL.Avail.Avg'] = dfkpih5g[['N.PRB.DL.Avail.Avg', 'AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp']].sum(axis=1, skipna=True)
    dfkpih5g.drop(columns=['AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp'], inplace=True)
    dfkpih5g['N.PRB.DL.Used.Avg'] = dfkpih5g[['N.PRB.DL.Used.Avg', 'eN.PRB.DL.Used.Avg']].sum(axis=1, skipna=True)
    dfkpih5g.drop(columns=['eN.PRB.DL.Used.Avg'], inplace=True)

    #dfkpih5g.rename(columns={'AVERAGE PRB LOAD DL: E5GPRBDL003: Num_DL_PRBs_Disp': 'N.PRB.DL.Avail.Avg'}, inplace=True)
    #dfkpih5g.rename(columns={'eN.PRB.DL.Used.Avg': 'N.PRB.DL.Used.Avg'}, inplace=True)

    return dfkpih5g

def browseFiles_import_excel():
    filename = filedialog.askopenfilenames(initialdir = os.getcwd(), 
    title = "Import excel tracking", 
    filetypes = (("xlsx file","*.xlsx*"), ("all files", "*.*"))) 
    return filename[0]

def browseFiles_import_csv(text):
    filename = filedialog.askopenfilenames(initialdir = os.getcwd(), 
    title = "Import csv tracking "+text, 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    return filename[0]

def import_excel():
    filenameci = browseFiles_import_csv('CI')
    try:
        dftrack_CI = pd.read_csv(filenameci,sep=';')
    except Exception as e:
        messagebox.showwarning(message=e)
        return 0
    return dftrack_CI

def browseFiles_celltable(): 
    global filename
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file celltable", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return
    return filename

def load_csv_celltable():
    filename = browseFiles_celltable()
    dfct = pd.read_csv(filename, sep=';')
    indexband = (dfct.STATUS == 1)&((dfct['OPERATOR NAME'] == 'ORANGE')|(dfct['OPERATOR NAME'] == 'Orange'))
    dfct['Band'] = ''
    bands = dfct.loc[indexband,'CELLNAMEX'].tolist()
    for i in range(len(bands)):
        if len(bands[i]) >= 11:
            bands[i] = bands[i][8]
    dfct.loc[indexband,'Band'] = bands
    return dfct

def checkformatsite(sites):
    sitesout = []
    for s in sites:
        if len(s)>=7:
            s = s[0:7]
            #check digits and upper
            if(s[0:3].isupper())&(s[3::].isdigit()):
                sitesout.append(s)
    return sitesout

def dateformatB(dateA):
    return dateA[6::]+'-'+dateA[3:5]+'-'+dateA[0:2]

def tech2freq(tech,zone):
    if zone == 'ZR':
        freqZR = {'S':700,'Q':700,'J':800,'G':900,'U':900,'K':1800,'M':2100,'W':2100,'L':2600,'X':2600,'P':3500}
        freq = np.zeros(len(tech))
        for i,t in enumerate(tech):
             freq[i] = freqZR[t]
    if zone == 'ZN':
        freqZN = {'Y':700,'Q':700,'M':800,'E':900,'F':900,'N':1800,'R':1800,'T':2100,'B':2100,'C':2100,'D':2100,'W':2100,'L':2600,'X':2600,'P':3500}
        freq = np.zeros(len(tech))
        for i,t in enumerate(tech):
            if t in freqZN.keys():
                freq[i] = freqZN[t]
    freq = freq[freq!=0]
    
    return np.unique(freq).astype('int16')

def extrae_info_clusters(clustersinput):
    if not 'dftrack_CI' in globals():
        #1. Read tracking
        global dftrack_CI
        dftrack_CI = import_excel()
    if not 'dfct' in globals():
        global dfct
        dfct = load_csv_celltable()
    clusters = np.array([])
    #Control of clusters and sites
    global dfclsit
    dfclsit = pd.DataFrame(columns=['Cluster','Seed Site','Corona Site'])
    for i,cluster in enumerate(clustersinput):
        dfclsit.loc[i,'Cluster'] = cluster
        #Search cluster
        sitesci = dftrack_CI.loc[dftrack_CI.loc[:,'CLUSTER NAME'] == cluster,'SITES CORONA'].tolist()
        sitessi = dftrack_CI.loc[dftrack_CI.loc[:,'CLUSTER NAME'] == cluster,'SITES SEMILLA'].tolist()
        if (len(sitesci) > 0) and (type(sitesci[0])!=float):
            sitesci = sitesci[0].split(",")
            sitesci=checkformatsite(sitesci)
            if len(sitesci) == 0:
                print('Not corona sites found for cluster '+cluster)
            #elif type(sitesci[0])==float:
            #    print('Not corona sites found for cluster '+cluster)
            #    sitesci = []
        else:
            if (len(sitesci) > 0):
                sitesci = []
            print('Not corona sites found for cluster '+cluster)
        
        if (len(sitessi) > 0) and (type(sitessi[0])!=float):
            sitessi = sitessi[0].split(",")
            sitessi=checkformatsite(sitessi)
            if len(sitessi) == 0:
                print('Not seed sites found for cluster '+cluster)
            #elif type(sitessi[0])==float:
            #    print('Not seed sites found for cluster '+cluster)
            #    sitessi = []
        else:
            if (len(sitessi) > 0):
                sitessi = []
            print('Not seed sites found for cluster '+cluster)
        sitessi = np.unique(np.sort(sitessi))
        sitesci = np.unique(np.sort([s for s in sitesci if s not in sitessi]))
        dfclsit.loc[i,'Seed Site'] = sitessi
        dfclsit.loc[i,'Corona Site'] = sitesci
        #Create Cluster
        clusters = np.append(clusters,Cluster(cluster))
        #Add Sites
        for s in sitesci:
            clusters[i].add_sitec(Site(s,[],[],[]))
        for s in sitessi:
            clusters[i].add_sites(Site(s,[],[],[]))
        #celltable cluster
        sites = np.concatenate((sitessi,sitesci))
        sitnfct = ''
        if len(sites) > 0:
            dfaux = pd.DataFrame()
            #Obtain data frame
            for s in sites:
                if ')' in s:
                    break
                indexdf = (dfct.SITE==s) & (dfct.STATUS == 1) & ((dfct['OPERATOR NAME'] == 'ORANGE')|(dfct['OPERATOR NAME'] == 'Orange')) 
                if not any(indexdf):
                    sitnfct+=s+'\n'
                dfaux=pd.concat([dfaux,dfct.loc[indexdf,:]],ignore_index=True)
            #delete cells B
            #indexbclassok = np.full((len(dfaux)), True)
            #bandclass = [i for i,e in enumerate(dfaux.CELLNAMEX.tolist()) if e[-1] == 'B']
            #indexbclassok[bandclass] = False
            #dfaux = dfaux[indexbclassok]
            clusters[i].add_dfct(dfaux)
        if len(sitnfct) > 0:
            savetxt(sitnfct,'Sites_not_found_in_'+cluster)
    return clusters

def filtkpi4gMR(dfkpi4g,cdfct):
    #filter
    missingsiteskpis4g = ''
    sites = np.unique(cdfct.SITE.tolist())
    indok = [False for x in range(len(dfkpi4g))]
    for s in sites:
        indit = dfkpi4g.Site == s
        if any(indit):
            indok = indok| indit
        else:
            missingsiteskpis4g += s+'\n'
    dfkpi4gfilt = dfkpi4g.loc[indok,:].copy()
    dfkpi4gfilt.reset_index(inplace=True)
    #fixformat
    """
    for k in dfkpi4gfilt.keys():
        if (type(dfkpi4gfilt[k][0])==str) and ((dfkpi4gfilt[k].str.replace(',','')[0].isnumeric())or (dfkpi4gfilt[k][0] == '/0')):
            dfkpi4gfilt[k] = dfkpi4gfilt[k].str.replace(',','.')
            dfkpi4gfilt[k] = pd.to_numeric(dfkpi4gfilt[k],errors='coerce')  
    """  
    dfkpi4gfilt.fillna('/0', inplace = True)
    return dfkpi4gfilt,missingsiteskpis4g  


def process_cellscoring_provincial():
    #filenamethorkeys = 'cell_score_provincial\\ThorKeysProvincial.csv'
    #thorkeys = load_csv_thorkeys(filenamethorkeys)
    #if len(thorkeys) == 0:
    #    return 0
    dfcellscoring = load_csv_cellscoring()
    #dlt errors
    poserror = [x for x in range(len(dfcellscoring)) if len(dfcellscoring.loc[x,'cellname_sin_X'])<8]
    okcellscoring = np.array([True for x in range(len(dfcellscoring))])
    okcellscoring[poserror] = False
    dfcellscoring = dfcellscoring.loc[okcellscoring,:]
    dfcellscoring.reset_index(inplace=True)
    dfcellscoring.loc[:,'Site'] = [x[0:7] for x in dfcellscoring['cellname_sin_X'].tolist()]
    dfcellscoring.loc[:,'Sector'] = [x[8] for x in dfcellscoring['cellname_sin_X'].tolist()]
    dfcellscoring.loc[:,'Band'] = [x[7] for x in dfcellscoring['cellname_sin_X'].tolist()]

    indexzona = ((dfcellscoring.loc[:,'ZONA'] == 'Zona ESTE')|(dfcellscoring.loc[:,'ZONA'] == 'Zona SUR')|(fcellscoring.loc[:,'ZONA'] == 'Zona CENTRO'))

    indexgisko = (dfcellscoring.loc[:,'GISKO'] != 0) & indexzona
    indexCore_Inner_num = (dfcellscoring.loc[:,'Core_Inner_num'] != 0) & indexzona
    indexCore_Outter_num = (dfcellscoring.loc[:,'Core_Outer_num'] != 0) & indexzona
    indexCore_Areas_num = (dfcellscoring.loc[:,'Core_Areas_num'] != 0) & indexzona
    indexCore_Locno_num = (dfcellscoring.loc[:,'Core_Locno_num'] != 0) & indexzona
    indexMercator_isolated_2Gbsc_num = (dfcellscoring.loc[:,'Mercator_isolated_2Gbsc_num'] != 0) & indexzona
    indexMercator_isolated_2Glrac_num = (dfcellscoring.loc[:,'Mercator_isolated_2Glrac_num'] != 0) & indexzona
    indexMercator_isolated_3Grnc_num = (dfcellscoring.loc[:,'Mercator_isolated_3Grnc_num'] != 0) & indexzona
    indexMercator_isolated_3Glrac_num = (dfcellscoring.loc[:,'Mercator_isolated_3Glrac_num'] != 0) & indexzona
    indexMercator_isolated_4G_num = (dfcellscoring.loc[:,'Mercator_isolated_4Gtac_num'] != 0) & indexzona
    indexSkadi_xfeeder_CLUSTER_UMTS_num = (dfcellscoring.loc[:,'Skadi_xfeeder_UMTS_num'] != 0) & indexzona
    indexSkadi_xfeeder_CLUSTER_LTE_num = (dfcellscoring.loc[:,'Skadi_xfeeder_LTE_num'] != 0) & indexzona
    colfixed = ['node','node_category','cellnamex','cellname_sin_X','cellname','clusters','PROVINCIA','ZONA','TECH']

    with pd.ExcelWriter('GISKO_CORE_ISOLATED_XFEEDER_W'+str(datetime.datetime.today().isocalendar()[1])+'.xlsx') as writer: 
        dfcellscoring.loc[indexgisko,colfixed+['GISKO']].to_excel(writer,sheet_name='GISKO',index=False)
        dfcellscoring.loc[indexCore_Inner_num,colfixed+['Core_Inner_num']].to_excel(writer,sheet_name='Core_Inner_num',index=False)
        dfcellscoring.loc[indexCore_Outter_num,colfixed+['Core_Outer_num']].to_excel(writer,sheet_name='Core_Outer_num',index=False)
        dfcellscoring.loc[indexCore_Areas_num,colfixed+['Core_Areas_num']].to_excel(writer,sheet_name='Core_Areas_num',index=False)
        dfcellscoring.loc[indexCore_Locno_num,colfixed+['Core_Locno_num']].to_excel(writer,sheet_name='Core_Locno_num',index=False)
        dfcellscoring.loc[indexMercator_isolated_2Gbsc_num,colfixed+['Mercator_isolated_2Gbsc_num']].to_excel(writer,sheet_name='Mercator_isolated_2Gbsc_num',index=False)
        dfcellscoring.loc[indexMercator_isolated_2Glrac_num,colfixed+['Mercator_isolated_2Glrac_num']].to_excel(writer,sheet_name='Mercator_isolated_2Glrac_num',index=False)
        dfcellscoring.loc[indexMercator_isolated_3Grnc_num,colfixed+['Mercator_isolated_3Grnc_num']].to_excel(writer,sheet_name='Mercator_isolated_3Grnc_num',index=False)
        dfcellscoring.loc[indexMercator_isolated_3Glrac_num,colfixed+['Mercator_isolated_3Glrac_num']].to_excel(writer,sheet_name='Mercator_isolated_3Glrac_num',index=False)
        dfcellscoring.loc[indexMercator_isolated_4G_num,colfixed+['Mercator_isolated_4Gtac_num']].to_excel(writer,sheet_name='Mercator_isolated_4Gtac_num',index=False)
        dfcellscoring.loc[indexSkadi_xfeeder_CLUSTER_UMTS_num,colfixed+['Skadi_xfeeder_UMTS_num']].to_excel(writer,sheet_name='Skadi_xfeeder_UMTS_num',index=False)
        dfcellscoring.loc[indexSkadi_xfeeder_CLUSTER_LTE_num,colfixed+['Skadi_xfeeder_LTE_num']].to_excel(writer,sheet_name='Skadi_xfeeder_LTE_num',index=False)
    messagebox.showinfo(message="Processed Cell Scoring", title="Export Clusters")

def export_clusters():
    if not 'clusters' in globals():
        global clusters
        clusters = extrae_info_clusters(clustersinput) 
    fname = 'Clusters PRBs Huella'
    create_folder(fname)
    create_folder(fname+'\\ALL')
    if (not os.path.isfile(fname+'\\ALL\\clustersites.xlsx')):
        dfclsit.to_excel(fname+'\\ALL\\clustersites.xlsx',index=False)
    if okHuella.get():
        filenamethorkeys = 'ThorKeys.csv'
        thorkeys = load_csv_thorkeys(filenamethorkeys)
        if len(thorkeys) == 0:
            return 0
        dfcellscoring = load_csv_cellscoring()
        #dlt errors
        poserror = [x for x in range(len(dfcellscoring)) if len(dfcellscoring.loc[x,'cellname_sin_X'])<8]
        okcellscoring = np.array([True for x in range(len(dfcellscoring))])
        okcellscoring[poserror] = False
        dfcellscoring = dfcellscoring.loc[okcellscoring,:]
        dfcellscoring.reset_index(inplace=True)
        dfcellscoring.loc[:,'Site'] = [x[0:7] for x in dfcellscoring['cellname_sin_X'].tolist()]
        dfcellscoring.loc[:,'Sector'] = [x[8] for x in dfcellscoring['cellname_sin_X'].tolist()]
        dfcellscoring.loc[:,'Band'] = [x[7] for x in dfcellscoring['cellname_sin_X'].tolist()]
        dfkpi3g = load_KPIs_3G()
        #Aquí hay que distinguir y meter en los kpis de micro la fecha+hora
        '''
        for i in len(dfkpi3g):
            try:
                dfkpi3g.Date = pd.to_datetime(dfkpi3g.Date)
                dictprov = {'V': 'VAL', 'B': 'BAL', 'U': 'MUR', 'A': 'AND', 'E': 'EXT', 'R': 'ARA', 'K': 'CYM', 'X': 'CLM', 'M': 'MAD', 'W': 'CAN'}
                dfkpi3g.loc[:, 'Site'] = [dictprov[x[0]] + x[1:5] for x in dfkpi3g['Cell Name'].tolist()]
                dfkpi3g.loc[:, 'Sector'] = [x[-1] for x in dfkpi3g['Cell Name'].tolist()]
                dfkpi3g.loc[:, 'Band'] = [x[-2] for x in dfkpi3g['Cell Name'].tolist()]
            except:
                if i == 0:
                    printf('No hay datos 3g:HW')

            try:
                dfkpi3g['HORA'] = dfkpi3g['HORA'].astype(str).str.zfill(2)
                dfkpi3g[:, 'Date'] = pd.to_datetime(dfkpi3g['DIA'].astype(str) + dfkpi3g['HORA'], format='%Y%m%d%H')
                dfkpi3g.rename(columns={'SITE':'Site'}, inplace=True)
                dfkpi3g.loc[:, 'Sector'] = [x[-2] for x in dfkpi3g['3G_UTRANCELL'].tolist()]
                dfkpi3g.loc[:, 'Band'] = [x[-3] for x in dfkpi3g['3G_UTRANCELL'].tolist()]
            except:
                if i == 1:
                    printf('No hay datos 3g:Eric')
        '''
        dfdb = load_csv('DataBase')
        if len(dfdb) == 0:
            dfdb = pd.DataFrame(columns=['CLUSTER','CELLNAMEX','DIST','COMMENTS','DATE'])
    if okPRBs.get() | okHuella.get() | okgmr.get():
        dfkpi4g = load_KPIs_4G()
    '''
        dfkpi4g.Date = pd.to_datetime(dfkpi4g.Date,dayfirst=True)
        dfkpi4g.loc[:,'Site'] = [x[0:3]+x[4:8] for x in dfkpi4g['Cell Name'].tolist()]
        dfkpi4g.loc[:,'Sector'] = [x[9] for x in dfkpi4g['Cell Name'].tolist()]
        dfkpi4g.loc[:,'Site-Sector'] = [x[0:3]+x[4:8]+'-'+x[9] for x in dfkpi4g['Cell Name'].tolist()]
        dfkpi4g.loc[:,'Band'] = [x[8] for x in dfkpi4g['Cell Name'].tolist()]
    '''
    """
    if okPRBs.get():
        dfkpi5g = load_KPIs_5G()
        dfkpi5g.Date = pd.to_datetime(dfkpi5g.Date)
        dfkpi5g.loc[:,'Site'] = [x[0:3]+x[4:8] for x in dfkpi5g['Cell Name'].tolist()]
        dfkpi5g.loc[:,'Sector'] = [x[9] for x in dfkpi5g['Cell Name'].tolist()]
        dfkpi5g.loc[:,'Site-Sector'] = [x[0:3]+x[4:8]+'-'+x[9] for x in dfkpi5g['Cell Name'].tolist()]
        dfkpi5g.loc[:,'Band'] = [x[8] for x in dfkpi5g['Cell Name'].tolist()]
    """
    if okPRBs.get() | okHuella.get():
        dfkpih5g = load_KPIs_Huella_5G()
        '''
        dfkpih5g.Date = pd.to_datetime(dfkpih5g.Date)
        dfkpih5g.loc[:,'Site'] = [x[0:3]+x[4:8] for x in dfkpih5g['Cell Name'].tolist()]
        dfkpih5g.loc[:,'Sector'] = [x[9] for x in dfkpih5g['Cell Name'].tolist()]
        dfkpih5g.loc[:,'Site-Sector'] = [x[0:3]+x[4:8]+'-'+x[9] for x in dfkpih5g['Cell Name'].tolist()]
        dfkpih5g.loc[:,'Band'] = [x[8] for x in dfkpih5g['Cell Name'].tolist()]
        dfkpih5g['N.PRB.DL.Used.Avg'] = dfkpih5g['N.PRB.DL.Used.Avg'].str.replace(',','.')
        dfkpih5g['N.PRB.DL.Avail.Avg'] = dfkpih5g['N.PRB.DL.Avail.Avg'].str.replace(',','.')
        dfkpih5g['N.PRB.DL.Used.Avg'] = pd.to_numeric(dfkpih5g['N.PRB.DL.Used.Avg'],errors='coerce')
        dfkpih5g['N.PRB.DL.Avail.Avg'] = pd.to_numeric(dfkpih5g['N.PRB.DL.Avail.Avg'],errors='coerce')
        dfkpih5g['PRB.DL(%)'] = 100*dfkpih5g['N.PRB.DL.Used.Avg']/dfkpih5g['N.PRB.DL.Avail.Avg']
        '''
    #dfkpi4g['PRB.DL.Usage.RATE(%)']
    #dfkpi5g['5G_Hua_PRB_Use_DL(%)']
    if okexpret.get():
        if not 'dfretheidis' in globals():
            messagebox.showwarning(message="HeidiSQL RETs file .csv not loaded", title="Warning")
            global dfretheidis
            dfretheidis = pd.DataFrame(columns=['Fecha','neid','subname','deviceno','tilt'])
        dfrets,dicttext,dfret3500 = process_RET(fname)
        dfreportret = pd.DataFrame(columns=dfrets.keys())
        dfreportret["ERROR"] = ""
        dfreportret["CLUSTER"] = ""
    if okpbal.get():
        if not 'dfbal' in globals():
            dfbal = process_bal(fname)
    dfctall = pd.DataFrame(columns = clusters[0].dfct.keys())
    for i,c in enumerate(clusters):
        if not c.dfct.empty:
            create_folder(fname+'\\'+c.name)
            #===============MR==============
            nameMR = '4G_MR_'
            filename = fname+'\\'+c.name+"\\"+nameMR+c.name+".xlsx"
            if (okPRBs.get() | okHuella.get() | okgmr.get()) & (not os.path.isfile(filename)):
                dfkpi4gfilt,missingsiteskpis4g = filtkpi4gMR(dfkpi4g,c.dfct)
                if len(missingsiteskpis4g) > 0:
                    print('Missing KPIs cluster '+c.name+':\n'+missingsiteskpis4g)
                    savetxt(missingsiteskpis4g,fname+'\\'+c.name+'\\missingsiteskpis4gMR_'+c.name)
                #Copy file
                dfkpi4gfilt = dfkpi4gfilt.applymap(
                    lambda x: f"{x:.2f}".replace('.', ',') if isinstance(x, (int, float)) else x
                )
                dfkpi4gfilt.to_csv(fname+'\\'+c.name+'\\DataMR_'+c.name+'.csv',sep=';',encoding='utf-8',index=False)
                shutil.copy(nameMR+".xlsx", fname+'\\'+c.name+"\\"+nameMR+".xlsx")
                filename = fname+'\\'+c.name+"\\"+nameMR+c.name+".xlsx"
                os.rename(fname+'\\'+c.name+"\\"+nameMR+".xlsx", filename)
            #===============PRBs==============
            if okPRBs.get():
                process_PRB_cluster(c.dfct,c.name,c.sitess,dfkpi4g,dfkpih5g,fname)
            #===============FOOTPRINT==============
            if okHuella.get():
                process_Huella_cluster(c.dfct,c.name,dfkpi3g,dfkpi4g,dfkpih5g,fname,thorkeys,dfcellscoring,dfdb)
            #===============IOM TMP==============
            if okiomtmp.get():
                generate_tmp_iom(c.dfct,c.name,fname)
            #===============Query RETs==============
            if okqrets.get():
                query_rets_cluster(c.dfct,c.name,fname)
            #===============CT comb=================
            if okiomtmp.get() | okqrets.get() | okqbal.get() | okqbalcosector.get():
                dfctall = pd.concat([dfctall,c.dfct])
            #=================EXPORT RETs=============
            if okexpret.get():
                dfreportretsi = export_clusters_RETs(c,fname,dfrets,dicttext,dfret3500)
                if len(dfreportretsi)>0:
                    dfreportretsi['CLUSTER'] = c.name
                    dfreportret = pd.concat([dfreportret,dfreportretsi])
            #================Query BAL===========================
            if okqbal.get():
                query_bal_cluster(c.dfct,c.name,fname)
            if okqbalcosector.get():
                query_bal_cluster_cosector(c.dfct,c.name,fname)
            if okpbal.get():
                process_bal_cluster(c.dfct,c.name,fname,dfbal)

    if okiomtmp.get():
        generate_tmp_iom(dfctall,'ALL',fname)
    if okqrets.get():
        query_rets_cluster(dfctall,'ALL',fname)
    if okqbalcosector.get():
        query_bal_cluster_cosector(dfctall,'ALL',fname)
    if okqbal.get():
        query_bal_cluster(dfctall,'ALL',fname)
    if okexpret.get():
        dfreportret.to_excel(fname+'\\ALL\\reportrets.xlsx',index = False)
    messagebox.showinfo(message="Export clusters finished", title="Export Clusters")

def calcdfbal(dfbal,cdfct,dfkpi4g,dateend,datestart):
    sitesectors = np.unique(cdfct['Site-Sector'].tolist())
    for i,ss in enumerate(sitesectors):
        bands = np.unique(cdfct.loc[(cdfct['Site-Sector']==ss)&(cdfct.TECH=='4G'),'Band'])
        dfbal.loc[i,'SITE-SECTOR'] = ss
        dfbal.loc[i,'ENODEB'] = cdfct.loc[(cdfct['Site-Sector'] == ss) &(cdfct.TECH=='4G'),'NODE'].tolist()[0]
        for j,b in enumerate(bands):
            pos = i
            indexKPI = (dfkpi4g['Site-Sector'] == ss) & (dfkpi4g['Band'] == b) 
            if (datestart != '')&(dateend != ''):
                indexKPIdate = (dfkpi4g.Date >= datestart)&(dfkpi4g.Date <= dateend)
                #indexKPIdate = (dfkpi4g.Date >= datestart)&(dfkpi4g.Date <= str(datetime.datetime.strptime(dateend, '%Y-%m-%d').date() + datetime.timedelta(days=1)))
            else:
                indexKPIdate = indexKPI
            if any(indexKPI & indexKPIdate):
                #prbs = pd.to_numeric(dfkpi4g.loc[indexKPI & indexKPIdate,'PRB.DL.Usage.RATE(%)'],errors='coerce').tolist()
                prbs = [float(x) for x in dfkpi4g.loc[indexKPI & indexKPIdate,'PRB.DL.Usage.RATE(%)'].tolist() if (type(x) == str)&(x != '/0')&(x != '-')]
                prbs = prbs + [x for x in dfkpi4g.loc[indexKPI & indexKPIdate,'PRB.DL.Usage.RATE(%)'].tolist() if (type(x) != str)]
                dfbal.loc[pos,b] = np.mean(prbs)

    #dfbal = dfbal.merge(cdfct[['Site-Sector', 'NODE']], left_on='SITE-SECTOR', right_on='Site-Sector', how='left')
    #dfbal['ENODEB'] = dfbal['NODE']
    #dfbal.drop(columns='NODE', inplace=True)

    return dfbal

def calcdfprb4g(dfprb4g,cdfct,dfkpi4g,dateend,datestart):
    dfth4g = pd.DataFrame(columns=['Thmid','Thhload'])
    sitesectors = np.unique(cdfct['Site-Sector'].tolist())
    for i,ss in enumerate(sitesectors):
        bands = np.unique(cdfct.loc[(cdfct['Site-Sector']==ss)&(cdfct.TECH=='4G'),'Band'])
        for j,b in enumerate(bands):
            pos = len(dfprb4g)
            indexKPI = (dfkpi4g['Site-Sector'] == ss) & (dfkpi4g['Band'] == b) 
            if (datestart != '')&(dateend != ''):
                indexKPIdate = (dfkpi4g.Date >= datestart)&(dfkpi4g.Date <= dateend)
            else:
                indexKPIdate = indexKPI
            dfprb4g.loc[pos,'EUTRANCELL'] = cdfct.loc[(cdfct['Site-Sector'] == ss)&((cdfct['Band'] == b)),'CELLNAMEX'].tolist()[0]
            if (b == 'M')|(b == 'Y'):
                limitth = 4
                dfprb4g.loc[pos,'LB/HB'] = 'LB'
            else:
                limitth = 6
                dfprb4g.loc[pos,'LB/HB'] = 'HB'
            if any(indexKPI & indexKPIdate):
                throughput = dfkpi4g.loc[indexKPI & indexKPIdate,'4G_User_DL_Throughput(Mbps)(Mbps)'].tolist()
                pdcch = dfkpi4g.loc[indexKPI & indexKPIdate,'PDCCH.Usage.RATE(%)'].tolist()
                prbs = dfkpi4g.loc[indexKPI & indexKPIdate,'PRB.DL.Usage.RATE(%)'].tolist()
                for el in range(len(prbs)):
                    if (type(throughput[el]) == str) and (throughput[el].replace(',', '').isdigit()):
                        throughput[el] = float(throughput[el].replace(',', '.'))
                    elif (type(throughput[el]) == str):
                        throughput[el] = np.nan
                    if (type(pdcch[el]) == str) and (pdcch[el].replace(',', '').isdigit()):
                        pdcch[el] = float(pdcch[el].replace(',', '.'))
                    elif (type(pdcch[el]) == str):
                        pdcch[el] = np.nan
                    if (type(prbs[el]) == str) and (prbs[el].replace(',', '').isdigit()):
                        prbs[el] = float(prbs[el].replace(',', '.'))
                    elif (type(prbs[el]) == str):
                        prbs[el] = np.nan
                prbs = np.array(prbs)
                throughput = np.array(throughput)
                if len(np.unique(prbs))>1:
                    dfprb4g.loc[pos,'MAX PDCCH'] = np.nanmax(pdcch)
                    posth = np.nanargmax(prbs)
                    dfprb4g.loc[pos,'PRB HC'] = np.nanmax(prbs)
                    if (dfprb4g.loc[pos,'PRB HC'] >= 70):
                        dfprb4g.loc[pos,'TH_HC'] = np.nanmean(throughput[prbs>70])
                    else:
                        dfprb4g.loc[pos,'TH_HC'] = np.nanmean(throughput)
                    if (dfprb4g.loc[pos,'PRB HC'] > 70) & (dfprb4g.loc[pos,'TH_HC']<limitth):
                        dfprb4g.loc[pos,'OK/NOK'] = 'NOK'
                    else:
                        dfprb4g.loc[pos,'OK/NOK'] = 'OK'
                    posth = len(dfth4g)
                    dfth4g.loc[posth,'Thmid'] = np.nanmean(throughput)
                    dfth4g.loc[posth,'Thhload'] = np.nanmean(throughput[prbs>70])#It's going to be empty lots of times
                else:
                    dfprb4g.loc[pos,'PRB HC'] = -1
                    dfprb4g.loc[pos,'TH_HC'] = -1
                
    dfprb4g = dfprb4g.reset_index(drop=True)
    return dfprb4g,dfth4g

def calcdfprb5g(dfprb5g,cdfct,dfkpi5g,dateend,datestart):
    sitesectors = np.unique(cdfct['Site-Sector'].tolist())
    for i,ss in enumerate(sitesectors):
        bands = np.unique(cdfct.loc[(cdfct['Site-Sector']==ss)&(cdfct.TECH=='5G'),'Band'])
        for j,b in enumerate(bands):
            pos = len(dfprb5g)
            indexKPI = (dfkpi5g['Site-Sector'] == ss) & (dfkpi5g['Band'] == b) 
            if (datestart != '')&(dateend != ''):
                indexKPIdate = (dfkpi5g.Date >= datestart)&(dfkpi5g.Date <= dateend)
            else:
                indexKPIdate = indexKPI
            dfprb5g.loc[pos,'EUTRANCELL'] = cdfct.loc[(cdfct['Site-Sector'] == ss)&((cdfct['Band'] == b)),'CELLNAMEX'].tolist()[0]
            if (b == 'Q'):
                dfprb5g.loc[pos,'LB/HB'] = 'LB'
            else:
                dfprb5g.loc[pos,'LB/HB'] = 'HB'
            if any(indexKPI & indexKPIdate):
                nprbs = dfkpi5g.loc[indexKPI & indexKPIdate,'N.PRB.DL.Avail.Avg'].tolist()
                prbs = dfkpi5g.loc[indexKPI & indexKPIdate,'N.PRB.DL.Used.Avg'].tolist()
                if len(np.unique(prbs))>1:
                    for el in range(len(prbs)):
                        if (type(prbs[el]) == str) and (prbs[el].replace(',', '').isdigit()):
                            prbs[el] = float(prbs[el].replace(',', '.'))
                        elif (type(prbs[el]) == str):
                            prbs[el] = np.nan
                        if (type(nprbs[el]) == str) and (nprbs[el].replace(',', '').isdigit()):
                            nprbs[el] = float(nprbs[el].replace(',', '.'))
                        elif (type(nprbs[el]) == str):
                            nprbs[el] = np.nan
                    indexnprbs = np.nanargmax(prbs)
                    dfprb5g.loc[pos,'DEGRADACION PRB/TH'] = 100*np.nanmax(prbs)/nprbs[indexnprbs]
                else:
                    dfprb5g.loc[pos,'DEGRADACION PRB/TH'] = -1
    dfprb5g = dfprb5g.reset_index(drop=True)
    return dfprb5g

def dss5g24g(b4g):
    t5gto4g = {'Q':'Y','W':'T','X':'L'}
    if b4g in t5gto4g.keys():
        return t5gto4g[b4g]
    else:
        return ''

def calcdfprb4g5g(dfprb4g5g,cdfct,dfkpi4g,dfkpi5g,dateend,datestart):
    sitesectors = np.unique(cdfct['Site-Sector'].tolist())
    for i,ss in enumerate(sitesectors):
        bands5g = [x for x in np.unique(cdfct.loc[(cdfct['Site-Sector']==ss)&(cdfct.TECH=='5G'),'Band']) if x!='P']
        bands4g = np.unique(cdfct.loc[(cdfct['Site-Sector']==ss)&(cdfct.TECH=='4G'),'Band']) 
        for j,b in enumerate(bands5g):
            pos = len(dfprb4g5g)
            if b == 'Q':
                limth = 4
            else:
                limth = 6
            dfprb4g5g.loc[pos,'Site_Sector'] = ss
            indexKPI5g = (dfkpi5g['Site-Sector'] == ss) & (dfkpi5g['Band'] == b) 
            if (datestart != '')&(dateend != ''):
                indexKPI5gdate = (dfkpi5g.Date >= datestart)&(dfkpi5g.Date <= dateend)
            else:
                indexKPI5gdate = indexKPI5g
            if any(indexKPI5g & indexKPI5gdate):
                nprbs5g = dfkpi5g.loc[indexKPI5g & indexKPI5gdate,'N.PRB.DL.Avail.Avg'].tolist()
                prbs5g = dfkpi5g.loc[indexKPI5g & indexKPI5gdate,'N.PRB.DL.Used.Avg'].tolist()
                if len(np.unique(prbs5g))>1:
                    for el in range(len(prbs5g)):
                        if (type(prbs5g[el]) == str) and (prbs5g[el].replace(',', '').isdigit()):
                            prbs5g[el] = float(prbs5g[el].replace(',', '.'))
                        elif (type(prbs5g[el]) == str):
                            prbs5g[el] = np.nan
                        if (type(nprbs5g[el]) == str) and (nprbs5g[el].replace(',', '').isdigit()):
                            nprbs5g[el] = float(nprbs5g[el].replace(',', '.'))
                        elif (type(nprbs5g[el]) == str):
                            nprbs5g[el] = np.nan
                    if dss5g24g(b) in bands4g:
                        b4g = dss5g24g(b)
                        dfprb4g5g.loc[pos,'Bandas'] = b+b4g
                        indexKPI4g = (dfkpi4g['Site-Sector'] == ss) & (dfkpi4g['Band'] == b4g) 
                        if (datestart != '')&(dateend != ''):
                            indexKPI4gdate = (dfkpi4g.Date >= datestart)&(dfkpi4g.Date <= dateend)
                        else:
                            indexKPI4gdate = indexKPI4g
                        nprbs4g = dfkpi4g.loc[indexKPI4g & indexKPI4gdate,'L.ChMeas.PRB.DL.Avail'].tolist()
                        prbs4g = dfkpi4g.loc[indexKPI4g & indexKPI4gdate,'L.ChMeas.PRB.DL.Used.Avg'].tolist()
                        throughput4g = dfkpi4g.loc[indexKPI4g & indexKPI4gdate,'4G_User_DL_Throughput(Mbps)(Mbps)'].tolist()
                        if len(np.unique(prbs4g))>1:
                            for el in range(len(prbs4g)):
                                if (type(nprbs4g[el]) == str) and (nprbs4g[el].replace(',', '').isdigit()):
                                    nprbs4g[el] = float(nprbs4g[el].replace(',', '.'))
                                elif (type(nprbs4g[el]) == str):
                                    nprbs4g[el] = np.nan
                                if (type(prbs4g[el]) == str) and (prbs4g[el].replace(',', '').isdigit()):
                                    prbs4g[el] = float(prbs4g[el].replace(',', '.'))
                                elif (type(prbs4g[el]) == str):
                                    prbs4g[el] = np.nan
                                if (type(throughput4g[el]) == str) and (throughput4g[el].replace(',', '').isdigit()):
                                    throughput4g[el] = float(throughput4g[el].replace(',', '.'))
                                elif (type(throughput4g[el]) == str):
                                    throughput4g[el] = np.nan
                            prbs4g = np.array(prbs4g)
                            throughput4g = np.array(throughput4g)
                            dfprb4g5g.loc[pos,'PRB_SUM'] = 100*(np.nanmean(prbs5g)+np.nanmean(prbs4g))/(np.nanmean(nprbs5g)+np.nanmean(nprbs4g))
                            if dfprb4g5g.loc[pos,'PRB_SUM'] > 70:
                                dfprb4g5g.loc[pos,'4gth_hc'] = np.nanmean(throughput4g[prbs4g>70])
                            else:
                                dfprb4g5g.loc[pos,'4gth_hc'] = np.nanmean(throughput4g)
                            if (dfprb4g5g.loc[pos,'PRB_SUM'] > 70) & (dfprb4g5g.loc[pos,'4gth_hc']<limth):
                                dfprb4g5g.loc[pos,'OK/NOK'] = 'NOK'
                            else:
                                dfprb4g5g.loc[pos,'OK/NOK'] = 'OK'
                    else:
                        dfprb4g5g.loc[pos,'Bandas'] = b
                        dfprb4g5g.loc[pos,'PRB_SUM'] = 100*(np.nanmean(prbs5g))/(np.nanmean(nprbs5g))
                        if (dfprb4g5g.loc[pos,'PRB_SUM'] > 70):
                            dfprb4g5g.loc[pos,'OK/NOK'] = 'NOK'
                        else:
                            dfprb4g5g.loc[pos,'OK/NOK'] = 'OK'
                
    dfprb4g5g=dfprb4g5g.reset_index(drop=True)
    return dfprb4g5g

def process_PRB_cluster(cdfct,cname,seedsites,dfkpi4g,dfkpi5g,fname):
    datestr = datetime.datetime.now().strftime("%m_%d_%Y")
    cdfct.loc[:,'Sector']=[x[-2] for x in cdfct.CELLNAMEX.tolist()]
    cdfct.loc[:,'Site-Sector']=[x[0:3]+x[4:8]+'-'+x[-2] for x in cdfct.CELLNAMEX.tolist()]
    dfct4g = cdfct.loc[cdfct.loc[:,'TECH']=='4G',:]
    dfct5g = cdfct.loc[cdfct.loc[:,'TECH']=='5G',:]
    colprb4g = ['EUTRANCELL','LB/HB','MAX PDCCH','PRB HC','TH_HC','OK/NOK']
    dfprb4g = pd.DataFrame(columns=colprb4g)
    colprb4g5g = ['Site_Sector','Bandas','PRB_SUM','4gth_hc','OK/NOK']
    dfprb4g5g = pd.DataFrame(columns=colprb4g5g)
    colprb5g = ['EUTRANCELL','LB/HB','DEGRADACION PRB/TH']
    dfprb5g = pd.DataFrame(columns=colprb5g)
    colbal = ['ENODEB','SITE-SECTOR','M-Y','M-N','M-T','M-L','N-T','N-L','T-L','Y','M','N','T','L']
    dfbal = pd.DataFrame(columns=colbal)
    sitessecan = np.unique(dfct4g['Site-Sector'].tolist())
    datestart = entry_datestart.get()
    dateend = entry_dateend.get()
    if (len(datestart)!=10) | (len(dateend)!=10):
        print('Error Date '+cname)
        return 0
    else:
        datestart = pd.to_datetime(datestart)
        dateend = pd.to_datetime(dateend+ ' 23:00:00')
    #Go for 4G
    #============BALANCE==============
    dfbal = calcdfbal(dfbal,cdfct,dfkpi4g,dateend,datestart)
    #============PRB 4G================
    dfprb4g,dfth4g = calcdfprb4g(dfprb4g,cdfct,dfkpi4g,dateend,datestart)
    #============PRB 5G================
    dfprb5g = calcdfprb5g(dfprb5g,cdfct,dfkpi5g,dateend,datestart)
    #===========PRB 4G5G===============
    dfprb4g5g = calcdfprb4g5g(dfprb4g5g,cdfct,dfkpi4g,dfkpi5g,dateend,datestart)
    #========Create folder cluster========
    create_folder(fname+'\\'+cname)
    #======Copy tmp,change name,load file===
    folder_tmp_PRB = 'TMP_PRBs'
    folder_PRBs = 'PRBs_'+datestr
    create_folder(fname+'\\'+cname+'\\'+folder_PRBs)
    #Copy file
    shutil.copy(folder_tmp_PRB+"\\JUST_PRB_.docx", fname+'\\'+cname+'\\'+folder_PRBs+"\\JUST_PRB_.docx")
    filename = fname+'\\'+cname+'\\'+folder_PRBs+"\\JUST_PRB_"+cname+".docx"
    os.rename(fname+'\\'+cname+'\\'+folder_PRBs+"\\JUST_PRB_.docx", filename)
    #Copy file
    shutil.copy(folder_tmp_PRB+"\\PRB_THP_.xlsx", fname+'\\'+cname+'\\'+folder_PRBs+"\\PRB_THP_.xlsx")
    filename = fname+'\\'+cname+'\\'+folder_PRBs+"\\PRB_THP_"+cname+".xlsx"
    os.rename(fname+'\\'+cname+'\\'+folder_PRBs+"\\PRB_THP_.xlsx", filename)
    #Load file
    doctofill = openpyxl.load_workbook(filename)
    # Cluster cells sheet
    sheet = doctofill['PRB_THP']

    #=======FILL BALANCE===================
    keysbal = ['Y','M','N','T','L','ENODEB','SITE-SECTOR']
    posfillbal = np.array([18,19,20,21,22,23,24])+3 #RSTUV
    for i in range(len(dfbal)):
        for j,pos in enumerate(posfillbal):
            sheet.cell(row=i+5, column=pos, value=dfbal.loc[i,keysbal[j]])
    
    #=======FILL PRB 4G===================
    keys4g = ['EUTRANCELL','LB/HB','MAX PDCCH','PRB HC','TH_HC','OK/NOK']
    posfill4g = [1,2,3,4,5,6] 
    for i in range(len(dfprb4g)):
        for j,pos in enumerate(posfill4g):
            sheet.cell(row=i+6, column=pos, value=dfprb4g.loc[i,keys4g[j]])
    for i in range(len(dfth4g)):
        for j in range(len(dfth4g.keys())):
            sheet.cell(row=i+6, column=j+8, value=dfth4g.iloc[i,j])

    #=======FILL PRB 4G5G===================
    dfprb4g5g['PRB_SUM'] = dfprb4g5g['PRB_SUM']/100
    keys4g5g = ['Site_Sector','Bandas','PRB_SUM','4gth_hc','OK/NOK']
    posfill4g5g = np.array([8,9,10,11,12])+3 
    for i in range(len(dfprb4g5g)):
        for j,pos in enumerate(posfill4g5g):
            sheet.cell(row=i+5, column=pos, value=dfprb4g5g.loc[i,keys4g5g[j]])

    #=======FILL PRB 5G===================
    keys5g = ['EUTRANCELL','LB/HB','DEGRADACION PRB/TH']
    posfill5g = np.array([14,15,16])+3
    for i in range(len(dfprb5g)):
        for j,pos in enumerate(posfill5g):
            sheet.cell(row=i+4, column=pos, value=dfprb5g.loc[i,keys5g[j]])

    doctofill.save(filename)

    #=========================DEGRADED==========================
    filename = fname+'\\'+cname+'\\'+folder_PRBs+'\\Degraded\\'
    create_folder(filename)
    #============BALANCE==============
    dfbal.loc[:,'M-Y'] = np.abs(dfbal.M-dfbal.Y)
    dfbal.loc[:,'M-N'] = (dfbal.M-dfbal.N)
    dfbal.loc[:,'M-T'] = (dfbal.M-dfbal.loc[:,'T'])
    dfbal.loc[:,'M-L'] = (dfbal.M-dfbal.L)
    dfbal.loc[:,'N-T'] = np.abs(dfbal.N-dfbal.loc[:,'T'])
    dfbal.loc[:,'N-L'] = np.abs(dfbal.N-dfbal.L)
    dfbal.loc[:,'T-L'] = np.abs(dfbal.loc[:,'T']-dfbal.L)
    indexbaldeg = ((dfbal.loc[:,'M-Y'] > 20)|(dfbal.loc[:,'M-N'] > 20)|(dfbal.loc[:,'M-T'] > 20)|(dfbal.loc[:,'M-L'] > 20)|(dfbal.loc[:,'M-N'] < -20)|(dfbal.loc[:,'M-T'] < -40)|(dfbal.loc[:,'M-L'] < -40)|
                   (dfbal.loc[:,'N-T'] > 20)|(dfbal.loc[:,'N-L'] > 20)|(dfbal.loc[:,'T-L'] > 20))
    sitesectorbaldeg = dfbal.loc[indexbaldeg,'SITE-SECTOR'].tolist()
    tsitesectorbaldeg = 'NO BAL 4G: \n'+','.join(sitesectorbaldeg)
    if len(sitesectorbaldeg) > 0:
        filenamebal = filename+'BAL\\'
        create_folder(filenamebal)
    for ss in sitesectorbaldeg:
        dfaux = dfbal.loc[dfbal.loc[:,'SITE-SECTOR'] == ss,:]
        sectori = ss[-1]
        sitei = ss[0:7]
        namexi = [x for x in dfct.loc[(dfct.SITE == sitei),'CELLNAMEX'].tolist() if len(x) == 11][0][0:8]
        nodei = np.unique(dfct.loc[(dfct.SITE == sitei),'NODE'].tolist())
        for nodeii in nodei:
            plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenamebal+'4G_BAL_'+ss)
    #============PRB 4G================

    indexprb4gdeg = (dfprb4g.loc[:,'OK/NOK'] == 'NOK')
    indexpdcchdeg = (dfprb4g.loc[:,'MAX PDCCH'] > 60)
    cnmxprb4gdeg = dfprb4g.loc[indexprb4gdeg,'EUTRANCELL'].tolist()
    tcnmxprb4gdeg = 'PRB DEG 4G: \n'+','.join(cnmxprb4gdeg)
    cnmxpdcchdeg = dfprb4g.loc[indexpdcchdeg,'EUTRANCELL'].tolist()
    tcnmxpdcchdeg = 'PDCCH DEG 4G: \n'+','.join(cnmxpdcchdeg)
    if len(cnmxprb4gdeg) > 0:
        filenameprb4g = filename+'PRB4G\\'
        create_folder(filenameprb4g)
    if len(cnmxpdcchdeg) > 0:
        filenamepdcch4g = filename+'PDCCH4G\\'
        create_folder(filenamepdcch4g)

    for cni in cnmxprb4gdeg:
        sectori = cni[-2]
        bandi = cni[-3]
        sitei = cni[0:3]+cni[4:8]
        nodei = np.unique(dfct.loc[(dfct.SITE == sitei),'NODE'].tolist())
        for nodeii in nodei:
            plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[bandi],'(%)',[0,100],filenameprb4g+'4G_PRB_'+cni)
            plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenameprb4g+'4G_PRB_'+cni[0:8]+'-'+sectori)
    for cni in cnmxpdcchdeg:
        sectori = cni[-2]
        bandi = cni[-3]
        sitei = cni[0:3]+cni[4:8]
        nodei = np.unique(dfct.loc[(dfct.SITE == sitei),'NODE'].tolist())
        for nodeii in nodei:
            plotandsave(dfkpi4g,'PDCCH.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,60,[sectori],[bandi],'(%)',[0,100],filenamepdcch4g+'4G_PDCCH_'+cni)
            plotandsave(dfkpi4g,'PDCCH.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,60,[sectori],[],'(%)',[0,100],filenamepdcch4g+'4G_PDCCH_'+cni[0:3]+cni[3:8]+'-'+sectori)
            plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[bandi],'(%)',[0,100],filenamepdcch4g+'4G_PDCCH_PRB_'+cni)
            plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenamepdcch4g+'4G_PDCCH_PRB_'+cni[0:3]+cni[3:8]+'-'+sectori)

    #============PRB 5G================
    indexprb5gdeg = (dfprb5g.loc[:,'DEGRADACION PRB/TH']>70)
    cnmxprb5gdeg = dfprb5g.loc[indexprb5gdeg,'EUTRANCELL'].tolist()
    tcnmxprb5gdeg = 'PRB DEG 5G: \n'+','.join(cnmxprb5gdeg)
    if len(cnmxprb5gdeg) > 0:
        filenameprb5g = filename+'PRB5G\\'
        create_folder(filenameprb5g)
    for cni in cnmxprb5gdeg:
        sectori = cni[-2]
        bandi = cni[-3]
        sitei = cni[0:3]+cni[4:8]
        nodei = np.unique(dfct.loc[(dfct.SITE == sitei),'NODE'].tolist())
        for nodeii in nodei:
            plotandsave(dfkpi5g,'PRB.DL(%)','gNodeB Name','Cell Name',nodeii,70,[sectori],[bandi],'(%)',[0,100],filenameprb5g+'5G_PRB_'+cni)
            plotandsave(dfkpi5g,'PRB.DL(%)','gNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenameprb5g+'5G_PRB_'+cni[3:8]+'-'+sectori)
    #===========PRB 4G5G===============
    indexprb4g5gdeg = (dfprb4g5g.loc[:,'OK/NOK'] == 'NOK')
    sitesectorprb4g5gdeg = dfprb4g5g.loc[indexprb4g5gdeg,'Site_Sector'].tolist()
    sitesectorbandsprb4g5gdeg = [sitesectorprb4g5gdeg[el]+'-'+x for el,x in enumerate(dfprb4g5g.loc[indexprb4g5gdeg,'Bandas'].tolist())]
    tsitesectorbandsprb4g5gdeg = 'PRB DEG 4G5G: \n'+','.join(sitesectorbandsprb4g5gdeg)
    if len(sitesectorbandsprb4g5gdeg) > 0:
        filenameprb4g5g = filename+'PRB4G5G\\'
        create_folder(filenameprb4g5g)
    for ssb in sitesectorbandsprb4g5gdeg:
        sitei = ssb[0:7]
        sectori = ssb[8]
        bandsi = ssb[10::]
        sitei = cni[0:3]+cni[4:8]
        nodei = np.unique(dfct.loc[(dfct.SITE == sitei),'NODE'].tolist())
        if 'Y' in bandsi:
            band4gi = 'Y'
            #namex4gi = dfct.loc[(dfct.SITE == sitei)&(dfct.Band==band4gi),'CELLNAMEX'].tolist()[0][0:8]
        elif 'T' in bandsi:
            band4gi = 'T'
            #namex4gi = dfct.loc[(dfct.SITE == sitei)&(dfct.Band==band4gi),'CELLNAMEX'].tolist()[0][0:8]
        elif 'L' in bandsi:
            band4gi = 'L'
        else:
            band4gi = ''
        if 'Q' in bandsi:
            band5gi = 'Q'
            #namex5gi = dfct.loc[(dfct.SITE == sitei)&(dfct.Band==band5gi),'CELLNAMEX'].tolist()[0][0:8]
        elif 'W' in bandsi:
            band5gi = 'W'
        elif 'X' in bandsi:
            band5gi = 'X'
            #namex5gi = dfct.loc[(dfct.SITE == sitei)&(dfct.Band==band5gi),'CELLNAMEX'].tolist()[0][0:8]
        for nodeii in nodei:
            plotandsave(dfkpi5g,'PRB.DL(%)','gNodeB Name','Cell Name',nodeii,70,[sectori],[bandi],'(%)',[0,100],filenameprb4g5g+'4G5G_PRB5G_'+sitei+band5gi+sectori+'A')
            plotandsave(dfkpi5g,'PRB.DL(%)','gNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenameprb4g5g+'4G5G_PRB5G_'+sitei)
            if len(band4gi) > 0:
                plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[bandi],'(%)',[0,100],filenameprb4g5g+'4G5G_PRB4G_'+sitei+band4gi+sectori+'A')
                plotandsave(dfkpi4g,'PRB.DL.Usage.RATE(%)','eNodeB Name','Cell Name',nodeii,70,[sectori],[],'(%)',[0,100],filenameprb4g5g+'4G5G_PRB4G_'+sitei)
    
    savetxt(tsitesectorbaldeg+'\n\n'+tcnmxprb4gdeg+'\n\n'+tcnmxpdcchdeg+'\n\n'+tcnmxprb5gdeg+'\n\n'+tsitesectorbandsprb4g5gdeg,fname+'\\'+cname+'\\'+folder_PRBs+'\\Degraded\\Degraded_'+cname)

    #--------------OPT PROPOSAL------------------

    # obtain seed nodes
    seednodes = ','.join([dfct.loc[(dfct.TECH == '4G') & (dfct.SITE == ss.name),'NODE'].tolist()[0] for ss in seedsites if len(dfct.loc[(dfct.TECH == '4G') & (dfct.SITE == ss.name),'NODE'].tolist())>1])

    #======Copy tmp,change name,load file===
    #Copy file
    shutil.copy(folder_tmp_PRB+"\\OPT_PRO_.xlsx", fname+'\\'+cname+'\\'+folder_PRBs+"\\OPT_PRO_.xlsx")
    filename = fname+'\\'+cname+'\\'+folder_PRBs+"\\OPT_PRO_"+cname+".xlsx"
    os.rename(fname+'\\'+cname+'\\'+folder_PRBs+"\\OPT_PRO_.xlsx", filename)
    #Load file
    doctofill = openpyxl.load_workbook(filename)
    # Cluster cells sheet
    sheet = doctofill['OPT_PROPOSAL']

    rowZN = 3
    rowZR = 3
    sitesectpro = []
    for cell in cnmxprb4gdeg:
        sitei = cell[0:3]+cell[4:8]
        sectori = cell[9]
        if sitei+sectori not in sitesectpro:
            nodei = cdfct.loc[cdfct.CELLNAMEX == cell,'NODE'].tolist()[0]
            cellsi = ','.join(cdfct.loc[(cdfct.SITE == sitei)&(cdfct.Sector == sectori)&(cdfct.TECH == '4G'),'CELLNAMEX'].tolist())
            #Check ZN
            if cell[0:3] == nodei[0:3]:
                sheet.cell(row=rowZN, column=1, value=seednodes)
                sheet.cell(row=rowZN, column=2, value=cname)
                sheet.cell(row=rowZN, column=3, value=sitei)
                sheet.cell(row=rowZN, column=4, value=cellsi)
                sitesectpro.append(sitei+sectori)
                rowZN+=1
            else:
                provi = cdfct.loc[cdfct.CELLNAMEX == cell,'PROVINCIA'].tolist()[0]
                sheet.cell(row=rowZR, column=8, value=cname)
                sheet.cell(row=rowZR, column=10, value=nodei)
                sheet.cell(row=rowZR, column=11, value=provi)
                sheet.cell(row=rowZR, column=12, value=sitei)
                sheet.cell(row=rowZR, column=13, value=cellsi)
                sheet.cell(row=rowZR, column=14, value=sectori)
                sitesectpro.append(sitei+sectori)
                rowZR+=1
    doctofill.save(filename)

def filtkpi3g(dfkpi3g,cdfct,datestart,dateend,ik3gfix=-1):
    #filter
    missingsiteskpis3g = ''
    sites = np.unique(cdfct.SITE.tolist())
    indok = [False for x in range(len(dfkpi3g))]
    for s in sites:
        indit = dfkpi3g.Site == s
        if any(indit):
            indok = indok| indit
        else:
            missingsiteskpis3g += s+'\n'
    if (datestart != '')&(dateend != ''):
        indexKPIdate = (dfkpi3g.Date >= datestart)&(dfkpi3g.Date <= dateend)
    else:
        indexKPIdate = indok
    indok = indok & indexKPIdate
    dfkpi3gfilt = dfkpi3g.loc[indok,:].copy()
    dfkpi3gfilt.reset_index(inplace=True)
    #fixformat
    #dfkpi3gfilt['3G_MeanDistance(#)'] = dfkpi3gfilt['3G_MeanDistance(#)'].str.replace(',','.')
    #dfkpi3gfilt['3G_MeanDistance(#)'] = pd.to_numeric(dfkpi3gfilt['3G_MeanDistance(#)'],errors='coerce')
    #if len(dfkpi3gfilt)>0:
    #    for k3f in dfkpi3g.keys()[ik3gfix::]:
    #        if type(dfkpi3gfilt[k3f][0])==str:
    #            dfkpi3gfilt[k3f] = dfkpi3gfilt[k3f].str.replace(',','.')
    #            dfkpi3gfilt[k3f] = pd.to_numeric(dfkpi3gfilt[k3f],errors='coerce')
    return dfkpi3gfilt,missingsiteskpis3g

def filtkpi4g(dfkpi4g,cdfct,datestart,dateend,ik4gfix=-1):
    #filter
    missingsiteskpis4g = ''
    sites = np.unique(cdfct.SITE.tolist())
    indok = [False for x in range(len(dfkpi4g))]
    for s in sites:
        indit = dfkpi4g.Site == s
        if any(indit):
            indok = indok| indit
        else:
            missingsiteskpis4g += s+'\n'
    if (datestart != '')&(dateend != ''):
        indexKPIdate = (dfkpi4g.Date >= datestart)&(dfkpi4g.Date <= dateend)
    else:
        indexKPIdate = indok
    indok = indok & indexKPIdate
    dfkpi4gfilt = dfkpi4g.loc[indok,:].copy()
    dfkpi4gfilt.reset_index(inplace=True)
    #fixformat
    #dfkpi4gfilt['PRB.DL.Usage.RATE(%)'] = dfkpi4gfilt['PRB.DL.Usage.RATE(%)'].str.replace(',','.')
    #dfkpi4gfilt['PRB.DL.Usage.RATE(%)'] = pd.to_numeric(dfkpi4gfilt['PRB.DL.Usage.RATE(%)'],errors='coerce')
    #dfkpi4gfilt['Average coverage distance(m)'] = dfkpi4gfilt['Average coverage distance(m)'].str.replace(',','.')
    #dfkpi4gfilt['Average coverage distance(m)'] = pd.to_numeric(dfkpi4gfilt['Average coverage distance(m)'],errors='coerce')
    #if len(dfkpi4gfilt)>0:
    #    for k4f in dfkpi4g.keys()[ik4gfix::]:
    #        if type(dfkpi4gfilt[k4f][0])==str:
    #            dfkpi4gfilt[k4f] = dfkpi4gfilt[k4f].str.replace(',','.')
    #            dfkpi4gfilt[k4f] = pd.to_numeric(dfkpi4gfilt[k4f],errors='coerce')
    return dfkpi4gfilt,missingsiteskpis4g  

def filtkpi5g(dfkpi5g,cdfct,datestart,dateend,ik5gfix=-1):
    #filter
    missingsiteskpis5g = ''
    sites = np.unique(cdfct.SITE.tolist())
    indok = [False for x in range(len(dfkpi5g))]
    for s in sites:
        indit = dfkpi5g.Site == s
        if any(indit):
            indok = indok| indit
        else:
            missingsiteskpis5g += s+'\n'
    if (datestart != '')&(dateend != ''):
        indexKPIdate = (dfkpi5g.Date >= datestart)&(dfkpi5g.Date <= dateend)
    else:
        indexKPIdate = indok
    indok = indok & indexKPIdate
    dfkpi5gfilt = dfkpi5g.loc[indok,:].copy()
    dfkpi5gfilt.reset_index(inplace=True)
    #fixformat
    #dfkpi5gfilt['5G_Hua_PRB_Use_DL(%)'] = dfkpi5gfilt['5G_Hua_PRB_Use_DL(%)'].str.replace(',','.')
    #dfkpi5gfilt['5G_Hua_PRB_Use_DL(%)'] = pd.to_numeric(dfkpi5gfilt['5G_Hua_PRB_Use_DL(%)'],errors='coerce')

    #if len(dfkpi5gfilt)>0:
    #    for k5f in dfkpi5g.keys()[ik5gfix::]:
    #        if type(dfkpi5gfilt[k5f][0])==str:
    #            dfkpi5gfilt[k5f] = dfkpi5gfilt[k5f].str.replace(',','.')
    #            dfkpi5gfilt[k5f] = pd.to_numeric(dfkpi5gfilt[k5f],errors='coerce')
    return dfkpi5gfilt,missingsiteskpis5g  

def filtcellscore(thorkeys,dfcellscoring,cdfct,idxffor=9):
    missingsitescellscore = ''
    #filter
    sites = np.unique(cdfct.SITE.tolist())
    indok = [False for x in range(len(dfcellscoring))]
    for s in sites:
        indit = dfcellscoring.Site == s
        if any(indit):
            indok = indok| indit
        else:
            missingsitescellscore += s+'\n'
    dfcellscorefilt = dfcellscoring.loc[indok,:].copy()
    dfcellscorefilt.reset_index(inplace=True)
    #fixformat
    for k in thorkeys[idxffor::]:
        if type(dfcellscorefilt[k][0])==str:
            dfcellscorefilt[k] = dfcellscorefilt[k].str.replace(',','.')
            dfcellscorefilt[k] = pd.to_numeric(dfcellscorefilt[k],errors='coerce')  
    return dfcellscorefilt,missingsitescellscore

def checkmissingsites(cdfct,mstxt,tech):
    ikeep = []
    mk3gl = mstxt.split('\n')
    for ism,sm in enumerate(mk3gl):
        if any((cdfct.SITE == sm)&(cdfct.TECH == tech)&(cdfct.STATUS == 1)):
            ikeep.append(ism)
    mk3gl = np.array(mk3gl)[ikeep]
    return '\n'.join(mk3gl)
        
def process_Huella_cluster(cdfct,cname,dfkpi3g,dfkpi4g,dfkpih5g,fname,thorkeys,dfcellscoring,dfdb):
    #========Create folder cluster========
    datestr = datetime.datetime.now().strftime("%d_%m_%Y")
    create_folder(fname+'\\'+cname)
    tmpfname = 'TMP_Huella'
    fpfname = 'Huella_'+datestr
    create_folder(fname+'\\'+cname+'\\'+fpfname)
    #======Copy tmp,change name,load file=====
    #Copy file excel
    namefill = 'HUELLA_'
    shutil.copy(tmpfname+'\\'+namefill+".xlsx", fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".xlsx")
    filename = fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+cname+".xlsx"
    os.rename(fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".xlsx", filename)
    #Copy file word
    namefill = 'HUELLA_'
    shutil.copy(tmpfname+'\\'+namefill+".docx", fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".docx")
    filename = fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+cname+".docx"
    os.rename(fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".docx", filename)
    #Copy file plantilla
    namefill = 'PLANTILLA_HUELLA_'
    shutil.copy(tmpfname+'\\'+namefill+".xlsx", fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".xlsx")
    filename = fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+cname+".xlsx"
    os.rename(fname+'\\'+cname+"\\"+fpfname+'\\'+namefill+".xlsx", filename)
    datestart = entry_datestart.get()
    dateend = entry_dateend.get()
    if (len(datestart)!=10) | (len(dateend)!=10):
        print('Error Date '+cname)
        return 0
    else:
        datestart = pd.to_datetime(datestart)
        dateend = pd.to_datetime(dateend+ ' 23:00:00')
    # =======================3G sheet==================================
    dfkpi3gfilt,missingsiteskpis3g = filtkpi3g(dfkpi3g,cdfct,datestart,dateend,8)
    if len(missingsiteskpis3g) > 0:
        missingsiteskpis3g = checkmissingsites(cdfct,missingsiteskpis3g,'3G')
        if len(missingsiteskpis3g) > 0:
            savetxt(missingsiteskpis3g,fname+'\\'+cname+'\\'+fpfname+'\\missingsiteskpis3g_'+cname)
    keys3g = ['Date','Cell Name','VS.TP.UE.0','VS.TP.UE.1','VS.TP.UE.2','VS.TP.UE.3','VS.TP.UE.4','VS.TP.UE.5','VS.TP.UE.6.9',
              'VS.TP.UE.10.15','VS.TP.UE.16.25','VS.TP.UE.26.35','VS.TP.UE.36.55','VS.TP.UE.More55','3G_MeanDistance(#)']
    keysadd3g = ['VS.TP.UE.0(%)','VS.TP.UE.1(%)','VS.TP.UE.2(%)','VS.TP.UE.3(%)','VS.TP.UE.4(%)','VS.TP.UE.5(%)','VS.TP.UE.6.9(%)',
              'VS.TP.UE.10.15(%)','VS.TP.UE.16.25(%)','VS.TP.UE.26.35(%)','VS.TP.UE.36.55(%)','VS.TP.UE.More55(%)']
    """
    for ii in range(len(keys3g[2:-1])):
        dfkpi3gfilt.loc[:,keysadd3g[ii]] = 100*dfkpi3gfilt[keys3g[2+ii]]/dfkpi3gfilt[keys3g[2:-1]].sum(axis = 1)
    for x in np.unique(dfkpi3gfilt['Cell Name'].tolist()):
        bandi = x[-3]
        sectori = x[-2]
        plotandsave(dfkpi3gfilt,keysadd3g,'Cell Name',x,'',sectori,bandi,'%',[0,100],fname+'\\'+cname+"\\"+fpfname+'\\'+'VSTPUE_'+x+'_'+datestr)
    """
    # =======================4G sheet==================================
    dfkpi4gfilt,missingsiteskpis4g = filtkpi4g(dfkpi4g,cdfct,datestart,dateend,6)
    if len(missingsiteskpis4g) > 0:
        missingsiteskpis4g = checkmissingsites(cdfct,missingsiteskpis4g,'4G')
        if len(missingsiteskpis4g) > 0:
            savetxt(missingsiteskpis4g,fname+'\\'+cname+'\\'+fpfname+'\\missingsiteskpis4g_'+cname)
    keys4g = ['Date','Cell Name','PRB.DL.Usage.RATE(%)','L.TA.UE.Index0','L.TA.UE.Index1','L.TA.UE.Index2','L.TA.UE.Index3','L.TA.UE.Index4',
              'L.TA.UE.Index5','L.TA.UE.Index6','Average coverage distance(m)']
    keysadd4g = ['L.TA.UE.Index0(%)','L.TA.UE.Index1(%)','L.TA.UE.Index2(%)','L.TA.UE.Index3(%)','L.TA.UE.Index4(%)',
              'L.TA.UE.Index5(%)','L.TA.UE.Index6(%)','Average coverage distance(m)']
    """
    for ii in range(len(keys4g[2:-1])):
        dfkpi4gfilt.loc[:,keysadd4g[ii]] = 100*dfkpi4gfilt[keys4g[2+ii]]/dfkpi4gfilt[keys4g[2:-1]].sum(axis = 1)
    for x in np.unique(dfkpi4gfilt['Cell Name'].tolist()):
        bandi = x[-3]
        sectori = x[-2]
        plotandsave(dfkpi4gfilt,keysadd4g,'Cell Name',x,'',sectori,bandi,'%',[0,100],fname+'\\'+cname+"\\"+fpfname+'\\'+'TA4G_'+x+'_'+datestr)
    """
    # =======================5G sheet==================================
    dfkpi5gfilt,missingsiteskpis5g = filtkpi5g(dfkpih5g,cdfct,datestart,dateend,6)
    if len(missingsiteskpis5g) > 0:
        missingsiteskpis5g = checkmissingsites(cdfct,missingsiteskpis5g,'5G')
        if len(missingsiteskpis5g) > 0:
            savetxt(missingsiteskpis5g,fname+'\\'+cname+'\\'+fpfname+'\\missingsiteskpis5g_'+cname)
    keys5g = ['Date','Cell Name','5G_Hua_PRB_Use_DL(%)','N.RA.TA.UE.Index0','N.RA.TA.UE.Index1','N.RA.TA.UE.Index2','N.RA.TA.UE.Index3',
              'N.RA.TA.UE.Index4','N.RA.TA.UE.Index5','N.RA.TA.UE.Index6','N.RA.TA.UE.Index7','N.RA.TA.UE.Index8','N.RA.TA.UE.Index9',
              'N.RA.TA.UE.Index10','N.RA.TA.UE.Index11','N.RA.TA.UE.Index12']
    keysadd5g = ['N.RA.TA.UE.Index0(%)','N.RA.TA.UE.Index1(%)','N.RA.TA.UE.Index2(%)','N.RA.TA.UE.Index3(%)','N.RA.TA.UE.Index4(%)','N.RA.TA.UE.Index5(%)',
              'N.RA.TA.UE.Index6(%)','N.RA.TA.UE.Index7(%)','N.RA.TA.UE.Index8(%)','N.RA.TA.UE.Index9(%)','N.RA.TA.UE.Index10(%)','N.RA.TA.UE.Index11(%)',
              'N.RA.TA.UE.Index12(%)']
    """
    for ii in range(len(keys5g[3::])):
        dfkpi5gfilt.loc[:,keysadd5g[ii]] = 100*dfkpi5gfilt[keys5g[3+ii]]/dfkpi5gfilt[keys5g[3::]].sum(axis = 1)
    dfkpi5gfilt.loc[:,keysadd5g] = 100*dfkpi5gfilt[keys5g[3::]]/dfkpi5gfilt[keys5g[3::]].sum()
    for x in np.unique(dfkpi5gfilt['Cell Name'].tolist()):
        bandi = x[-3]
        sectori = x[-2]
        plotandsave(dfkpi5gfilt,keysadd5g,'Cell Name',x,'',sectori,bandi,'%',[0,100],fname+'\\'+cname+"\\"+fpfname+'\\'+'TA5G_'+x+'_'+datestr)   
    """
    # =======================Thor sheet==================================
    dfcellscorefilt,missingsitescellscore = filtcellscore(thorkeys,dfcellscoring,cdfct)
    if len(missingsitescellscore) > 0:
        savetxt(missingsitescellscore,fname+'\\'+cname+'\\missingsitescellscore_'+cname)
    
    #==========================CELLNAMEX DB=================================
    icellnx = (cdfct.TECH == '3G')|(cdfct.TECH == '4G')|(cdfct.TECH == '5G')
    dfdbexp = pd.DataFrame(columns=['CLUSTER','CELLNAMEX','DIST','COMMENTS','DATE'])
    dfdbexp2 = pd.DataFrame(columns=['CLUSTER','CELLNAMEX','DIST','COMMENTS','DATE'])
    cellnamexc = [x for x in cdfct.loc[icellnx,'CELLNAMEX'].tolist() if x[-1] in ('A','B')]
    dfdb.sort_values('DATE', inplace=True)
    for cnxc in cellnamexc:
        idb = dfdb.CELLNAMEX == cnxc
        if any(idb):
            dfdbexp = pd.concat([dfdbexp,dfdb.loc[idb,:]])
            idate = dfdb.DATE == dfdb.loc[idb,'DATE'].max()
            dfdbexp2 = pd.concat([dfdbexp2,dfdb.loc[idb&idate,:]])
        else:
            posadd = len(dfdbexp)
            dfdbexp.loc[posadd,'CELLNAMEX'] = cnxc
            dfdbexp.loc[posadd,'CLUSTER'] = cname
            posadd = len(dfdbexp2)
            dfdbexp2.loc[posadd,'CELLNAMEX'] = cnxc
            dfdbexp2.loc[posadd,'CLUSTER'] = cname

    #===============Save===============
    with pd.ExcelWriter(fname+'\\'+cname+'\\'+fpfname+'\\Data_footprint_'+cname+'.xlsx') as writer:  
        dfkpi3gfilt.to_excel(writer, sheet_name='3G',index=False)
        dfkpi4gfilt.to_excel(writer, sheet_name='4G',index=False)
        dfkpi5gfilt.to_excel(writer, sheet_name='5G',index=False)
        dfcellscorefilt[thorkeys].to_excel(writer, sheet_name='Thor',index=False)
        cdfct.to_excel(writer, sheet_name='CT',index=False)
        dfdbexp.to_excel(writer, sheet_name='DB',index=False)
        dfdbexp2.to_excel(writer, sheet_name='DB_filtered',index=False)



#==========================Process CELL SCORING===========================
        
def process_cell_scoring_only_footprint():
    dirct = browseFiles_celltable()
    dfct = pd.read_csv(dirct,sep = ';')
    load_csv_cellscoring()
    filenamethorkeys = 'ThorKeys.csv'
    thorkeys = load_csv_thorkeys(filenamethorkeys)
    celllist = pd.read_excel('cells.xlsx')
    celllist = celllist['Cell Name'].tolist()
    indexct = [x in celllist for x in dfct['CELLNAMEX']]
    indexctoperator = dfct['OPERATOR NAME'].str.upper() == 'ORANGE'
    indexstatus = dfct['STATUS'] == 1
    cdfct = dfct.loc[indexct&indexctoperator&indexstatus,:].reset_index()
    dfcellscoring['Site'] = [x[0:3]+x[4:8] for x in dfcellscoring['cellnamex']]
    dfcellscorefilt,missingsitescellscore = filtcellscore(thorkeys,dfcellscoring,cdfct)
    if len(missingsitescellscore) > 0:
        savetxt(missingsitescellscore,'missingsitescellscore')
    dfcellscorefilt.to_excel('cellscoringfilter')

#==============================RETs==========================================

def tech2freq2(tech):
    tech2freqdict = {'Y':'700','M':'800','E':'900','F':'900','N':'1800','B':'2100','T':'2100','L':'2600','Q':'700','W':'2100','X':'2600','P':'3500'}
    if tech in tech2freqdict.keys():
        freq = tech2freqdict[tech]
    else:
        freq = ''
    return freq

def ZRnode(node):
    ok = False
    if (len(node)<8):
        ok = True
    return ok

def comfreq(fa,fb):
    ok = False
    for f in fa:
        if any(f == fb):
            ok = True
    return ok

def getindexnewline(text,n,i):
    if n  >= 0:
        for ni in range(n):
            i = text[i::].find('\n')+i+1
    else:
        text = text[0:i]
        text = text[::-1]
        n = -n+1
        cnt = 0
        for ni in range(n):
            pos = text.find('\n')+1
            cnt += pos 
            text = text[pos::]
        i = i-cnt+1     
    return i

def getindexnextspace(text,n,i):
    for ni in range(n):
        i = text[i::].find(' ')+i
    return i

def istech(node,tech):
    return any(dfct.TECH[dfct.NODE == node] == tech)

def checktf(f,df):
    ok = False
    fsearch = []
    coln = df.columns.values
    i = 0
    while 'Freq_'+str(i) in coln:
        fsearch.append('Freq_'+str(i))
        i+=1 
    flist = np.array([])
    for n in fsearch:
        flist = np.append(flist,df.loc[:,n])
    flist = np.unique(flist)
    flist = flist.astype('int16') 
    flist = flist[flist != 0]
    #delete 3500
    f = f[f!=3500]
    if len(f) == len(flist):
        f = np.sort(f)
        flist = np.sort(flist)
        if all(f == flist):
            ok = True
    return ok,flist

def browsetextfile():
    path = filedialog.askopenfilenames(initialdir = os.getcwd(), 
        title = "Select MML Command text file",
        filetypes = (("txt file","*.txt*"), ("all files", "*.*"))) 
    #return path[0]
    return list(path)

def browsetextfilesave():
    path = filedialog.asksaveasfilename(initialdir = os.getcwd(), 
        title = "Select MML Command text file",
        filetypes = (("txt file","*.txt*"), ("all files", "*.*"))) 
    #return path[0]
    return path

def join_text():
    if 'text' in globals():
        path = browsetextfilesave()
        if (len(path)>4)and(path[-4::] == '.txt'):
            path = path[0:-4]
        savetxt(text,path)

def readtextfile():
    global text
    text = ''
    paths = browsetextfile()
    for path in paths:
        if (path == '') | (path[-4::] != '.txt'):
            messagebox.showwarning(message="Please select a valid text file", title="Warning")
            return
        try:
            with open(path,encoding="utf-8") as f:
                text += f.read()+'\n'
        except IOError:
            messagebox.showwarning('Error: File does not appear to exist.')
            return 0
    #global filenameread
    #filenameread = os.path.split(path)[1][0:-4]
    return text

def process_txt(command,column_names):
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    global text
    textan = text
    dfexp = pd.DataFrame(columns=column_names)
    indexini = textan.find(command)
    while indexini!=-1:
        #check Corresponding results not found
        iend = indexini+textan[indexini::].find('END')
        if not 'Corresponding results not found' in textan[indexini:iend]:
            #generate dfroe in the iteration
            dfexpi = pd.DataFrame(columns=column_names)
            #NE line
            indexnode = getindexnewline(textan,1,indexini)
            #Extract node
            node = textan[indexnode+5:getindexnewline(textan,1,indexnode)-1]
            #1. search line, look for number of results and subtrac lines
            indexnor = indexini
            nlinesadd = 0
            while textan[indexnor:indexnor+10] != '(Number of':
                indexnor = getindexnewline(textan,1,indexnor)
                nlinesadd+=1
                if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
                    break
            if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
                #delete and new search
                textan = textan[indexnor+29::]
                indexini = textan.find(command)
            else:
                #2. point to the first
                indexnorend = getindexnewline(textan,1,indexnor)-1
                nres = int(textan[indexnorend-3:indexnorend-1])
                indexani = getindexnewline(textan,nlinesadd-nres,indexini)
                #3. extract columns and index of columns
                indexcols = getindexnewline(textan,nlinesadd-nres-2,indexini)
                indexcolsend = getindexnewline(textan,1,indexcols)-1
                textcolsi = textan[indexcols:indexcolsend]
                indexcolsi = np.ones([len(column_names),2],dtype='int16')*-1
                for i in range(len(column_names)):
                    colnamei = column_names[i]
                    indexcolsi[i,0]=int(textcolsi.find(colnamei))
                    if indexcolsi[i,0]!=-1:
                        indexcolsi[i,1] = indexcolsi[i,0]+len(colnamei)
                #4. Extract info
                for i in range(nres):
                    for j in range(len(column_names)):
                        if indexcolsi[j,0] != -1:
                            #search next space
                            indexendj = getindexnextspace(textan,1,indexani+indexcolsi[j,0])
                            dfexpi.loc[i,column_names[j]] = textan[indexani+indexcolsi[j,0]:indexendj]
                    indexani = getindexnewline(textan,1,indexani)
                #5. Save dataframe
                dfexpi.loc[:,'NE'] = node
                dfexp = pd.concat([dfexp,dfexpi],ignore_index = True)
                #new search
                textan = textan[indexnor+29::]
                indexini = textan.find(command)
        else:
            textan = textan[iend::]
            indexini = textan.find(command)
    return dfexp

def process_txt_verthorz(command,column_names):
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    global text
    textan = text
    dfexp = pd.DataFrame(columns=column_names)
    indexini = textan.find(command)
    while indexini!=-1:
        #generate dfroe in the iteration
        dfexpi = pd.DataFrame(columns=column_names)
        #NE line
        indexnode = getindexnewline(textan,1,indexini)
        #Extract node
        node = textan[indexnode+5:getindexnewline(textan,1,indexnode)-1]
        #check UCELL
        #1. search line, look for number of results and subtrac lines
        indexnor = indexini
        nlinesadd = 0
        while textan[indexnor:indexnor+10] != '(Number of':
            indexnor = getindexnewline(textan,1,indexnor)
            nlinesadd+=1
            if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
                break
        if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
            #delete and new search
            textan = textan[indexnor+29::]
            indexini = textan.find(command)
        else:
            if textan[indexnor+21:indexnor+23] != '1)':
                #2. point to the first
                indexnorend = getindexnewline(textan,1,indexnor)-1
                nres = int(textan[indexnorend-3:indexnorend-1])
                indexani = getindexnewline(textan,nlinesadd-nres,indexini)
                #3. extract columns and index of columns
                indexcols = getindexnewline(textan,nlinesadd-nres-2,indexini)
                indexcolsend = getindexnewline(textan,1,indexcols)-1
                textcolsi = textan[indexcols:indexcolsend]
                indexcolsi = np.ones([len(column_names),2],dtype='int16')*-1
                for i in range(len(column_names)):
                    colnamei = column_names[i]
                    indexcolsi[i,0]=int(textcolsi.find(colnamei))
                    if indexcolsi[i,0]!=-1:
                        indexcolsi[i,1] = indexcolsi[i,0]+len(colnamei)
                #4. Extract info
                for i in range(nres):
                    for j in range(len(column_names)):
                        if indexcolsi[j,0] != -1:
                            #search next space
                            indexendj = getindexnextspace(textan,1,indexani+indexcolsi[j,0])
                            dfexpi.loc[i,column_names[j]] = textan[indexani+indexcolsi[j,0]:indexendj]
                    indexani = getindexnewline(textan,1,indexani)
            else:
                for j in range(len(column_names)):
                    posname = textan.find(column_names[j])
                    posdata = posname + textan[posname::].find('=')+3
                    addposdata = textan[posdata::].find('\n')
                    dfexpi.loc[0,column_names[j]] = textan[posdata:posdata+addposdata]
            #5. Save dataframe
            dfexpi.loc[:,'NE'] = node
            dfexp = pd.concat([dfexp,dfexpi],ignore_index = True)
            #new search
            textan = textan[indexnor+29::]
            indexini = textan.find(command)
    return dfexp

def process_txt_verthorz_multiple(command,column_names):
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    global text
    textan = text
    dfexp = pd.DataFrame(columns=column_names)
    indexini = textan.find(command)
    while indexini!=-1:
        #generate df in the iteration
        dfexpi = pd.DataFrame(columns=column_names)
        #NE line
        indexnode = getindexnewline(textan,1,indexini)
        #Extract node
        node = textan[indexnode+5:getindexnewline(textan,1,indexnode)-1]
        if node == 'CS23AL':
            node
        #check UCELL
        #1. search line, look for number of results and subtrac lines
        indexnor = indexini
        indexmult = getindexnewline(textan,2,indexnor)
        nlinesadd = 0
        while textan[indexnor:indexnor+10] != '(Number of':
            indexnor = getindexnewline(textan,1,indexnor)
            nlinesadd+=1
            if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
                break
        if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
            #delete and new search
            textan = textan[indexnor+29::]
            indexini = textan.find(command)
        else:
            if textan[indexnor+21:indexnor+23] != '1)':
                #if multiple
                indexend = getindexnewline(textan,3,indexnor)
                if textan[indexend:indexend+10] != '---    END':
                    #report to report
                    while textan[indexend:indexend+10] == 'To be cont':
                        posnof = textan[indexnor::].find('(Number of results = ')+21+indexnor
                        posendnof = textan[posnof::].find(')')+posnof
                        nres = int(textan[posnof:posendnof])
                        #3. extract columns and index of columns
                        indexcols = getindexnewline(textan,-nres-2,indexnor)
                        indexani = getindexnewline(textan,2,indexcols)
                        indexcolsend = getindexnewline(textan,1,indexcols)-1
                        textcolsi = textan[indexcols:indexcolsend]
                        indexcolsi = np.ones([len(column_names),2],dtype='int16')*-1
                        for i in range(len(column_names)):
                            colnamei = column_names[i]
                            indexcolsi[i,0]=int(textcolsi.find(colnamei))
                            if indexcolsi[i,0]!=-1:
                                indexcolsi[i,1] = indexcolsi[i,0]+len(colnamei)
                        #4. Extract info
                        for i in range(nres):
                            posdf = len(dfexpi)
                            for j in range(len(column_names)):
                                if indexcolsi[j,0] != -1:
                                    #search next space
                                    indexendj = getindexnextspace(textan,1,indexani+indexcolsi[j,0])
                                    dfexpi.loc[posdf,column_names[j]] = textan[indexani+indexcolsi[j,0]:indexendj]
                            indexani = getindexnewline(textan,1,indexani)
                        indexnor = textan[indexend::].find('(Number of')+indexend
                        indexend = getindexnewline(textan,3,indexnor)
                    posnof = textan[indexnor::].find('(Number of results = ')+21+indexnor
                    posendnof = textan[posnof::].find(')')+posnof
                    nres = int(textan[posnof:posendnof])
                    #3. extract columns and index of columns
                    indexcols = getindexnewline(textan,-nres-2,indexnor)
                    indexani = getindexnewline(textan,2,indexcols)
                    indexcolsend = getindexnewline(textan,1,indexcols)-1
                    textcolsi = textan[indexcols:indexcolsend]
                    indexcolsi = np.ones([len(column_names),2],dtype='int16')*-1
                    for i in range(len(column_names)):
                        colnamei = column_names[i]
                        indexcolsi[i,0]=int(textcolsi.find(colnamei))
                        if indexcolsi[i,0]!=-1:
                            indexcolsi[i,1] = indexcolsi[i,0]+len(colnamei)
                    #4. Extract info
                    for i in range(nres):
                        posdf = len(dfexpi)
                        for j in range(len(column_names)):
                            if indexcolsi[j,0] != -1:
                                #search next space
                                indexendj = getindexnextspace(textan,1,indexani+indexcolsi[j,0])
                                dfexpi.loc[posdf,column_names[j]] = textan[indexani+indexcolsi[j,0]:indexendj]
                        indexani = getindexnewline(textan,1,indexani)                 
                else:
                    #2. point to the first
                    indexnorend = getindexnewline(textan,1,indexnor)-1
                    nres = int(textan[indexnorend-3:indexnorend-1])
                    indexani = getindexnewline(textan,nlinesadd-nres,indexini)
                    #3. extract columns and index of columns
                    indexcols = getindexnewline(textan,nlinesadd-nres-2,indexini)
                    indexcolsend = getindexnewline(textan,1,indexcols)-1
                    textcolsi = textan[indexcols:indexcolsend]
                    indexcolsi = np.ones([len(column_names),2],dtype='int16')*-1
                    for i in range(len(column_names)):
                        colnamei = column_names[i]
                        indexcolsi[i,0]=int(textcolsi.find(colnamei))
                        if indexcolsi[i,0]!=-1:
                            indexcolsi[i,1] = indexcolsi[i,0]+len(colnamei)
                    #4. Extract info
                    for i in range(nres):
                        for j in range(len(column_names)):
                            if indexcolsi[j,0] != -1:
                                #search next space
                                indexendj = getindexnextspace(textan,1,indexani+indexcolsi[j,0])
                                dfexpi.loc[i,column_names[j]] = textan[indexani+indexcolsi[j,0]:indexendj]
                        indexani = getindexnewline(textan,1,indexani)
            else:
                for j in range(len(column_names)):
                    posname = indexini+textan[indexini::].find(column_names[j])
                    posdata = posname + textan[posname::].find('=')+3
                    addposdata = textan[posdata::].find('\n')
                    dfexpi.loc[0,column_names[j]] = textan[posdata:posdata+addposdata]
            #5. Save dataframe
            dfexpi.loc[:,'NE'] = node
            dfexp = pd.concat([dfexp,dfexpi],ignore_index = True)
            #new search
            textan = textan[indexnor+29::]
            indexini = textan.find(command)
    return dfexp

def get_NE_txt(command):
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    global text
    textan = text
    indexini = textan.find(command)
    dictext = {}
    while indexini!=-1:
        #NE line
        indexnode = getindexnewline(textan,1,indexini)
        #Extract node
        NE = textan[indexnode+5:getindexnewline(textan,1,indexnode)-1]
        #1. search line, look for number of results and subtrac lines
        indexnor = indexini
        nlinesadd = 0
        while textan[indexnor:indexnor+10] != '(Number of':
            indexnor = getindexnewline(textan,1,indexnor)
            nlinesadd+=1
            if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
                break
        if (textan[indexnor:indexnor+29] == 'Report : Ne is not connected.')|(textan[indexnor:indexnor+29] == 'No matching result is found\n\n'):
            #delete and new search
            textan = textan[indexnor+29::]
            indexini = textan.find(command)
        else:
            indexend = textan[indexini::].find('END')+indexini+4
            dictext[NE] = textan[indexini:indexend]
            #new search
            textan = textan[indexend::]
            indexini = textan.find(command)
    return dictext

def complete_dfret(dfret):
    nodes = dfret.NE.tolist()
    for el,n in enumerate(nodes):
        indexnode = (dfct.loc[:,'NODE'] == n) & ((dfct.loc[:,'OPERATOR NAME'] == 'ORANGE')|(dfct.loc[:,'OPERATOR NAME'] == 'Orange'))
        dfret.loc[el,'Site'] = dfct.loc[indexnode,'SITE'].tolist()[0]
        if not ZRnode(n):
            dfret.loc[el,'Sector'] = dfret.loc[el,'Device Name'][-1]
        else:
            possect = dfret.loc[el,'Device Name'].find('_')-1
            dfret.loc[el,'Sector'] = dfret.loc[el,'Device Name'][possect]
    
    dfret['Tilt'] = pd.to_numeric(dfret['Actual Tilt(0.1degree)'],errors='coerce')
    freqZN = np.array([700,800,900,1800,2100,2600]).astype('str')
    dictfreqZR = {'S':'700','Q':'700','J':'800','G':'900','U':'900','K':'1800','M':'2100','W':'2100','L':'2600','X':'2600'}
    bandsZR = dictfreqZR.keys()
    for el,n in enumerate(dfret.NE.tolist()):
        devnameel = dfret.loc[el,'Device Name']
        if ZRnode(n):
            dfret.loc[el,'ZONE'] = 'ZR'
            devnameel = devnameel[(devnameel.find('_')+1)::]
            if devnameel.find('OSP'):
                devnameel = devnameel[0:devnameel.find('OSP')]+devnameel[devnameel.find('OSP')+2::]
            fok = True
            for b in bandsZR:
                if b in devnameel:
                    if fok:
                        dfret.loc[el,'Freq'] = [[dictfreqZR[b]]]
                        fok = False
                    else:
                        dfret.loc[el,'Freq'].append(dictfreqZR[b])
        else:
            dfret.loc[el,'ZONE'] = 'ZN'
            fok = True
            for f in freqZN:
                if fok:
                    if (f == '800')&(f in devnameel)&('1800'not in devnameel):
                        dfret.loc[el,'Freq'] = [[f]]
                        fok = False                
                    elif (f != '800')&(f in devnameel):
                        dfret.loc[el,'Freq'] = [[f]]
                        fok = False
                else:
                    if (f == '800')&(f in devnameel)&('1800'not in devnameel):
                        dfret.loc[el,'Freq'].append(f)              
                    elif (f != '800')&(f in devnameel):
                        dfret.loc[el,'Freq'].append(f)
    notev = dfret.loc[:,'Freq'].isnull()
    posev = np.array([el for el,x in enumerate(notev) if not x])
    indexndiplexok = [el for el,x in enumerate(dfret.loc[~notev,'Freq'].tolist()) if len(x)>1]
    indexndiplexnok = [el for el,x in enumerate(dfret.loc[~notev,'Freq'].tolist()) if len(x)<=1]
    dfret.loc[posev[indexndiplexok],'DIPLEX'] = 'D'
    dfret.loc[posev[indexndiplexnok],'DIPLEX'] = 'N'
    return dfret
def complete_dfret3500(dfret3500):
    dfret3500 = dfret3500.dropna(subset=['NR DU Cell TRP ID'])
    dfret3500['NR DU Cell TRP ID'] = dfret3500['NR DU Cell TRP ID'].astype(int)
    dfct.loc[:,'NE'] = dfct.NODE
    dfct.loc[:,'NR DU Cell TRP ID'] = dfct['CELLID']
    dfret3500 = pd.merge(dfret3500,dfct[['NE','SITE','NR DU Cell TRP ID','CELLNAMEX']],how = 'left',on=['NE','NR DU Cell TRP ID'])
    return dfret3500

def complete_dfretheidis(dfret):
    #fill subnames
    indexsubname = dfretheidis.loc[:,'subname'].isnull()
    nodessubname = dfretheidis.loc[indexsubname,'neid'].tolist()
    devicenosubname = dfretheidis.loc[indexsubname,'deviceno'].tolist()
    addsubnames = []
    for el,n in enumerate(nodessubname):
        indexdfret = (dfret.NE == n) & (dfret['Device No.']==str(devicenosubname[el]))
        addsubnames.append(dfret.loc[indexdfret,'Device Name'].tolist()[0])
    dfretheidis.loc[indexsubname,'subname'] = addsubnames
    nodes = dfretheidis.neid.tolist()
    freqZN = np.array([700,800,900,1800,2100,2600]).astype('str')
    dictfreqZR = {'S':'700',
                  'Q':'700',
                  'J':'800',
                  'G':'900',
                  'U':'900',
                  'K':'1800',
                  'M':'2100',
                  'W':'2100',
                  'L':'2600',
                  'X':'2600'}
    bandsZR = dictfreqZR.keys()
    for el,n in enumerate(nodes):
        dfretheidis.loc[el,'Site'] = dfct.loc[dfct.loc[:,'NODE'] == n,'SITE'].tolist()[0]
        devnameel = dfretheidis.loc[el,'subname']
        if ZRnode(n):
            dfretheidis.loc[el,'ZONE'] = 'ZR'
            possect = dfretheidis.loc[el,'subname'].find('_')-1
            dfretheidis.loc[el,'Sector'] = dfretheidis.loc[el,'subname'][possect]
            devnameel = [x[x.find('_')+1::] for x in devnameel]
            fok = True
            for b in bandsZR:
                if b in devnameel:
                    if fok:
                        dfretheidis.loc[el,'Freq'] = [[dictfreqZR[b]]]
                        fok = False
                    else:
                        dfretheidis.loc[el,'Freq'].append(dictfreqZR[b])
        else:
            dfretheidis.loc[el,'Sector'] = dfretheidis.loc[el,'subname'][-1]
            dfretheidis.loc[el,'ZONE'] = 'ZN'
            fok = True
            for f in freqZN:
                if fok:
                    if (f == '800')&(f in devnameel)&('1800'not in devnameel):
                        dfretheidis.loc[el,'Freq'] = [[f]]
                        fok = False                
                    elif (f != '800')&(f in devnameel):
                        dfretheidis.loc[el,'Freq'] = [[f]]
                        fok = False
                else:
                    if (f == '800')&(f in devnameel)&('1800'not in devnameel):
                        dfretheidis.loc[el,'Freq'].append(f)              
                    elif (f != '800')&(f in devnameel):
                        dfretheidis.loc[el,'Freq'].append(f)

def fixformatfreq(strlistfreq):
    freqlist = []
    if type(strlistfreq) == str:
        posini = strlistfreq.find("'")
        while posini != -1:
            posend = strlistfreq[posini+1::].find("'")+posini+1
            freqlist.append(strlistfreq[posini+1:posend])
            strlistfreq = strlistfreq[posend+1::]
            posini = strlistfreq.find("'")
    else:
        freqlist.append(str(strlistfreq))
    return freqlist
def fixformattilt(tl):
    tlok = ''
    if tl.find('.')!=-1:
        tlok = tl[0:tl.find('.')]
    else:
        tlok = 'NULL'
    return tlok
def formatdfret():
    freqlist = dfret.Freq.tolist()
    for el,fl in enumerate(freqlist):
        freqlist[el] = fixformatfreq(fl)
    dfret.Freq = freqlist
    tiltlist = dfret.loc[:,'Actual Tilt(0.1degree)'].astype(str).tolist()
    for el,tl in enumerate(tiltlist):
        tiltlist[el] = fixformattilt(tl)
    dfret.loc[:,'Actual Tilt(0.1degree)'] = tiltlist

def process_RET(fname):
    if (not 'text' in globals()):
        messagebox.showwarning(message="Load a .txt ret file", title="Warning")
        return
    if not 'dfct' in globals():
        messagebox.showwarning(message="Load a celltable .csv", title="Warning")
        return
    if (not 'dfret' in globals()):
        global dfret
        command = 'MML Command-----DSP RETSUBUNIT:;'
        column_names = ['Device No.','Device Name','Subunit No.','Subunit Name','Online Status',
            'Actual Tilt(0.1degree)','Actual Sector ID','RET Configuration Data File Name',
            'Configuration Data File Load Time']
        dfret = process_txt_verthorz_multiple(command,column_names)
        dfret = complete_dfret(dfret)
        dictext = get_NE_txt(command)
    else:
        command = 'MML Command-----DSP RETSUBUNIT:;'
        dictext = get_NE_txt(command)
        formatdfret()

    command = 'MML Command-----LST NRDUCELLTRPBEAM:;'
    column_names = ['NR DU Cell TRP ID','Coverage Scenario','Tilt(degree)','Azimuth(degree)','Max SSB Power Offset(dB)'] 
    dfret3500 = process_txt_verthorz_multiple(command,column_names)
    dfret3500 = complete_dfret3500(dfret3500)
    #complete_dfretheidis(dfret)
    create_folder(fname+'\\ALL')
    if (not os.path.isfile(fname+'\\ALL\\rets_all.xlsx')):
        dfret.to_excel(fname+'\\ALL\\rets_all.xlsx',index=False)
        dfret3500.to_excel(fname+'\\ALL\\rets_all_3500.xlsx',index=False)
    return dfret,dictext,dfret3500


def export_image(text,path,ext = '.png'):
    font = font_manager.FontProperties(family='Lucida Console', weight='bold')
    file = font_manager.findfont(font)
    font = ImageFont.truetype(file, 11)
    nrows = text.count('\n') 
    if len(text) == 0:
        spaces = []
        textim = 'RET not found'
        img = Image.new('RGB',(200,100),color = (255, 255, 255))
    else:
        spaces = np.array([m.start() for m in re.finditer('\n', text)])
        if len(spaces) == 0:
            img = Image.new('RGB',(200,100),color = (255, 255, 255))
            textim = text
        else:
            if (text.find('No matching result is found') == -1) & (text.find('Report : Ne is not connected.') == -1):
                sizerows = np.diff(spaces)
                rowmax = np.argmax(sizerows)
                if text[text.find('Actual Tilt')+24] == '=':
                    ncols = text[spaces[rowmax]+1:spaces[rowmax+1]].find('Actual Tilt')+60
                else:
                    ncols = text[spaces[rowmax]+1:spaces[rowmax+1]].find('Actual Tilt')+22
                #dlt unnecesary text 
                textim = text[0:spaces[0]]
                for i in range(len(sizerows)-1):
                    if sizerows[i+1] > ncols:
                        textim = textim+text[spaces[i+1]:spaces[i+1]+ncols+1]
                    else:
                        textim = textim+text[spaces[i+1]:spaces[i+2]]
            else:
                sizerows = np.diff(spaces)
                ncols = np.max(sizerows)
                textim = text
            if text.find('Report : Ne is not connected.') != -1:
                    img = Image.new('RGB', (int(np.round(8*(ncols+1))),int(np.round((nrows+2)*15))),color = (255, 255, 255))
            else:
                img = Image.new('RGB', (int(np.round(8*(ncols))),int(np.round((nrows)*15))),color = (255, 255, 255))
    d = ImageDraw.Draw(img)
    d.text((20,20), textim, fill=(0,0,0),font = font)
    img.save(path+ext, dpi=(1000, 1000))
    return img

def insert_image_sheet(sheet,img):
    image_parts = img.split()
    r = image_parts[0]
    g = image_parts[1]
    b = image_parts[2]
    img = Image.merge("RGB", (r, g, b))
    fo = BytesIO()
    img.save(fo, format='bmp')
    sheet.insert_bitmap_data(fo.getvalue(),2,0)
    return sheet

def modsubunitname(df,path,name):
    df.loc[:,'MOD_SUNAME'] = 'MOD RETSUBUNIT:DEVICENO='+df.loc[:,'Device No.']+',SUBUNITNO='+df.loc[:,'Subunit No.']+',SUBNAME="'+df.loc[:,'Device Name']+'";'
    textlist = df.loc[:,'MOD_SUNAME'].tolist()
    if (len(textlist) > 0) and (type(textlist[0])==type('asdf')):
        text = '\n'.join(textlist)
        with open(path+'\\'+'modsubunitname_'+name+'.txt', 'a') as f:
            f.write(text)
    else:
        df

def export_csv_rets():
    fname = 'Clusters PRBs Huella'
    if not 'dfct' in globals():
        global dfct
        dfct = load_csv_celltable()
    dfrets,dicttext,dfret3500 = process_RET(fname)
    dfrets.to_csv(fname+'\\ALL\\RETs.csv',index=False,sep=';', encoding='utf-8')
    messagebox.showinfo(message="Export RETs finished", title="Export RETs")

def format_file_tmp_rets(df,output_file,sheet_name):

    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    # Combinación de celdas en la columna 'SITE' cada 10 filas
    start_row = 2  # Empieza en la fila 2, porque la fila 1 es la cabecera
    column_letter = get_column_letter(df.columns.get_loc('SITE') + 1)  # Obtener la letra de la columna de 'SITE'

    for i in range(0, len(df), 11):
        start = start_row + i
        end = start + 10
        ws.merge_cells(f'{column_letter}{start}:{column_letter}{end}')

    # Guardar el archivo Excel con las combinaciones de celdas
    wb.save(output_file)

def export_clusters_RETs(c,fncontainer,dfrets,dicttext,dfrets3500):
    indexgeneral = ((c.dfct.loc[:,'OPERATOR NAME']== 'ORANGE')|(c.dfct.loc[:,'OPERATOR NAME']== 'Orange'))&(c.dfct.loc[:,'STATUS']== 1)
    cdfct = c.dfct.loc[indexgeneral,:].copy()
    dfreportrets = pd.DataFrame(columns=dfrets.keys())
    dfreportrets["ERROR"] = ""
    sites = np.unique(cdfct.SITE.tolist())
    sitess = [x.name for x in c.sitess]
    dictkeys = dicttext.keys()
    #Create a container folder
    # Check whether the specified path exists or not
    fnret = 'RETs'
    create_folder(fncontainer)
    #Create a cluster folder
    create_folder(fncontainer+'\\'+c.name+'\\'+fnret)
    #Create excel 
    book=xlwt.Workbook(encoding="utf-8",style_compression=0)
    #create dataframe for template
    colnameretfp = ['SITE','TECH','CONCAT','DIPLEX']
    dfretfp = pd.DataFrame(columns=colnameretfp)
    dfretseedfp = pd.DataFrame(columns=colnameretfp)
    techs = ['Y','Q','M','F','N','T','W','B','L','X','P']
    for s in np.unique(sites):
        create_folder(fncontainer+'\\'+c.name+'\\'+fnret+'\\'+s)  
        #index dataframe
        indexi  = dfrets.Site == s
        #if empty, add to report
        if not any(indexi):
            pos = len(dfreportrets)
            dfreportrets.loc[pos,'Site'] = s
            dfreportrets.loc[pos,'ERROR'] = 'Not found'
            if s in sitess:
                dfreportrets.loc[pos,'Semilla'] = 'Y'
            else:
                dfreportrets.loc[pos,'Semilla'] = 'N'
        #Export dataframe
        dfrets.loc[indexi,:].to_excel(fncontainer+'\\'+c.name+'\\'+fnret+'\\'+s+'\\'+'rets.xlsx',index=False)
        #Export image
        indexcti = (cdfct.SITE == s)
        #indexcti = (cdfct.SITE == s)&(cdfct.STATUS == 1)&((cdfct.loc[:,'OPERATOR NAME'] == 'ORANGE')|(cdfct.loc[:,'OPERATOR NAME'] == 'Orange'))
        nodesi = np.unique(cdfct.loc[indexcti,'NODE'].tolist())
        inodesi = 0
        while (inodesi < len(nodesi))and(not nodesi[inodesi] in dictkeys):
            inodesi+=1
        if inodesi < len(nodesi):
            nodei = nodesi[inodesi]
            img = export_image(dicttext[nodei],fncontainer+'\\'+c.name+'\\'+fnret+'\\'+s+'\\'+s,'.png') 
        else:
            img = export_image('',fncontainer+'\\'+c.name+'\\'+fnret+'\\'+s+'\\'+s,'.png')
        if s in sitess:
            #Create sheet 
            sheet = book.add_sheet(s, cell_overwrite_ok=True)
            #fill excel sheet
            sheet.write(0,0,s)
            #instert image
            sheet = insert_image_sheet(sheet,img)
            okseed = True
        else:
            okseed = False
        cellnamexi = [x for x in cdfct.loc[indexcti,'CELLNAMEX'].tolist() if (len(x) == 11) & (x[3]!='B')]
        bandsi = [x[8] for x in cellnamexi]
        sectorsi = [x[9] for x in cellnamexi]
        cellnamei = [x[0:3]+x[4:10] for x in cellnamexi]
        pos = len(dfretfp)
        dfretfp.loc[pos,'SITE'] = s
        if okseed:
            posseed = len(dfretseedfp)
            dfretseedfp.loc[posseed,'SITE'] = s
        for tech in techs:
            dfretfp.loc[pos,'TECH'] = tech
            if okseed:
                dfretseedfp.loc[posseed,'TECH'] = tech
            if tech in bandsi:
                freqret = tech2freq2(tech)
                dfretfp.loc[pos,'CONCAT'] = s+tech
                if okseed:
                    dfretseedfp.loc[posseed,'CONCAT'] = s+tech
                sectori = 1
                while s+tech+str(sectori) in cellnamei:
                    if tech == 'P':
                        dfretfp.loc[pos,'DIPLEX'] = 'N'
                        index3500 = dfrets3500.CELLNAMEX == s[0:3]+'X'+s[3:7]+tech+str(sectori)+'A'
                        if any(index3500):
                            dfretfp.loc[pos,'S'+str(sectori)] = float(dfrets3500.loc[index3500,'Tilt(degree)'].tolist()[0])
                    else:
                        indexretsitesect = (dfrets.Site == s)&((dfrets.Sector == str(sectori))|(dfrets.Sector == sectori))
                        indexretsfreq = np.array([False for x in range(len(indexretsitesect))])
                        posretsfreq = [el for el,x in enumerate(dfrets.Freq.tolist()) if (type(x) == list) and (freqret in x)]
                        indexretsfreq[posretsfreq] = True
                        indexrets = indexretsitesect & indexretsfreq
                        if any(indexrets):
                            if len(dfrets.loc[indexrets,'DIPLEX']) > 1:
                                cret = np.array(dfrets.loc[indexrets,'Actual Tilt(0.1degree)'].tolist())
                                cdiplex = np.array(dfrets.loc[indexrets,'DIPLEX'].tolist())
                                if any(cdiplex == 'D'):
                                    dfretfp.loc[pos,'DIPLEX'] = 'D'
                                else:
                                    dfretfp.loc[pos,'DIPLEX'] = 'N'
                                if any(cret == 'NULL'):
                                    dfretfp.loc[pos,'S'+str(sectori)] = cret[0]
                                    dfretfp.loc[pos,'Comments'] = 'Error NULL'
                                    dfrets.loc[indexrets,'ERROR'] = 'Error NULL'
                                    if okseed:
                                        dfrets.loc[indexrets,'Semilla'] = 'Y'
                                    else:
                                        dfrets.loc[indexrets,'Semilla'] = 'N'
                                    dfreportrets = pd.concat([dfreportrets,dfrets.loc[indexrets,:]])
                                else:
                                    dfretfp.loc[pos,'S'+str(sectori)] = int(cret[0])
                                #error diplex
                                if (dfretfp.loc[pos,'DIPLEX'] == 'D') & any(np.diff(dfrets.loc[indexrets,'Tilt'].tolist())):
                                    dfretfp.loc[pos,'Comments'] = 'Error Diplex'
                                    dfrets.loc[indexrets,'ERROR'] = 'Error Diplex'
                                    if okseed:
                                        dfrets.loc[indexrets,'Semilla'] = 'Y'
                                    else:
                                        dfrets.loc[indexrets,'Semilla'] = 'N'
                                    dfreportrets = pd.concat([dfreportrets,dfrets.loc[indexrets,:]])
                                if okseed:
                                    dfretseedfp.loc[posseed,'DIPLEX'] = dfretfp.loc[pos,'DIPLEX']
                                    nodeheidis = dfrets.loc[indexrets,'NE'].tolist()[0]
                                    devicenoheidis = int(dfrets.loc[indexrets,'Device No.'].tolist()[0])
                                    dftiltheidis = dfretheidis.loc[(dfretheidis.neid == nodeheidis)&(dfretheidis.deviceno == devicenoheidis),['Fecha','tilt']].copy().reset_index()
                                    if not dftiltheidis.empty:
                                        dfretseedfp.loc[posseed,'S'+str(sectori)] = int(dftiltheidis.loc[dftiltheidis.Fecha.idxmin(),'tilt'])
                            else:
                                dfretfp.loc[pos,'DIPLEX'] = dfrets.loc[indexrets,'DIPLEX'].tolist()[0]
                                cret = dfrets.loc[indexrets,'Actual Tilt(0.1degree)'].tolist()[0]
                                if cret == 'NULL':
                                    dfretfp.loc[pos,'S'+str(sectori)] = cret
                                    dfretfp.loc[pos,'Comments'] = 'Error NULL'
                                    dfrets.loc[indexrets,'ERROR'] = 'Error NULL'
                                    if okseed:
                                        dfrets.loc[indexrets,'Semilla'] = 'Y'
                                    else:
                                        dfrets.loc[indexrets,'Semilla'] = 'N'
                                    dfreportrets = pd.concat([dfreportrets,dfrets.loc[indexrets,:]])
                                else:
                                    dfretfp.loc[pos,'S'+str(sectori)] = int(cret)
                                if okseed:
                                    dfretseedfp.loc[posseed,'DIPLEX'] =dfretfp.loc[pos,'DIPLEX']
                                    nodeheidis = dfrets.loc[indexrets,'NE'].tolist()[0]
                                    devicenoheidis = int(dfrets.loc[indexrets,'Device No.'].tolist()[0])
                                    dftiltheidis = dfretheidis.loc[(dfretheidis.neid == nodeheidis)&(dfretheidis.deviceno == devicenoheidis),['Fecha','tilt']].copy().reset_index()
                                    if not dftiltheidis.empty:
                                        dfretseedfp.loc[posseed,'S'+str(sectori)] = int(dftiltheidis.loc[dftiltheidis.Fecha.idxmin(),'tilt'])

                            dfretfp.loc[pos,'CONCAT'] = s+tech
                    sectori+=1 
            pos = len(dfretfp)
            if okseed:
                posseed = len(dfretseedfp)

  
    #Save export template
    output_file = fncontainer+'\\'+c.name+'\\'+fnret+'\\tilt_tmp.xlsx'
    sheet_name_dfretfp = 'ALL SITES'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:                
        dfretfp.to_excel(writer,sheet_name = sheet_name_dfretfp,index=False)
        dfretseedfp.to_excel(writer,sheet_name = 'SEED DESIGN',index=False) 
        dfreportrets.to_excel(writer,sheet_name = 'REPORT',index=False) 
    format_file_tmp_rets(dfretfp,output_file,sheet_name_dfretfp)

    #Save book
    book.save(fncontainer+'\\'+c.name+'\\'+fnret+'\\'+'RETs.xls')
    dfreportrets.CLUSTER = c.name
    return dfreportrets


def generate_tmp_iom(cdfct,cname,fname):
    indexgeneral = ((cdfct.loc[:,'OPERATOR NAME']== 'ORANGE')|(cdfct.loc[:,'OPERATOR NAME']== 'Orange'))&(cdfct.loc[:,'STATUS']== 1)&(cdfct.loc[:,'VENDOR'] =='HUAWEI')
    cdfct2g = cdfct.loc[indexgeneral&(cdfct.loc[:,'TECH']=='2G'),:].copy().reset_index()
    cdfct3g = cdfct.loc[indexgeneral&(cdfct.loc[:,'TECH']=='3G'),:].copy().reset_index()
    cdfct4g = cdfct.loc[indexgeneral&(cdfct.loc[:,'TECH']=='4G'),:].copy().reset_index()
    cdfct5g = cdfct.loc[indexgeneral&(cdfct.loc[:,'TECH']=='5G'),:].copy().reset_index()
    #========Create folder cluster========
    create_folder(fname+'\\'+cname)
    sheetname = 'template'
    tmpfoldername = 'TMP_IOM'
    create_folder(fname+'\\'+cname+'\\'+tmpfoldername)
    #==============2G==============
    if not cdfct2g.empty:
        #Load iom export file
        df2giom = pd.read_csv("TMP_IOM\\HUAWEI_GSM_CUSTOMIZED_CELL_exportCellObjectForSelect.csv",sep=";")
        #======Copy tmp,change name,load file===
        #Copy file
        shutil.copy(tmpfoldername+"\\HUAWEI_GSM_CUSTOMIZED_CELL_.xlsx", fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_GSM_CUSTOMIZED_CELL_.xlsx")
        filename = fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_GSM_CUSTOMIZED_CELL_"+cname+".xlsx"
        os.rename(fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_GSM_CUSTOMIZED_CELL_.xlsx", filename)
        #Load file
        doctofill = openpyxl.load_workbook(filename)
        # Cluster cells sheet
        sheet = doctofill[sheetname]
        keysfill = ['CONTROLLER','CELLNAME','CELLID']
        poskeysfill = [7,8,9]
        df2giom.rename(columns={"Cell Name": "CELLNAME"},inplace=True)
        df2giom = pd.merge(df2giom,cdfct2g['CELLNAME'],how = 'right',on=['CELLNAME'])
        for i in range(len(df2giom)):
            for j in range(len(df2giom.keys())):
                sheet.cell(row=i+2, column=j+1, value=df2giom.iloc[i,j])
        doctofill.save(filename)
    #==============3G==============
    if not cdfct3g.empty:
        #======Copy tmp,change name,load file===
        #Copy file
        shutil.copy(tmpfoldername+"\\HUAWEI_UMTS_CUSTOMIZED_CELL_.xlsx", fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_UMTS_CUSTOMIZED_CELL_.xlsx")
        filename = fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_UMTS_CUSTOMIZED_CELL_"+cname+".xlsx"
        os.rename(fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_UMTS_CUSTOMIZED_CELL_.xlsx", filename)
        #Load file
        doctofill = openpyxl.load_workbook(filename)
        # Cluster cells sheet
        sheet = doctofill[sheetname]
        keysfill = ['RNCID','CELLNAME','CELLID']
        poskeysfill = [7,8,9]
        for i in range(len(cdfct3g)):
            for j,k in enumerate(keysfill):
                sheet.cell(row=i+2, column=poskeysfill[j], value=cdfct3g.loc[i,k])
        doctofill.save(filename)
    #==============4G==============
    if not cdfct4g.empty:
        #======Copy tmp,change name,load file===
        #Copy file
        shutil.copy(tmpfoldername+"\\HUAWEI_LTE_CUSTOMIZED_CELL_.xlsx", fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_LTE_CUSTOMIZED_CELL_.xlsx")
        filename = fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_LTE_CUSTOMIZED_CELL_"+cname+".xlsx"
        os.rename(fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_LTE_CUSTOMIZED_CELL_.xlsx", filename)
        #Load file
        doctofill = openpyxl.load_workbook(filename)
        # Cluster cells sheet
        sheet = doctofill[sheetname]
        keysfill = ['ENODEB_ID','CELLNAME','CELLID']
        poskeysfill = [7,8,9]
        for i in range(len(cdfct4g)):
            for j,k in enumerate(keysfill):
                sheet.cell(row=i+2, column=poskeysfill[j], value=cdfct4g.loc[i,k])
        doctofill.save(filename)
    #==============5G==============
    if not cdfct5g.empty:
        #======Copy tmp,change name,load file===
        #Copy file
        shutil.copy(tmpfoldername+"\\HUAWEI_NR_CUSTOMIZED_CELL_.xlsx", fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_NR_CUSTOMIZED_CELL_.xlsx")
        filename = fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_NR_CUSTOMIZED_CELL_"+cname+".xlsx"
        os.rename(fname+'\\'+cname+"\\"+tmpfoldername+"\\HUAWEI_NR_CUSTOMIZED_CELL_.xlsx", filename)
        #Load file
        doctofill = openpyxl.load_workbook(filename)
        # Cluster cells sheet
        sheet = doctofill[sheetname]
        keysfill = ['ENODEB_ID','CELLNAME','CELLID']
        poskeysfill = [7,8,9]
        for i in range(len(cdfct5g)):
            for j,k in enumerate(keysfill):
                sheet.cell(row=i+2, column=poskeysfill[j], value=cdfct5g.loc[i,k])
        doctofill.save(filename)

        #==============Guardar sites ERICSSON==================
        ericsson_sites = cdfct.loc[cdfct['VENDOR'] == 'ERICSSON', 'SITE'].dropna().unique()
        ericsson_sites_str = ','.join(ericsson_sites)

        # Guardar en archivo .txt dentro de TMP_IOM
        ericsson_txt_path = os.path.join(fname, cname, tmpfoldername, 'ERICSSON_SITES.txt')
        with open(ericsson_txt_path, 'w', encoding='utf-8') as f:
            f.write(ericsson_sites_str)


def nodezone(n):
    if n[2].isdigit():
        zone = 'ZR'
    else:
        zone = 'ZN'
    return zone

def query_rets_cluster(cdfct,cname,fname):
    #========Create folder cluster========
    create_folder(fname+'\\'+cname)
    foldername = 'RETs'
    create_folder(fname+'\\'+cname+'\\'+foldername)
    indexgeneral = ((cdfct.loc[:,'OPERATOR NAME']== 'ORANGE')|(cdfct.loc[:,'OPERATOR NAME']== 'Orange'))&(cdfct.loc[:,'STATUS']== 1)
    indextech = (cdfct.loc[:,'TECH']=='3G')|(cdfct.loc[:,'TECH']=='4G')|(cdfct.loc[:,'TECH']=='5G')
    cdfctfilter = cdfct.loc[indexgeneral&indextech].copy().reset_index()
    #==============Query current RETs ATAE================
    queriesatae = ['DSP RETSUBUNIT:;','DSP RETDEVICEDATA:;']
    q3500 = 'LST NRDUCELLTRPBEAM:;'
    nodes = np.unique(cdfctfilter.NODE)
    Zones = ['ZN','ZR']
    #indexnodes = np.unique(cdfctfilter.NODE,return_index=True)
    #ataes = cdfctfilter.loc[indexnodes[1],'OSS'].tolist()
    ataes = []
    for n in nodes:
        atae = np.unique(cdfctfilter.loc[cdfctfilter.loc[:,'NODE'] == n,'ZONA'])
        atae = atae[np.argmax([len(x) for x in atae])]
        if '\\' in atae:
            atae = atae.replace('\\','')
        ataes.append(atae)
    for atae in np.unique(ataes):
        for zone in Zones:
            qret = ''
            for nel,n in enumerate(nodes):
                if (nodezone(n) == zone)&(ataes[nel]==atae):
                    for q in queriesatae:
                        qret = qret+q+'{'+n+'}\n'
                    if any(cdfctfilter.loc[cdfctfilter.NODE == n,'Band']=='P'):
                        qret = qret+q3500+'{'+n+'}\n'   
            if len(qret) > 0:
                savetxt(qret,fname+'\\'+cname+'\\'+foldername+'\\QRETs_'+cname+'_'+atae+'_'+zone)
    #==============Query RETs HEIDIs================
    querieheidis = "SELECT Fecha, neid, subname,deviceno,tilt FROM nodemodule_retsubunit WHERE neid IN ('"+"','".join(nodes)+"') AND Fecha <= '"+datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+"';"
    savetxt(querieheidis,fname+'\\'+cname+'\\'+foldername+'\\QHEIDIsRETs_'+cname)

def browsecsv_retheidis(): 
    global filename
    filename = filedialog.askopenfilename(initialdir = os.getcwd(), 
    title = "Select a csv file (RETs HEIDIs)", 
    filetypes = (("csv file","*.csv*"), ("all files", "*.*"))) 
    if len(filename) == 0:
        return 0
    global dfretheidis
    dfretheidis= pd.read_csv(filename, sep=';')
    dfretheidis.Fecha = pd.to_datetime(dfretheidis.Fecha)

#===================================PLOT KPIs======================================


def plotkpi(df,colsel,dfkeysite,dfkeycellnamex,namexsel,target,sectorssel=[],bandssel=[],ylabel='',yinter=[],saveok = False):
   
    #check inputs

    if colsel == '':
        messagebox.showwarning(message="Select a KPI from the KPI list", title="Warning")
        return
    if namexsel == '':
        messagebox.showwarning(message="Select a site from the site list", title="Warning")
        return
    if type(colsel) == str:
        dfplot = df.loc[df.loc[:,dfkeysite] == namexsel,['Date',colsel,'Sector','Band']]
    else:
        dfplot = df.loc[df.loc[:,dfkeysite] == namexsel,colsel+['Date','Sector','Band']]

    if sectorssel == []: #select all
       sectorssel = np.unique(dfplot.Sector.tolist()) 

    if bandssel == []: #select all
        bandssel = np.unique(dfplot.Band.tolist()) 

    
    #constants
    Colours = ['b','r','g','y','m']

    #let's work
    
    #Delete data not available
    if type(colsel) == str:
        dfplot.loc[dfplot.loc[:,colsel] == '/0',colsel] = np.nan
        dfplot.loc[dfplot.loc[:,colsel] == '-',colsel] = np.nan
    else:
        for cs in colsel:
            dfplot.loc[dfplot.loc[:,cs] == '/0',cs] = np.nan
            dfplot.loc[dfplot.loc[:,cs] == '-',cs] = np.nan

    #check sectors
    okcheck = True
    for sectorsel in sectorssel:
        #check bands
        for bandsel in bandssel:
            if not any((dfplot.Sector == sectorsel)&(dfplot.Band == bandsel)):
                print('plot not found Sector '+sectorsel+' Band '+bandsel+' KPI '+colsel)
                okcheck = False
    if (not okcheck) | (len(dfplot) == 0):
        return [],[]       

    #Turn to float data
    if type(colsel) == str:
        if any(np.array([type(x) for x in dfplot[colsel].values])==str):
            dfplot[colsel] = dfplot[colsel].str.replace(",", ".")
            dfplot[colsel] = pd.to_numeric(dfplot[colsel],errors='coerce')       
    else:
        for cs in colsel:
            if any(np.array([type(x) for x in dfplot[cs].values])==str):
                dfplot[cs] = dfplot[cs].str.replace(",", ".")
                dfplot[cs] = pd.to_numeric(dfplot[cs],errors='coerce') 
    
    okcreateplt = True
    for i in range(len(sectorssel)):
        if type(colsel) == str:
            dfploti = dfplot.loc[(dfplot.loc[:,'Sector']==sectorssel[i]),['Date',colsel,'Band']]
        else:
            dfploti = dfplot.loc[(dfplot.loc[:,'Sector']==sectorssel[i]),colsel+['Date','Band']]
        #Prepare legend
        cellnamex = np.unique(df.loc[dfploti.index,dfkeycellnamex].tolist())
        dlt = np.array([])
        for j in range(len(cellnamex)):
            if len(cellnamex[j])==7:
                if (cellnamex[j][-2] not in bandssel) | (cellnamex[j][-1] not in sectorssel):
                    dlt = np.append(dlt,j)
            else:
                if (cellnamex[j][-3] not in bandssel) | (cellnamex[j][-2] not in sectorssel):
                    dlt = np.append(dlt,j)
        legend = np.delete(cellnamex,dlt.astype('int'))
        dfploti.reset_index(drop=True, inplace=True)
        if okcreateplt:
            if type(colsel) == str:
                fig,ax = plot(dfploti,bandssel,len(sectorssel),i,sectorssel[i],colsel,target,ylabel,legend,Colours[0:len(bandssel)],saveok=saveok,yinter=yinter)
                okcreateplt = False
            else:
                fig,ax = plotfill(dfploti,bandssel,len(sectorssel),i,sectorssel[i],colsel,target,ylabel,legend,Colours[0:len(bandssel)],saveok=saveok,yinter=yinter)
                okcreateplt = False
        else:
            if type(colsel) == str:
                plot(dfploti,bandssel,len(sectorssel),i,sectorssel[i],colsel,target,ylabel,legend,Colours[0:len(bandssel)],fig = fig,ax = ax,saveok=saveok,yinter=yinter)
            else:
                plotfill(dfploti,bandssel,len(sectorssel),i,sectorssel[i],colsel,target,ylabel,legend,Colours[0:len(bandssel)],fig = fig,ax = ax,saveok=saveok,yinter=yinter)
    return fig,ax

def discard_plot(dfplot,colname):
    typedata = np.array([type(x) for x in dfplot.loc[:,colname].tolist()]) 
    discardstr = (typedata == str)
    dfplot.loc[discardstr,colname] = np.nan
    #dfplot = dfplot.loc[(dfplot.loc[:,colname] != np.nan),:]
    #dfplot = dfplot.loc[~(dfplot.loc[:,colname].isnull()),:]
    return dfplot

def plot(dfplot,bandssel,nrows,rowsel,sector,colname,target='',ylabel='',legends='',colour = 'b',fig = '',ax = [''],saveok=False,yinter=[]): 
    # the figure that will contain the plot 
    if (fig == '') | (ax[0] == ''):
        fig, ax = plt.subplots(nrows,1)
    fig.suptitle(colname,fontsize = 10, fontweight ='bold')
    total = 25
    inter = int(str(((dfplot.Date.max()-dfplot.Date.min())/total).round('H'))[7:9])+((dfplot.Date.max()-dfplot.Date.min())/total).round('d').days*24
    if nrows > 1:
        ax[rowsel].xaxis.set_major_locator(mdates.HourLocator(interval=inter))
        ax[rowsel].xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%Y %H'))
        if yinter != []:
            ax[rowsel].set_ylim([yinter[0], yinter[1]])
        for i in range(len(bandssel)):
            index = (dfplot.loc[:,'Band']==bandssel[i])
            dfploti = dfplot.loc[index].copy()
            dfploti = discard_plot(dfploti,colname)
            ax[rowsel].plot(dfploti['Date'], dfploti[colname], color = colour[i])
        if target != '':
            ax[rowsel].hlines(y = float(target), xmin=np.min(dfplot.iloc[:,0]), xmax=np.max(dfplot.iloc[:,0]),color = 'r', linestyle = '--')
        ax[rowsel].set_xlabel('Dates', color = 'black',fontsize=12)
        ax[rowsel].set_ylabel(ylabel, color = 'black',fontsize=12)
        if target != '':
            ax[rowsel].legend(np.append(legends,'Target'))
        else:
            ax[rowsel].legend(legends)
        ax[rowsel].set_title('Sector ' + sector)
        ax[rowsel].grid(color='powderblue', linestyle='--',which='both')
    else:
        ax.xaxis.set_major_locator(mdates.HourLocator(interval=inter))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%Y %H'))
        if yinter != []:
            ax.set_ylim([yinter[0], yinter[1]])
        for i in range(len(bandssel)):
            index = (dfplot.loc[:,'Band']==bandssel[i])
            dfploti = dfplot.loc[index].copy()
            dfploti = discard_plot(dfploti,colname)
            #ax.plot(dfploti['Date'], dfploti[colname], color = colour[i])
            posord = np.argsort(dfploti['Date'].tolist())
            ax.plot(np.array(dfploti['Date'].tolist())[posord], np.array(dfploti[colname].tolist())[posord], color = colour[i])
        if target != '':
            ax.hlines(y = float(target), xmin=np.min(dfplot.iloc[:,0]), xmax=np.max(dfplot.iloc[:,0]),color = 'r', linestyle = '--')
        ax.set_xlabel('Dates', color = 'black',fontsize=12)
        ax.set_ylabel(ylabel, color = 'black',fontsize=12)
        if target != '':
            ax.legend(np.append(legends,'Target'))
        else:
            ax.legend(legends)
        #ax.set_title('Sector ' + sector)
        ax.grid(color='powderblue', linestyle='--',which='both')
    fig.autofmt_xdate()
    plt.xticks(rotation=90)
    #manager = plt.get_current_fig_manager()
    #manager.full_screen_toggle()
    if not saveok:
        fig.show()
    return fig,ax

def plotfill(dfplot,bandssel,nrows,rowsel,sector,colname,target='',ylabel='',legends='',colour = 'b',fig = '',ax = [''],saveok=False,yinter=[]): 
    # the figure that will contain the plot 
    if (fig == '') | (ax[0] == ''):
        fig, ax = plt.subplots(nrows,1)
    fig.suptitle(colname[0],fontsize = 10, fontweight ='bold')
    total = 25
    inter = int(str(((dfplot.Date.max()-dfplot.Date.min())/total).round('H'))[7:9])+((dfplot.Date.max()-dfplot.Date.min())/total).round('d').days*24
    if nrows > 1:
        ax[rowsel].xaxis.set_major_locator(mdates.HourLocator(interval=inter))
        ax[rowsel].xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%Y %H'))
        if yinter != []:
            ax[rowsel].set_ylim([yinter[0], yinter[1]])
        for i in range(len(bandssel)):
            index = (dfplot.loc[:,'Band']==bandssel[i])
            dfploti = dfplot.loc[index].copy()
            ax[rowsel].fill_between(dfploti['Date'], dfploti[colname], color = colour[i])
        if target != '':
            ax[rowsel].hlines(y = float(target), xmin=np.min(dfplot.iloc[:,0]), xmax=np.max(dfplot.iloc[:,0]),color = 'r', linestyle = '--')
        ax[rowsel].set_xlabel('Dates', color = 'black',fontsize=12)
        ax[rowsel].set_ylabel(ylabel, color = 'black',fontsize=12)
        if target != '':
            ax[rowsel].legend(np.append(legends,'Target'))
        else:
            ax[rowsel].legend(legends)
        ax[rowsel].set_title('Sector ' + sector)
        ax[rowsel].grid(color='powderblue', linestyle='--',which='both')
    else:
        ax.xaxis.set_major_locator(mdates.HourLocator(interval=inter))
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%Y %H'))
        if yinter != []:
            ax.set_ylim([yinter[0], yinter[1]])
        for i in range(len(bandssel)):
            index = (dfplot.loc[:,'Band']==bandssel[i])
            dfploti = dfplot.loc[index].copy()
            #ax.plot(dfploti['Date'], dfploti[colname], color = colour[i])
            posord = np.argsort(dfploti['Date'].tolist())
            ax.fill_between(np.array(dfploti['Date'].tolist())[posord], np.array(dfploti[colname].tolist())[posord], color = colour[i])
        if target != '':
            ax.hlines(y = float(target), xmin=np.min(dfplot.iloc[:,0]), xmax=np.max(dfplot.iloc[:,0]),color = 'r', linestyle = '--')
        ax.set_xlabel('Dates', color = 'black',fontsize=12)
        ax.set_ylabel(ylabel, color = 'black',fontsize=12)
        if target != '':
            ax.legend(np.append(legends,'Target'))
        else:
            ax.legend(legends)
        #ax.set_title('Sector ' + sector)
        ax.grid(color='powderblue', linestyle='--',which='both')
    fig.autofmt_xdate()
    plt.xticks(rotation=90)
    #manager = plt.get_current_fig_manager()
    #manager.full_screen_toggle()
    if not saveok:
        fig.show()
    return fig,ax

def plotandsave(df,colsel,dfkeysite,dfkeycellnamex,namexsel,target,sectorssel,bandssel,ylabel,yinter,filename):
    fig,ax = plotkpi(df,colsel,dfkeysite,dfkeycellnamex,namexsel,target,sectorssel,bandssel,ylabel,yinter,True)
    if fig != []:
        fig.set_size_inches(12, 7)
        save_plot(fig,filename)

def save_plot(fig,filename): 
    fig.savefig(filename+'.png')
    plt.close(fig)

#======================BAL OFFSET=============================   
def localcellid(tech,zone):
    if zone == 'ZR':
        ZR = {"M1":120,"N1":110,"T1":130,"L1":100,"LB1":140,"Y1":106,"M2":121,"N2":111,"T2":131,"L2":101,"LB2":141,"Y2":107,"M3":122,"N3":112,"T3":132,"L3":102,"LB3":142,"Y3":108,
        "M4":123,"N4":113,"T4":133,"L4":103,"LB4":143,"Y4":109,"M5":124,"N5":114,"T5":134,"L5":104,"LB5":144,"M6":125,"N6":115,"T6":135,"L6":105,"LB6":145}
        lci = -1
        if tech in ZR.keys():
            lci = ZR[tech]
        return lci
    else:
        ZN = {"M1":12,"N1":6,"T1":18,"L1":0,"LB1":40,"Y1":90,"M2":13,"N2":7,"T2":19,"L2":1,"LB2":41,"Y2":91,"M3":14,"N3":8,"T3":20,"L3":2,"LB3":42,"Y3":92,
        "M4":15,"N4":9,"T4":21,"L4":3,"LB4":43,"Y4":93,"M5":16,"N5":10,"T5":22,"L5":4,"LB5":44,"M6":17,"N6":11,"T6":23,"L6":5,"LB6":45}
        lci = -1
        if tech in ZN.keys():
            lci = ZN[tech]
        return lci
    
def techfromlci(lci):
    if lci > 100:
        ZR = {"M1":120,"N1":110,"T1":130,"L1":100,"LB1":140,"Y1":106,"M2":121,"N2":111,"T2":131,"L2":101,"LB2":141,"Y2":107,"M3":122,"N3":112,"T3":132,"L3":102,"LB3":142,"Y3":108,
        "M4":123,"N4":113,"T4":133,"L4":103,"LB4":143,"Y4":109,"M5":124,"N5":114,"T5":134,"L5":104,"LB5":144,"M6":125,"N6":115,"T6":135,"L6":105,"LB6":145}
        ZR_inv = {v: k for k, v in ZR.items()}
        tech = 'Not Found'
        if lci in ZR_inv.keys():
            tech = ZR_inv[lci]
        return tech
    else:
        ZN = {"M1":12,"N1":6,"T1":18,"L1":0,"LB1":40,"Y1":90,"M2":13,"N2":7,"T2":19,"L2":1,"LB2":41,"Y2":91,"M3":14,"N3":8,"T3":20,"L3":2,"LB3":42,"Y3":92,
        "M4":15,"N4":9,"T4":21,"L4":3,"LB4":43,"Y4":93,"M5":16,"N5":10,"T5":22,"L5":4,"LB5":44,"M6":17,"N6":11,"T6":23,"L6":5,"LB6":45}
        ZN_inv = {v: k for k, v in ZN.items()}
        tech = 'Not Found'
        if lci in ZN_inv.keys():
            tech = ZN_inv[lci]
        return tech
    
def techfromci(ci):
    if ci > 100:
        ZR = {"M1":120,"N1":110,"T1":130,"L1":100,"LB1":140,"Y1":106,"M2":121,"N2":111,"T2":131,"L2":101,"LB2":141,"Y2":107,"M3":122,"N3":112,"T3":132,"L3":102,"LB3":142,"Y3":108,
        "M4":123,"N4":113,"T4":133,"L4":103,"LB4":143,"Y4":109,"M5":124,"N5":114,"T5":134,"L5":104,"LB5":144,"M6":125,"N6":115,"T6":135,"L6":105,"LB6":145}
        ZR_inv = {v: k for k, v in ZR.items()}
        tech = 'Not Found'
        if ci in ZR_inv.keys():
            tech = ZR_inv[ci]
        return tech
    else:
        ZN = {"M1":20,"N1":10,"T1":30,"L1":0,"LB1":40,"Y1":90,"M2":21,"N2":11,"T2":31,"L2":1,"LB2":41,"Y2":91,"M3":22,"N3":12,"T3":32,"L3":2,"LB3":42,"Y3":92,
        "M4":23,"N4":13,"T4":33,"L4":3,"LB4":43,"Y4":93,"M5":24,"N5":14,"T5":34,"L5":4,"LB5":44,"M6":25,"N6":15,"T6":35,"L6":5,"LB6":45}
        ZN_inv = {v: k for k, v in ZN.items()}
        tech = 'Not Found'
        if ci in ZN_inv.keys():
            tech = ZN_inv[ci]
        return tech

def query_bal_cluster(cdfct,cname,fname):
    nfbal = 'BALANCEOS'
    create_folder(fname+'\\'+cname+'\\'+nfbal)
    filter4g = cdfct.TECH == '4G'
    nodes = cdfct.loc[filter4g,'NODE'].tolist()
    regions = cdfct.loc[filter4g,'ZONA'].tolist()
    nodes,indexnodes = np.unique(nodes,return_index=True)
    regions = [regions[x] for x in indexnodes]
    sczneste = ''
    scznsur = ''
    sczreste = ''
    sczrsur= ''
    sczncentro = ''
    sczrcentro = ''
    for i,node in enumerate(nodes):
        zone = nodezone(node)
        if zone == 'ZR':
            if (regions[i] == 'Zona ESTE'):
                sczreste +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                sczreste +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
            elif (regions[i] == 'Zona SUR'):
                sczrsur +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                sczrsur +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
            else: 
                sczrcentro +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                sczrcentro +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
        else:
            if (regions[i] == 'Zona ESTE'):
                sczneste +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                sczneste +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
            elif (regions[i] == 'Zona SUR'):
                scznsur +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                scznsur +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
            else: 
                sczncentro +='LST EUTRANINTERFREQNCELL:;{'+node+'}\n'
                sczncentro +='LST EUTRANINTRAFREQNCELL:;{'+node+'}\n'
    if len(sczreste) > 0:   
        savetxt(sczreste,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_ESTE')
    if len(sczrsur) > 0:   
        savetxt(sczrsur,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_SUR')
    if len(sczrcentro) > 0:   
        savetxt(sczrcentro,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_CENTRO')
    if len(sczneste) > 0:   
        savetxt(sczneste,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_ESTE')
    if len(scznsur) > 0:   
        savetxt(scznsur,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_SUR')
    if len(sczncentro) > 0:   
        savetxt(sczncentro,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_CENTRO')
    return


def query_bal_cluster_cosector(cdfct,cname,fname):
    nfbal = 'BALANCEOS'
    create_folder(fname+'\\'+cname+'\\'+nfbal)
    cdfct.loc[:,'Sector']=[x[-2] for x in cdfct.CELLNAMEX.tolist()]
    cdfct.loc[:,'Band']=[x[-3] for x in cdfct.CELLNAMEX.tolist()]
    indexct = (cdfct.TECH == '4G')&(cdfct.STATUS == 1)
    cdfctq = cdfct.loc[indexct,:]
    sites = np.unique(cdfctq.SITE.tolist())
    sczneste = ''
    scznsur = ''
    sczreste = ''
    sczrsur= ''
    sczncentro = ''
    sczrcentro = ''
    for site in sites:
        indexsite = cdfctq.SITE == site
        node = cdfctq.loc[indexsite,'NODE'].tolist()[0]
        zone = nodezone(node)
        region = cdfctq.loc[indexsite,'ZONA'].tolist()[0]
        vendor = cdfctq.loc[indexsite,'VENDOR'].tolist()[0]
        if (((region == 'Zona ESTE')|(region == 'Zona SUR')|(region == 'Zona CENTRO')) & (vendor == 'HUAWEI')):
            sectors = np.unique(cdfctq.loc[indexsite,'Sector'].tolist())
            for sector in sectors:
                indexsector = cdfctq.Sector == sector
                bands = np.unique(cdfctq.loc[indexsite & indexsector,'Band'].tolist())
                for el,band in enumerate(bands):
                    lcid = localcellid(band+str(sector),zone)
                    #check with other bands
                    if lcid != -1:
                        for i in range(len(bands)-1):
                            banddest = bands[np.mod(el+i+1,len(bands))]
                            indexbanddest = cdfctq.Band == banddest
                            cid = cdfctq.loc[indexsite&indexsector&indexbanddest,'CELLID'].tolist()[0]
                            enbid = cdfctq.loc[indexsite&indexsector&indexbanddest,'ENODEB_ID'].tolist()[0]
                            if zone == 'ZR':
                                if (region == 'Zona ESTE'):
                                    sczreste +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
                                elif (region == 'Zona SUR'):
                                    sczrsur +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
                                else: 
                                        sczrcentro +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
                            else:
                                if (region == 'Zona ESTE'):
                                    sczneste +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
                                elif (region == 'Zona SUR'):
                                    scznsur +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
                                else: 
                                    sczncentro +='LST EUTRANINTERFREQNCELL:LOCALCELLID='+str(lcid)+',ENODEBID='+str(enbid)+',CELLID='+str(cid)+';{'+node+'}\n'
        # Incluir comandos de Ericsson para obtener su QBAL
    if len(sczreste) > 0:   
        savetxt(sczreste,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_ESTE')
    if len(sczrsur) > 0:   
        savetxt(sczrsur,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_SUR')
    if len(sczrcentro) > 0:   
        savetxt(sczrcentro,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZR_CENTRO')
    if len(sczneste) > 0:   
        savetxt(sczneste,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_ESTE')
    if len(scznsur) > 0:   
        savetxt(scznsur,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_SUR')
    if len(sczncentro) > 0:   
        savetxt(sczncentro,fname+'\\'+cname+'\\'+nfbal+'\\Qbal_'+cname+'_ZN_CENTRO')
    return

def lcidtobandsect(lcid,zone):
    if zone == 'ZR':
        ZRI = {"M1":120,"N1":110,"T1":130,"L1":100,"LB1":140,"Y1":106,"M2":121,"N2":111,"T2":131,"L2":101,"LB2":141,"Y2":107,"M3":122,"N3":112,"T3":132,"L3":102,"LB3":142,"Y3":108,
        "M4":123,"N4":113,"T4":133,"L4":103,"LB4":143,"Y4":109,"M5":124,"N5":114,"T5":134,"L5":104,"LB5":144,"M6":125,"N6":115,"T6":135,"L6":105,"LB6":145}
        ZR = {v: k for k, v in ZRI.items()}
        return ZR[lcid]
    else:
        ZNI = {"M1":12,"N1":6,"T1":18,"L1":0,"LB1":40,"Y1":90,"M2":13,"N2":7,"T2":19,"L2":1,"LB2":41,"Y2":91,"M3":14,"N3":8,"T3":20,"L3":2,"LB3":42,"Y3":92,
        "M4":15,"N4":9,"T4":21,"L4":3,"LB4":43,"Y4":93,"M5":16,"N5":10,"T5":22,"L5":4,"LB5":44,"M6":17,"N6":11,"T6":23,"L6":5,"LB6":45}
        ZN = {v: k for k, v in ZNI.items()}
        return ZN[lcid]

def check_bal(df):

    keys = ['Cell individual offset(dB)','Cell offset(dB)','NE','NE neigh','BandSect FROM','BandSect TO']

    #Data preprocessing

    df['Cell individual offset(dB)'] = pd.to_numeric(df['Cell individual offset(dB)'].str.replace('dB',''),errors='coerce')
    df['Cell offset(dB)'] = pd.to_numeric(df['Cell offset(dB)'].str.replace('dB',''),errors='coerce')


    #=================CELL IND OFFSET = -CELL OFFSET=====================

    df['CHECK_CINDOFF_COFF'] = df['Cell individual offset(dB)'] == -df['Cell offset(dB)']

    #=================RECIPROCO=====================

    for i in range(len(df)):
        #coger valor cell individual offset
        ciofforigen = df.loc[i,'Cell individual offset(dB)']
        #buscar el indice donde este la relación inversa
        j = (df['NE']==df.loc[i,'NE neigh'])&(df['BandSect FROM']==df.loc[i,'BandSect TO'])&(df['NE neigh']==df.loc[i,'NE'])&(df['BandSect TO']==df.loc[i,'BandSect FROM'])
        #chequear si hay algún j
        if any(j):
            cioffneigh = df.loc[j,'Cell individual offset(dB)'].tolist()[0]
            if ciofforigen == -cioffneigh:#bien
                df.loc[i,'CHECK_CINDOFF_RECIPROCO'] = 'OK'
            else:
                df.loc[i,'CHECK_CINDOFF_RECIPROCO'] = cioffneigh
        else:
            df.loc[i,'CHECK_CINDOFF_RECIPROCO'] = 'N/D'

    return df

def complete_dfbal(dfbal):
    #data type
    dfbal['Local cell ID'] = pd.to_numeric(dfbal['Local cell ID'],errors='coerce')
    dfbal['eNodeB ID'] = pd.to_numeric(dfbal['eNodeB ID'],errors='coerce')
    dfbal['Cell ID'] = pd.to_numeric(dfbal['Cell ID'],errors='coerce')
    dfbal['BandSect FROM'] = [techfromlci(lci) for lci in dfbal['Local cell ID']]
    dfbal['Band FROM'] = [x[0] for x in dfbal['BandSect FROM']]
    dfbal['Sector FROM'] = [x[1] for x in dfbal['BandSect FROM']]
    dfbal['BandSect TO'] = [techfromci(ci) for ci in dfbal['Cell ID']]
    dfbal['Band TO'] = [x[0] for x in dfbal['BandSect TO']]
    dfbal['Sector TO'] = [x[1] for x in dfbal['BandSect TO']]
    conop = (dfct['OPERATOR NAME'] == 'ORANGE')|(dfct['OPERATOR NAME'] == 'Orange')
    dfct4g = dfct.loc[(dfct.TECH == '4G')&conop,:].copy().reset_index()
    dfct4g['NE'] = dfct4g['NODE']
    dfct4g['Cell ID'] = pd.to_numeric(dfct4g['CELLID'],errors='coerce')
    dfct4g['eNodeB ID'] = pd.to_numeric(dfct4g['ENODEB_ID'],errors='coerce')
    dfct4g['NE neigh'] = dfct4g['NE']
    dfct4g['Site neigh'] = dfct4g['SITE']
    #NE neigh
    dfbal = pd.merge(dfbal,dfct4g[['NE neigh','eNodeB ID','Cell ID']].drop_duplicates(),how = 'left',on=['eNodeB ID','Cell ID'])
    #Site
    dfbal = pd.merge(dfbal,dfct4g[['NE','SITE']].drop_duplicates(),how = 'left',on=['NE'])
    dfbal.rename(columns={"SITE": "Site"},inplace=True)
    #Site neigh
    dfbal = pd.merge(dfbal,dfct4g[['Site neigh','eNodeB ID','Cell ID']].drop_duplicates(),how = 'left',on=['eNodeB ID','Cell ID'])
    """
    nodes =  np.unique(dfbal['NE'].tolist())
    for n in nodes:
        ns = n[0:8]
        index = dfbal['NE'] == n
        if any(index):
            sitel = dfct.loc[dfct['NODE'] == ns,'SITE'].tolist()
            if len(sitel) > 0:
                site = sitel[0]
            else:
                site = 'Not found'
            dfbal.loc[index,'Site'] = site
    enodebids = np.unique(dfbal['eNodeB ID'].tolist())
    for enbi in enodebids:
        index = dfbal['eNodeB ID'] == enbi
        #Cuidado zona roja
        nei = dfct.loc[(dfct['ENODEB_ID'] == enbi)|(dfct['ENODEB_ID'] == str(enbi)),'NODE'].tolist()
        if len(nei) > 0:
            nei = nei[0]
        else:
            nei = 'Not found'
        dfbal.loc[index,'NE neigh'] = nei
    nodesdest =  np.unique(dfbal['NE neigh'].tolist())
    for n in nodesdest:
        ns = n[0:8]
        index = dfbal['NE neigh'] == n
        if any(index):
            site = dfct.loc[dfct['NODE'] == ns,'SITE'].tolist()
            if len(site) > 0:
                site = site[0]
            else:
                site = 'Not found'
            dfbal.loc[index,'Site neigh'] = site
    """
    """
    nodes = dfbal.NE.tolist()
    for el,n in enumerate(nodes):
        indexnode = (dfct.loc[:,'NODE'] == n) & ((dfct.loc[:,'OPERATOR NAME'] == 'ORANGE')|(dfct.loc[:,'OPERATOR NAME'] == 'Orange'))
        dfbal.loc[el,'Site'] = dfct.loc[indexnode,'SITE'].tolist()[0]
        bandsec = lcidtobandsect(dfbal.loc[el,'Local cell ID'],nodezone(n))
        dfbal.loc[el,'BandLocal'] = bandsec[0]
        dfbal.loc[el,'SectorLocal'] = bandsec[1]
        enbidi = dfbal.loc[el,'eNodeB ID']
        cellidi = dfbal.loc[el,'Cell ID']
        indexctdest = ((dfct.loc[:,'ENODEB_ID'] == enbidi)|(dfct.loc[:,'ENODEB_ID'] == str(enbidi)))&((dfct.loc[:,'CELLID'] == cellidi)|(dfct.loc[:,'CELLID'] == str(cellidi)))
        if any(indexctdest):
            cellnamexi = dfct.loc[indexctdest,'CELLNAMEX'].tolist()[0]
            dfbal.loc[el,'BandNeigh'] = cellnamexi[-3]
            dfbal.loc[el,'SectorNeigh'] = cellnamexi[-2]
        else:
            print('Not found neighbour for NODE: '+n+' Local Band-Sector: '+bandsec+' ENBID: '+str(enbidi)+' CELLID: '+str(cellidi))
    """
    return dfbal

def process_bal(fname = 'Clusters PRBs Huella'):
    if 'dfct' not in globals():
        load_csv_ctable()
    
    create_folder(fname+"\\ALL\\BALANCEOS")

    command = 'MML Command-----LST EUTRANINTERFREQNCELL'
    if okqbalsaving.get():
        column_names = ['Local cell ID','Mobile country code','Mobile network code','eNodeB ID','Cell ID','Cell individual offset(dB)','Cell offset(dB)','Local cell name','Neighbour cell name']    
    else:
        column_names = ['Local cell ID','Mobile country code','Mobile network code','eNodeB ID','Cell ID','Cell individual offset(dB)','Cell offset(dB)','No handover indicator',
                        'No remove indicator','Blind handover Priority','ANR flag','Local cell name','Neighbour cell name','Cell Measure Priority','Overlap Indicator',
                        'Overlap Range','High Speed Interference Avoid Flag','Neighbor Cell Classification Label','Aggregation Attribute','Overlap Rate(%)','Control Mode',
                        'Overlap Indicator Extension','Neighboring Cell Addition Time','NSA DC UE Cell Individual offset(dB)']
    dfbalinter = process_txt_verthorz_multiple(command,column_names)
    if len(dfbalinter) > 0:
        dfbalinter = complete_dfbal(dfbalinter)
        if okcheckbal.get():
            dfbalinter = check_bal(dfbalinter)
        dfbalinter.to_csv(fname+'\\ALL\\BALANCEOS\\dfbalinter.csv',index=False,sep=';')

    command = 'MML Command-----LST EUTRANINTRAFREQNCELL'
    if okqbalsaving.get():
        column_names = ['Local cell ID','Mobile country code','Mobile network code','eNodeB ID','Cell ID','Cell individual offset(dB)','Cell offset(dB)','Local cell name','Neighbour cell name']
    else:
        column_names = ['Local cell ID','Mobile country code','Mobile network code','eNodeB ID','Cell ID','Cell individual offset(dB)','Cell offset(dB)','No handover indicator',
                        'ANR flag','Local cell name','Neighbour cell name','Cell Measure Priority','Cell Range Expansion(dB)','Neighbor Cell Classification Label',
                        'Control Mode','Attach Cell Switch','High Speed Cell Individual Offset(dB)','Vector Cell Flag','NSA DC UE Cell Individual offset(dB)',
                        'Neighboring Cell Addition Time','Aggregation Attribute']
    dfbalintra = process_txt_verthorz_multiple(command,column_names)
    if len(dfbalintra) > 0:
        dfbalintra = complete_dfbal(dfbalintra)
        if okcheckbal.get():
            dfbalintra = check_bal(dfbalintra)
        dfbalintra.to_csv(fname+'\\ALL\\BALANCEOS\\dfbalintra.csv',index=False,sep=';')
    
    
    return dfbalinter

def process_bal_cluster(cdfct,cname,fname,dfbal):
    dfbalc = pd.DataFrame(columns=dfbal.keys())
    nfbal = 'BALANCEOS'
    create_folder(fname+'\\'+cname+'\\'+nfbal)
    cdfct.loc[:,'Sector']=[x[-2] for x in cdfct.CELLNAMEX.tolist()]
    cdfct.loc[:,'Band']=[x[-3] for x in cdfct.CELLNAMEX.tolist()]
    indexct = (cdfct.TECH == '4G')&(cdfct.STATUS == 1)
    cdfctq = cdfct.loc[indexct,:]
    sites = np.unique(cdfctq.SITE.tolist())
    for site in sites:
        indexsite = cdfctq.SITE == site
        node = cdfctq.loc[indexsite,'NODE'].tolist()[0]
        zone = nodezone(node)
        region = cdfctq.loc[indexsite,'ZONA'].tolist()[0]
        if (region == 'Zona ESTE')|(region == 'Zona SUR')|(region == 'Zona CENTRO'):
            sectors = np.unique(cdfctq.loc[indexsite,'Sector'].tolist())
            for sector in sectors:
                indexsector = cdfctq.Sector == sector
                bands = np.unique(cdfctq.loc[indexsite & indexsector,'Band'].tolist())
                for el,band in enumerate(bands):
                    lcid = localcellid(band+str(sector),zone)
                    #check with other bands
                    if lcid != -1:
                        for i in range(len(bands)-1):
                            banddest = bands[np.mod(el+i+1,len(bands))]
                            indexbanddest = cdfctq.Band == banddest
                            cid = cdfctq.loc[indexsite&indexsector&indexbanddest,'CELLID'].tolist()[0]
                            enbid = cdfctq.loc[indexsite&indexsector&indexbanddest,'ENODEB_ID'].tolist()[0]
                            indexdfbal = (dfbal['eNodeB ID'] == enbid)&(dfbal['Cell ID'] == cid)&(dfbal['Local cell ID'] == lcid)&(dfbal['NE'] == node)
                            if any(indexdfbal):
                                dfbalc = pd.concat([dfbalc,dfbal.loc[indexdfbal,:]])
                            else:
                                dfbalc.loc[len(dfbalc),['eNodeB ID','Cell ID','Local cell ID','NE']] = [enbid,cid,lcid,node]
    fieldsel = ['Site','Local cell name','Neighbour cell name','Cell individual offset(dB)','Cell offset(dB)','BandSect FROM','Band FROM','Sector FROM',
                'BandSect TO','Band TO','Sector TO','Local cell ID','Cell ID']
    dfbalc.loc[:,fieldsel].to_excel(fname+'\\'+cname+'\\'+nfbal+'\\current_offset_'+cname+'.xlsx',index=False)

#==============================POWERS=================================

def script_pot_5g(nodes):
    nrducelltrp = ''
    nrcell = ''
    nrducell = ''
    for n in nodes:
        nrducelltrp+='LST NRDUCELLTRP:;{'+n+'}\n'
        nrcell+='LST NRCELL:;{'+n+'}\n'
        nrducell+='LST NRDUCELL:;{'+n+'}\n'
    return nrducelltrp,nrcell,nrducell

def script_pot_4g(nodes):
    cell = ''
    pdschcfg = ''
    celldlpcpdschpa = ''
    cellchpwrcfg = ''
    for n in nodes:
        cellchpwrcfg += 'LST CELLCHPWRCFG:;{'+n+'}\n'
        cell+='LST CELL:;{'+n+'}\n'
        pdschcfg+='LST PDSCHCFG:;{'+n+'}\n'
        celldlpcpdschpa+='LST CELLDLPCPDSCHPA:;{'+n+'}\n'
    return cell,pdschcfg,celldlpcpdschpa,cellchpwrcfg

def script_pot_3g(nodes,controller3g,cellid3g):
    ulocell = ''
    ulocellcontr = ''
    ucell = ''
    upcpich = ''
    for i,n in enumerate(nodes):
        ucell+='LST UCELL:LSTTYPE=ByCellId,CELLID='+str(cellid3g[i])+';{'+controller3g[i]+'}\n'
        ulocell+='LST ULOCELLSECTOREQM:;{'+n+'}\n'
        #ulocell+= 'LST ULOCELL:MODE=ALLLOCALCELL;{'+n+'}\n'
        ulocellcontr+='LST ULOCELL:IDTYPE=BYNAME,NODEBNAME="'+n+'";{'+controller3g[i]+'}\n'
        upcpich+='LST UPCPICH:CELLID='+str(cellid3g[i])+';{'+controller3g[i]+'}\n'
    ulocell = '\n'.join(np.unique(ulocell.split("\n")))
    ulocellcontr ='\n'.join(np.unique(ulocellcontr.split("\n")))
    if len(ulocellcontr) > 2:
        ulocellcontr = ulocellcontr[1::]+'\n'
    if len(ulocell) > 2:
        ulocell = ulocell[1::]+'\n'
    return ucell,ulocell,upcpich,ulocellcontr

def script_pot_2g(cellname2g,controller2g):
    gtrxdev = ''
    for i,c in enumerate(cellname2g):
        gtrxdev+='LST GTRXDEV:IDTYPE=BYNAME,CELLNAME="'+c+'",TRXIDTYPE=BYID;{'+controller2g[i]+'}\n'
    return gtrxdev

def techfromlcid(lcid,zone):
    if zone == 'ZR':
        ZRa = {"M1":120,"N1":110,"T1":130,"L1":100,"LB1":140,"Y1":106,"M2":121,"N2":111,"T2":131,"L2":101,"LB2":141,"Y2":107,"M3":122,"N3":112,"T3":132,"L3":102,"LB3":142,"Y3":108,
        "M4":123,"N4":113,"T4":133,"L4":103,"LB4":143,"Y4":109,"M5":124,"N5":114,"T5":134,"L5":104,"LB5":144,"M6":125,"N6":115,"T6":135,"L6":105,"LB6":145}
        ZR = {v: k for k, v in ZRa.items()}
        return ZR[lcid]
    else:
        ZNa = {"M1":12,"N1":6,"T1":18,"L1":0,"LB1":40,"Y1":90,"M2":13,"N2":7,"T2":19,"L2":1,"LB2":41,"Y2":91,"M3":14,"N3":8,"T3":20,"L3":2,"LB3":42,"Y3":92,
        "M4":15,"N4":9,"T4":21,"L4":3,"LB4":43,"Y4":93,"M5":16,"N5":10,"T5":22,"L5":4,"LB5":44,"M6":17,"N6":11,"T6":23,"L6":5,"LB6":45}
        ZN = {v: k for k, v in ZNa.items()}
        return ZN[lcid]
    
def query_power_ct(ct,cname,scriptZN,scriptZR):
    if len(ct) > 0:
        #----------2G--------------
        controller2g = np.array(ct.loc[ct['TECH']=='2G','CONTROLLER'].tolist())
        controller2gu = np.unique(controller2g)
        zones2g = np.array(ct.loc[ct['TECH']=='2G','ZONA'].tolist())
        nodes2g = np.array(ct.loc[ct['TECH']=='2G','NODE'].tolist())
        cellname2g = np.array(ct.loc[ct['TECH']=='2G','CELLNAME'].tolist())
        #----------3G--------------
        controller3g = np.array(ct.loc[ct['TECH']=='3G','CONTROLLER'].tolist())
        controller3gu = np.unique(controller3g)
        nodes3g = np.array(ct.loc[ct['TECH']=='3G','NODE'].tolist())
        nodes3gu = np.unique(nodes3g)
        cellid3g = np.array(ct.loc[ct['TECH']=='3G','CELLID'].tolist())
        zones3g = np.array(ct.loc[ct['TECH']=='3G','ZONA'].tolist())
        #----------4G--------------
        nodes4g = np.array(ct.loc[ct['TECH']=='4G','NODE'].tolist())
        nodes4gu = np.unique(nodes4g)
        zones4g = np.array(ct.loc[ct['TECH']=='4G','ZONA'].tolist())
        #----------5G--------------
        nodes5g = np.array(ct.loc[ct['TECH']=='5G','NODE'].tolist())
        nodes5gu = np.unique(nodes5g)
        zones5g = np.array(ct.loc[ct['TECH']=='5G','ZONA'].tolist())
        #-----------Create queries
        #documento_sites = openpyxl.load_workbook('Sites.xlsx')
        #documento_sites3G = openpyxl.load_workbook('Sites3G.xlsx')
        #sheet_comando = documento_sites['Comando']
        #sheet_comando3G = documento_sites3G['Comando']
        #query5g = comando_potencias_5g(nodes5gu, sheet_comando, fecha)
        #query4g = comando_potencias_4g(nodes4gu, sheet_comando, fecha)
        #query3g = comando_potencias_3g(nodes3gu, sheet_comando, fecha)
        #query3gmaxtxpwr  = comando_maxtxpwr_3G(controller3gu, sheet_comando3G, fecha)
        #query2g = comando_potencias_2g(controller2gu, sheet_comando, fecha)
        #-------script potencias-----------
        zones = ['Zona ESTE','Zona SUR','Zona CENTRO','Zona NORTE','Zona 5']
        fname = 'Response queries powers'
        create_folder(fname)
        create_folder(fname+'\\'+cname)
        
        for iz,zone in enumerate(zones):
            zn2g = [i for i,n in enumerate(nodes2g[zone == zones2g]) if not ZRnode(n)]
            zr2g = [i for i,n in enumerate(nodes2g[zone == zones2g]) if ZRnode(n)]
            zn2g = np.where(zone == zones2g)[0][zn2g]
            zr2g = np.where(zone == zones2g)[0][zr2g]
            zn3g = [i for i,n in enumerate(nodes3g[zone == zones3g]) if not ZRnode(n)]
            zr3g = [i for i,n in enumerate(nodes3g[zone == zones3g]) if ZRnode(n)]
            zn3g = np.where(zone == zones3g)[0][zn3g]
            zr3g = np.where(zone == zones3g)[0][zr3g]
            zn4g = [i for i,n in enumerate(nodes4g[zone == zones4g]) if not ZRnode(n)]
            zr4g = [i for i,n in enumerate(nodes4g[zone == zones4g]) if ZRnode(n)]
            zn4g = np.where(zone == zones4g)[0][zn4g]
            zr4g = np.where(zone == zones4g)[0][zr4g]
            zn5g = [i for i,n in enumerate(nodes5g[zone == zones5g]) if not ZRnode(n)]
            zr5g = [i for i,n in enumerate(nodes5g[zone == zones5g]) if ZRnode(n)]
            zn5g = np.where(zone == zones5g)[0][zn5g]
            zr5g = np.where(zone == zones5g)[0][zr5g]

            #ZN
            gtrxdev = script_pot_2g(cellname2g[zn2g],controller2g[zn2g])
            ucell,ulocell,upcpich,ulocellcontr = script_pot_3g(nodes3g[zn3g],controller3g[zn3g],cellid3g[zn3g])
            cell,pdschcfg,celldlpcpdschpa,cellchpwrcfg = script_pot_4g(np.unique(nodes4g[zn4g]))
            nrducelltrp,nrcell,nrducell = script_pot_5g(np.unique(nodes5g[zn5g]))
            #save ZN
            scriptZN[iz]+=gtrxdev+ucell+ulocell+ulocellcontr+upcpich+cell+pdschcfg+celldlpcpdschpa+cellchpwrcfg+nrducelltrp+nrcell+nrducell
            savetxt(gtrxdev+ucell+ulocell+ulocellcontr+upcpich+cell+pdschcfg+celldlpcpdschpa+cellchpwrcfg+nrducelltrp+nrcell+nrducell,fname+'\\'+cname+'\\checkpower_'+cname+'_ZN_'+zone)
            #ZR
            gtrxdev = script_pot_2g(cellname2g[zr2g],controller2g[zr2g])
            ucell,ulocell,upcpich,ulocellcontr = script_pot_3g(nodes3g[zr3g],controller3g[zr3g],cellid3g[zr3g])
            cell,pdschcfg,celldlpcpdschpa,cellchpwrcfg = script_pot_4g(nodes4g[zr4g])
            nrducelltrp,nrcell,nrducell = script_pot_5g(nodes5g[zr5g])
            #save ZR
            scriptZR[iz]+=gtrxdev+ucell+ulocell+ulocellcontr+upcpich+cell+pdschcfg+celldlpcpdschpa+cellchpwrcfg+nrducelltrp+nrcell+nrducell
            savetxt(gtrxdev+ucell+ulocell+ulocellcontr+upcpich+cell+pdschcfg+celldlpcpdschpa+cellchpwrcfg+nrducelltrp+nrcell+nrducell,fname+'\\'+cname+'\\checkpower_'+cname+'_ZR_'+zone)
    return scriptZN,scriptZR

def Query_powers():
    if not 'clusters' in globals():
        global clusters
        clusters = extrae_info_clusters(clustersinput)
    scriptZN = ['','','','',''] 
    scriptZR = ['','','','',''] 
    for i,c in enumerate(clusters):
        scriptZN,scriptZR = query_power_ct(c.dfct,c.name,scriptZN,scriptZR)
    fname = 'Response queries powers'
    create_folder(fname)
    zones = ['Zona ESTE','Zona SUR','Zona CENTRO','Zona NORTE','Zona 5']
    for iz,zone in enumerate(zones): 
        scriptZN[iz] = '\n'.join(np.unique(scriptZN[iz].split("\n")))
        if len(scriptZN[iz]) > 2:
            scriptZN[iz] = scriptZN[iz][1::]+'\n'
        scriptZR[iz] = '\n'.join(np.unique(scriptZR[iz].split("\n")))
        if len(scriptZR[iz]) > 2:
            scriptZR[iz] = scriptZR[iz][1::]+'\n'        
        savetxt(scriptZN[iz],fname+'\\checkpowerALL_ZN_'+zone)
        savetxt(scriptZR[iz],fname+'\\checkpowerALL_ZR_'+zone)
    dfclsit.to_excel(fname+'\\clusters_sites.xlsx')
    messagebox.showinfo(message="Query finished", title="Query powers")

def get_CELLtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST CELL:'
    if okpowersaving.get():
        column_names = ['Local Cell ID','Cell Name','Frequency band','Uplink EARFCN indication','Uplink EARFCN','Downlink EARFCN','Uplink bandwidth',
                            'Downlink bandwidth','Cell ID','Physical cell ID','Ue max power allowed configure indicator',
                            'Max transmit power allowed(dBm)','Cell transmission and reception mode']
    else:
        column_names = ['Local Cell ID','Cell Name','Csg indicator','Uplink cyclic prefix length','Downlink cyclic prefix length',
                        'NB-IoT Cell Flag','Coverage Level Type ','Frequency band','Uplink EARFCN indication','Uplink EARFCN','Downlink EARFCN','Uplink bandwidth',
                        'Downlink bandwidth','Cell ID','Physical cell ID','Additional spectrum emission','Cell active state','Cell admin state',
                        'Cell Blocking Duration(min)','Cell FDD TDD indication','Subframe assignment','Special subframe patterns','SSP6 DwPTS Mode','Cell Standby Mode',
                        ' Cell specific offset(dB)','Frequency offset(dB)','Root sequence index','High speed flag','Preamble format','Cell radius(m)',
                        'Customized bandwidth configure indicator','Customized uplink bandwidth(0.1MHz)','Customized downlink bandwidth(0.1MHz)','Ue max power allowed configure indicator',
                        'Max transmit power allowed(dBm)','Flag of Multi-RRU Cell','Mode of Multi-RRU Cell','CPRI Ethernet Compression Ratio','CPRI Compression',
                        'Physical Cell Number of SFN Cell','CRS Port Number','Cell transmission and reception mode','CRS Antenna Port Mapping','User label','Work mode',
                        'CN Operator Sharing Group ID','Intra Frequency RAN Sharing Indication','IntraFreq ANR Indication','Cell Radius Start Location(m)','Specified Cell Flag',
                        'Downlink Punctured RB Number','SFN Master Cell Label','Multi Cell Share Mode','Standby Cell SFN Recovery Time(h)','Compact Bandwidth Control Interference Mode',
                        'Uplink Punctured RB Number Offset','Ultra High-Speed Cell Root Sequence Index']
    global dfcell
    dfcell = process_txt_verthorz(command,column_names)
    return dfcell

def get_NRCELLtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST NRCELL:'
    column_names = ['NR Cell ID','Cell Name','Cell ID','Frequency Band','Duplex Mode','User Label','Cell Activate State']
    global dfnrcell
    dfnrcell = process_txt_verthorz(command,column_names)
    return dfnrcell

def get_NRDUCELLtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST NRDUCELL:'
    if okpowersaving.get():
         column_names = ['NR DU Cell ID','NR DU Cell Name','Duplex Mode','Cell ID','Physical Cell ID','Frequency Band','Uplink NARFCN',
                    'Downlink NARFCN','Uplink Bandwidth','Downlink Bandwidth','Cell Radius(m)','Subcarrier Spacing(KHz)','Cyclic Prefix Length','SSB Subcarrier Spacing(KHz)']
    else:
        column_names = ['NR DU Cell ID','NR DU Cell Name','Duplex Mode','Cell ID','Physical Cell ID','Frequency Band','Uplink NARFCN',
                        'Downlink NARFCN','Uplink Bandwidth','Downlink Bandwidth','Cell Radius(m)','Subcarrier Spacing(KHz)','Cyclic Prefix Length',
                        'Slot Assignment','Slot Structure','RAN Notification Area ID','LampSite Cell Flag','Tracking Area ID','TA Offset','Cell Administration State'
                        ,'SSB Frequency Position Describe Method','SSB Frequency Position','SSB Period(ms)','SIB1 Period(ms)','NR DU Cell Networking Mode',
                        'SIB Config ID','Logical Root Sequence Index','PRACH Frequency Start Position','High Speed Flag','Supplementary Cell Indicator','SMTC Duration(ms)',
                        'SMTC Period(ms)','Additional Frequency Band','SSB Subcarrier Spacing(KHz)','Customized Bandwidth Config Indicator','Customized Downlink Bandwidth',
                        'Customized Uplink Bandwidth','Uplink Punctured RB Number Offset','Peer Cell Slot Assignment','SSB Time Position','Cell Barred','Deployment Position',
                        'IAB Node ID']
    global dfnrducell
    dfnrducell = process_txt_verthorz(command,column_names)
    return dfnrducell

def get_CELLCHPWRCFGtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST CELLCHPWRCFG:'
    if okpowersaving.get():
        column_names = ['Local cell ID','Antenna Output Power(W)']
    else:
        column_names = ['Local cell ID','PCFICH power(0.005dB)','PBCH power(0.005dB)','SCH power(0.005dB)','DBCH power(0.005dB)','PCH power(0.005dB)',
                        'Rach response power(0.005dB)','Prs power(0.005dB)','Antenna Output Power(W)','PMCH Power Offset(0.1dB)','Handover RAR Power Enhanced Switch',
                        'TCCH Power Offset(0.005dB)','TTCH Power Offset(0.005dB)','TPCCH Power Offset(0.005dB)','Output Power Rate(per mill)',
                        'PDCCH Power Increase Offset(0.1dB)','Power Check Switch','Power Assistance Switch']
    global dfcellchpwrcfg
    dfcellchpwrcfg = process_txt_verthorz(command,column_names)
    return dfcellchpwrcfg

def get_UCELLtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST UCELL:'
    if okpowersaving.get():
        column_names = ['NE','Cell ID','Cell Name','Logical RNC ID','Max Transmit Power of Cell','Band Indicator',
                    'Uplink UARFCN','Downlink UARFCN','NodeB Name','Local Cell ID','Location Area Code',
                    'Service Area Code','RAC Configuration Indication','Routing Area Code']
    else:
        column_names = ['NE','Cell ID','Cell Name','Logical RNC ID','Max Transmit Power of Cell','Band Indicator',
                        'Cn Operator Group Index','UL Frequency Ind','Uplink UARFCN','Downlink UARFCN','Time Offset',
                        'Num of Continuous in Sync Ind','Num of Continuous Out of Sync Ind','Radio Link Failure Timer Length',
                        'DL Power Control Mode 1','DL Primary Scrambling Code','TX Diversity Indication',
                        'Service Priority Group Identity','NodeB Name','Local Cell ID','Location Area Code',
                        'Service Area Code','RAC Configuration Indication','Routing Area Code','STTD Support Indicator',
                        'CP1 Support Indicator','Closed Loop Time Adjust Mode','DPCH Tx Diversity Mode for Other User',
                        'FDPCH Tx Diversity Mode for Other User','DPCH Tx Diversity Mode for MIMO User',
                        'FDPCH Tx Diversity Mode for MIMO User','Tx Diversity Mode for DC-HSDPA User','Cell Oriented Cell Individual Offset',
                        'Cell VP Limit Indicator','DSS Cell Flag','Maximum TX Power in Small DSS Coverage',
                        'Common Channel Bandwidth Operator Index','Hierarchy ID of Terminal Type','Subrack No.','Subrack name',
                        'Slot No.','Subsystem No.','Heterogeneous Cell Flag','Self Planning Flag','Remark','Split Cell Indicator',
                        'Cell Coverage Type','Network Layer ID For Flexible UE Group','Pole Site Identification','HostType',
                        'Validation indication','Cell administrative state','Cell MBMS state','Cell MIMO state','IPDL flag',
                        'Cell CBS state','Cell ERACH state']
    global dfucell
    dfucell = process_txt(command,column_names)
    return dfucell

def get_ULOCELLtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST ULOCELL:'
    column_names = ['Local Cell ID','Local Cell Type','Cell Scale Indication','Cell ID','Cell Name','UL Baseband Equipment ID',
                    'DL Baseband Equipment ID','Local Cell Radius(m)','Local Cell Inner Handover Radius(m)','Two Tx Way','Reserved Cell','UL Frequency Indicator']
    global dfulocell
    dfulocell = process_txt(command,column_names)
    return dfulocell

def get_ULOCELLSECTOREQMtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST ULOCELLSECTOREQM:'
    column_names = ['Local Cell ID','Sector Equipment ID','Max Output Power(0.1dBm)','Sector Equipment Property']
    global dfulocellsectoreqm
    dfulocellsectoreqm = process_txt_verthorz(command,column_names)
    return dfulocellsectoreqm

def get_UPCPICHtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST UPCPICH:'
    column_names = ['Cell ID','Cell Name','PCPICH ID','PCPICH Transmit Power','Max Transmit Power of PCPICH',
                    'Min Transmit Power of PCPICH','PCPICH TX Power in Small DSS Coverage']
    global dfupcpich
    dfupcpich = process_txt(command,column_names)
    return dfupcpich

def get_PDSCHCFGtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST PDSCHCFG:'
    column_names = ['Local cell ID','Reference signal power(0.1dBm)','PB','Reference Signal Power Margin(0.1dB)','Offset of Ant0 to Tx Power(0.1dB)',
                    'Offset of Ant1 to Tx Power(0.1dB)','Offset of Ant2 to Tx Power(0.1dB)','Offset of Ant3 to Tx Power(0.1dB)','TX Channel Power Config Switch',
                    'Cell Power Limit(0.01W)','EMF Power Limit Switch','CRS Power Boosting Amplitude','Logical Port Swap Switch','PDSCH Power Boosting Switch',
                    'CRS Power Reduction Amount']
    global dfpdschcfg
    dfpdschcfg = process_txt_verthorz(command,column_names)
    return dfpdschcfg

def get_CELLDLPCPDSCHPAtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST CELLDLPCPDSCHPA:'
    column_names = ['Local cell ID','UE-Level NPREO Adjustment Opt Switch','PA for even power distribution(dB)','Nominal PDSCH-to-RS-EPRE Offset(dB)',
                    'EPRE Offset Adjustment Threshold','TM9 Rank Correct Threshold','Cell-Level NPREO Adjustment Threshold','UE-Level NPREO Adjustment Threshold',
                    'UE-Level NPREO Adjustment Upper Limit','TM3 TM4 Rank Correct Threshold','UE NPREO Adjust Switch']
    global dfcelldlpcpdschpa
    dfcelldlpcpdschpa = process_txt_verthorz(command,column_names)
    return dfcelldlpcpdschpa

def get_NRDUCELLTRPtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    global text
    textan = text
    command = 'MML Command-----LST NRDUCELLTRP:'
    column_names = ['NR DU Cell TRP ID','NR DU Cell ID','Transmit and Receive Mode','Baseband Equipment ID','Power Config Mode','Max Transmit Power(0.1dBm)',
                    'Max EIRP(0.1dBm)','CPRI Compression',' Baseband Resource Mutual Aid Switch','Branch CPRI Compression','Antenna Polarization','TRP Type',
                    'Master TRP ID','Frequency Range and Duplex Mode','UL RX Power Attenuation Value(dB)','Secondary Baseband Equipment ID']
    global dfnrducelltrp
    dfnrducelltrp = process_txt_verthorz(command,column_names)
    return dfnrducelltrp

def get_GTRXDEVtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST GTRXDEV:'
    if okpowersaving.get():
        column_names = ['TRX ID','TRX Name','Power Level','GBTS Power Type(w)','eGBTS Power Type(0.1dBm)','TRX Priority','Cell Index']
    else:
        column_names = ['TRX ID','TRX Name','Power Level','GBTS Power Type(w)','eGBTS Power Type(0.1dBm)','TRX Priority','Shut Down Enabled','TCH Rate Adjust Allow',
                        'TRX 8PSK Level','Power Finetune','Receive Mode','Send Mode','Allow Dynamic Shutdown of TRX','Power Unit','Timeslot Power Reserve',
                        'Timeslot Power Reserve(0.1dBm)','Power Overload Threshold In Ho. Out','Power Overload Threshold In Ho. Out(0.1dBm)',
                        'Power Overload Threshold In Ho. In','Power Overload Threshold In Ho. In(0.1dBm)','Frequency Reuse Mode','Priority of Shut Down TRX',
                        '16QAM Transmitter Power Reduce Level','32QAM Transmitter Power Reduce Level','Spectrum Sharing Shutdown Allowed','Target CIR for IBCA MAIO Configured on TRX',
                        'TRX Logical Lock Switch','DL Precise PC Power Increase Level','Cell Index']
        global dfgtrxdev
    dfgtrxdev = process_txt(command,column_names)
    return dfgtrxdev

def get_ULOCELLCONTRtext():
    if not 'text' in globals():
        messagebox.showwarning(message="Introduzca txt file", title="Warning")
        return
    command = 'MML Command-----LST ULOCELL:IDTYPE=BYNAME,NODEBNAME='
    column_names = ['NodeB ID','Local Cell ID','NE','NODEBNAME']
    global text
    textan = text
    global dfulocellcontr
    dfulocellcontr = process_txt_verthorz(command,column_names)
    return dfulocellcontr


def get_techsec_from_cellnames(cellnames):
    techs = []
    for c in cellnames:
        if not c[-1].isdigit():
            techs.append(c[-3:-1])
        else:
            techs.append(c[-2::])
    return techs

def papcofftranslate(x):
    if (str == type(x))and(str.isnumeric(x)):
        x = int(x)
    #dict papcoff
    papcoffdict = {6:0,4:1,1:3,0:4,1:5,2:6,3:7}
    if x in papcoffdict:
        y = papcoffdict[x]
    else:
        y = -1
    return y
def script_modpower_3g(df3gexp,dfoptch):
    ipcpich = df3gexp[df3gexp['pcpichpower esperado'] != 'OK'].index
    imaxtxpower = df3gexp[df3gexp['maxtxpower esperado'] != 'OK'].index
    modpow3g = ''
    for el in imaxtxpower:
        #OPTCH
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df3gexp.loc[el,'cellnamex']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'MAXTXPOWER'
        dfoptch.loc[posopch,'AUX INI'] = df3gexp.loc[el,'maxtxpower']
        dfoptch.loc[posopch,'AUX FIN'] = df3gexp.loc[el,'maxtxpower esperado']
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df3gexp.loc[el,'cellnamex']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'MAXPOWER'
        dfoptch.loc[posopch,'AUX INI'] = df3gexp.loc[el,'maxtxpower']
        dfoptch.loc[posopch,'AUX FIN'] = df3gexp.loc[el,'maxtxpower esperado']
        #Script
        maxtxpowerel = str(df3gexp.loc[el,'maxtxpower esperado'])
        cellidel = str(df3gexp.loc[el,'cellid'])
        nodel = df3gexp.loc[el,'nodo']
        neid = df3gexp.loc[el,'neid']
        #modpow3g+='LST ALMAF:;{'+nodel+'}\n'
        #modpow3g+='LST UCELL:NODEBNAME="'+nodel+'";{'+neid+'}\n'
        modpow3g+='\nMAXTXPOWER '+df3gexp.loc[el,'cellnamex']+'\n'
        #if df3gexp.loc[el,'maxtxpw nodo'] < df3gexp.loc[el,'maxtxpower']:
        modpow3g+='MOD ULOCELL:ULOCELLID='+cellidel+',LOCELLTYPE=NORMAL_CELL,MAXPWR='+maxtxpowerel+';{'+nodel+'}\n'
        modpow3g+='MOD UCELL:CELLID='+cellidel+',MAXTXPOWER='+maxtxpowerel+';{'+neid+'}\n'
        #modpow3g+='LST UCELL:NODEBNAME="'+nodel+'";{'+neid+'}\n'
        #modpow3g+='LST ALMAF:;{'+nodel+'}\n'
    for el in ipcpich:
        #OPTCH
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df3gexp.loc[el,'cellnamex']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'PCPICHPOWER'
        dfoptch.loc[posopch,'AUX INI'] = df3gexp.loc[el,'pcpichpower']
        dfoptch.loc[posopch,'AUX FIN'] = df3gexp.loc[el,'pcpichpower esperado']
        #Script
        pcpichel = str(df3gexp.loc[el,'pcpichpower esperado'])
        cellidel = str(df3gexp.loc[el,'cellid'])
        nodel = df3gexp.loc[el,'nodo']
        neid = df3gexp.loc[el,'neid']
        #modpow3g+='LST ALMAF:;{'+nodel+'}\n'
        #modpow3g+='LST UPCPICH:CELLID='+cellidel+';{'+neid+'}\n'
        modpow3g+='\nCPICH '+df3gexp.loc[el,'cellnamex']+'\n'
        if df3gexp.loc[el,'pcpichpower esperado'] > df3gexp.loc[el,'maxpcpichpower']:
            modpow3g+='MOD UPCPICHPWR:CELLID='+cellidel+',MAXPCPICHPOWER='+pcpichel+';{'+neid+'}\n'
            posopch = len(dfoptch)
            dfoptch.loc[posopch,'ELEMENTO'] = df3gexp.loc[el,'cellnamex']
            dfoptch.loc[posopch,'AUX NOMBRE'] = 'MAXPCPICHPOWER'
            dfoptch.loc[posopch,'AUX INI'] = df3gexp.loc[el,'maxpcpichpower']
            dfoptch.loc[posopch,'AUX FIN'] = df3gexp.loc[el,'pcpichpower esperado']
        modpow3g+='MOD UCELL:CELLID='+cellidel+',PCPICHPOWER='+pcpichel+';{'+neid+'}\n'
        #modpow3g+='LST UPCPICH:CELLID='+cellidel+';{'+neid+'}\n'
        #modpow3g+='LST ALMAF:;{'+nodel+'}\n'
    return modpow3g,dfoptch
def checkpower3g(df3gexp,sitess,dfoptch):
    df3gexp['pcpichpower esperado'] = 'OK'
    df3gexp.loc[:,'maxtxpower esperado'] = 'OK'
    df3gexp.loc[:,'Check Diseño'] = 'OK'
    pcpich = df3gexp.pcpichpower.tolist()
    maxtxpower = df3gexp['maxtxpower'].tolist()
    maxpcpich = df3gexp['maxpcpichpower'].tolist()
    band = [x[-3] for x in df3gexp['cellnamex'].tolist()]
    for el,pcpichel in enumerate(pcpich):
        if (type(pcpichel) == int) & (type(maxpcpich[el]) == int)&(type(maxtxpower[el]) == int):
            if band[el] == 'F':
                if pcpichel != 360:
                    if df3gexp.loc[el,'site'] in sitess:
                        df3gexp.loc[el,'Comentarios'] = 'RRU no soporta más potencia.'
                        df3gexp.loc[el,'pcpichpower esperado'] = 360
                    df3gexp.loc[el,'Check Diseño'] = 'EXCEPCION'
                if maxtxpower[el] != 460:
                    df3gexp.loc[el,'Check Diseño'] = 'EXCEPCION'
                    if df3gexp.loc[el,'site'] in sitess:
                        df3gexp.loc[el,'Comentarios'] = 'RRU no soporta más potencia.' 
                        df3gexp.loc[el,'maxtxpower esperado'] = 460  
            else:
                if maxtxpower[el]-pcpichel > 100:
                    df3gexp.loc[el,'pcpichpower esperado'] = maxtxpower[el]-100
                if maxtxpower[el]-pcpichel < 100:   
                    df3gexp.loc[el,'maxtxpower esperado'] = pcpichel+100
    modpow3g,dfoptch = script_modpower_3g(df3gexp,dfoptch)

    return df3gexp,modpow3g,dfoptch
def script_modpower_4g(df4gexp,dfoptch):
    keys = df4gexp.keys()
    if 'pb esperado' in keys:
        ipb = df4gexp[df4gexp['pb esperado'] != 'OK'].index
    else:
        ipb = []
    if 'papcoff esperado' in keys:
        ipapcoff = df4gexp[df4gexp['papcoff esperado'] != 'OK'].index
    else:
        ipapcoff = []
    irefpow = df4gexp[df4gexp.loc[:,'referencesignalpwr esperado']!='OK'].index
    modpow4g = ''
    for el in ipb:
        cellidel = str(df4gexp.loc[el,'localcellid'])
        nodel = df4gexp.loc[el,'nodo']
        pbel = str(df4gexp.loc[el,'pb esperado'])
        #OPTCH
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df4gexp.loc[el,'cellname']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'PB'
        dfoptch.loc[posopch,'AUX INI'] = df4gexp.loc[el,'pb']
        dfoptch.loc[posopch,'AUX FIN'] = df4gexp.loc[el,'pb esperado']
        #Script
        modpow4g+='\nPB '+df4gexp.loc[el,'cellname']+'\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
        #modpow4g+='LST PDSCHCFG:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        modpow4g+='MOD PDSCHCFG:LOCALCELLID='+cellidel+',PB='+pbel+';{'+nodel+'}\n'
        #modpow4g+='LST PDSCHCFG:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
    for el in ipapcoff:
        #OPTCH
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df4gexp.loc[el,'cellname']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'PAPCOFF'
        dfoptch.loc[posopch,'AUX INI'] = df4gexp.loc[el,'papcoff']
        dfoptch.loc[posopch,'AUX FIN'] = df4gexp.loc[el,'papcoff esperado']
        #Script
        cellidel = str(df4gexp.loc[el,'localcellid'])
        nodel = df4gexp.loc[el,'nodo']
        modpow4g+='\nPA '+df4gexp.loc[el,'cellname']+'\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
        #modpow4g+='LST CELLDLPCPDSCHPA:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        modpow4g+='MOD CELLDLPCPDSCHPA:LOCALCELLID='+cellidel+',PAPCOFF=DB0_P_A;{'+nodel+'}\n'
        #modpow4g+='LST CELLDLPCPDSCHPA:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
    for el in irefpow:
        #OPTCH
        posopch = len(dfoptch)
        dfoptch.loc[posopch,'ELEMENTO'] = df4gexp.loc[el,'cellname']
        dfoptch.loc[posopch,'AUX NOMBRE'] = 'REFERENCESIGNALPWR'
        dfoptch.loc[posopch,'AUX INI'] = df4gexp.loc[el,'referencesignalpwr']
        dfoptch.loc[posopch,'AUX FIN'] = df4gexp.loc[el,'referencesignalpwr esperado']
        #Script
        cellidel = str(df4gexp.loc[el,'localcellid'])
        nodel = df4gexp.loc[el,'nodo']
        refpowel = str(df4gexp.loc[el,'referencesignalpwr esperado'])
        modpow4g+='\nPDSCHCFG '+df4gexp.loc[el,'cellname']+'\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
        #modpow4g+='LST PDSCHCFG:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        modpow4g+='MOD PDSCHCFG:LOCALCELLID='+cellidel+',REFERENCESIGNALPWR='+refpowel+';{'+nodel+'}\n'
        #modpow4g+='LST PDSCHCFG:LOCALCELLID='+cellidel+';{'+nodel+'}\n'
        #modpow4g+='LST ALMAF:;{'+nodel+'}\n'
    return modpow4g,dfoptch
def checkpower4g(df4gexp,sitess,dfoptch):
    if any(df4gexp.pb != 0):
        pos = int(np.where(df4gexp.keys()=='pb')[0][0]+1)
        df4gexp.insert(pos,"pb esperado",df4gexp.pb.tolist(),True)
        df4gexp.loc[df4gexp.pb == 0,'pb esperado'] = 'OK'
        df4gexp.loc[df4gexp.pb != 0,'pb esperado'] = 0
    papcofftrans = [papcofftranslate(x) for x in df4gexp.papcoff.tolist()]
    df4gexp.papcoff = papcofftrans
    if any(df4gexp.papcoff != 4):
        pos = int(np.where(df4gexp.keys()=='papcoff')[0][0]+1)
        df4gexp.insert(pos,"papcoff esperado",df4gexp.papcoff.tolist(),True)
        df4gexp.loc[df4gexp.pb == 4,'papcoff esperado'] = 'OK'
        df4gexp.loc[df4gexp.pb != 4,'papcoff esperado'] = 4
    df4gexp.loc[:,'referencesignalpwr esperado']='OK'
    df4gexp.loc[:,'Check Diseño'] = 'OK'
    #BAND Y
    limbandy = 152
    baselinebandy = 152
    df4gexp.loc[(df4gexp.BANDA == 'Y') & (df4gexp.referencesignalpwr != limbandy),'referencesignalpwr esperado'] = baselinebandy
    df4gexp.loc[(df4gexp.BANDA == 'Y') & (df4gexp.referencesignalpwr != limbandy),'Check Diseño'] = 'EXCEPCION'
    bwy = df4gexp.loc[(df4gexp.BANDA == 'Y') ,'dlbandwidth'].tolist()
    for i,bwi in enumerate(bwy):
        if bwi != '10M':
            bwy[i] = '10M'
    df4gexp.loc[(df4gexp.BANDA == 'Y') ,'dlbandwidth'] = bwy
    #BAND M
    limbandm = 147
    baselinebandm = 149
    df4gexp.loc[(df4gexp.BANDA == 'M') & (df4gexp.referencesignalpwr < limbandm),'referencesignalpwr esperado'] = baselinebandm
    df4gexp.loc[(df4gexp.BANDA == 'M') & (df4gexp.referencesignalpwr != baselinebandm),'Check Diseño'] = 'EXCEPCION'
    #Other bands
    limmin = 122
    baselinemin = 122
    df4gexp.loc[(df4gexp.BANDA != 'M') & (df4gexp.BANDA != 'Y') & (df4gexp.referencesignalpwr < limmin),'referencesignalpwr esperado'] = baselinemin
    df4gexp.loc[(df4gexp.BANDA != 'M') & (df4gexp.BANDA != 'Y') & (df4gexp.referencesignalpwr < limmin),'Check Diseño'] = 'EXCEPCION'
    #check ZR
    df4gexp.loc[(df4gexp.ZONA == 'ZR')&((df4gexp.BANDA == 'N') | (df4gexp.BANDA == 'L')) & (df4gexp.referencesignalpwr != limmin),'referencesignalpwr esperado'] = baselinemin
    #Check Excepcion
    for el in range(len(df4gexp)):
        if (df4gexp.loc[el,'Check Diseño'] == 'EXCEPCION') and (df4gexp.site[el] in sitess):
            df4gexp.loc[el,'Comentarios'] = 'RRU no soporta más potencia.'
    #ADD 4G script changes
    modpow4g,dfoptch = script_modpower_4g(df4gexp,dfoptch)
    return df4gexp,modpow4g,dfoptch

def process_power_ct(ct,cname,sitess,dfgtrxdev,dfucell,dfulocell,dfulocellsectoreqm,dfulocellcontr,dfupcpich,dfcell,dfpdschcfg,dfcelldlpcpdschpa,dfcellchpwrcfg,dfnrducelltrp,dfnrcell,dfnrducell):
    fname = 'Response queries powers'
    create_folder(fname)
    create_folder(fname+'\\'+cname)
    ct = ct.loc[ct.STATUS == 1 & ((ct['OPERATOR NAME'] == 'ORANGE')|(ct['OPERATOR NAME'] == 'Orange')),:]
    #Optimization changes
    colnamesopch = ['ELEMENTO','AUX NOMBRE','AUX INI','AUX FIN']
    dfoptch = pd.DataFrame(columns=colnamesopch)
    #-------------------------------2G------------------------------------
    sites2g = ct.loc[ct.loc[:,'TECH'] == '2G','SITE'].tolist()
    nodes2g = ct.loc[ct.loc[:,'TECH'] == '2G','NODE'].tolist()
    cellnamex2g = ct.loc[ct.loc[:,'TECH'] == '2G','CELLNAMEX'].tolist()
    cellname2g = ct.loc[ct.loc[:,'TECH'] == '2G','CELLNAME'].tolist()
    controller2g = ct.loc[ct.loc[:,'TECH'] == '2G','CONTROLLER'].tolist()
    cellid2g = ct.loc[ct.loc[:,'TECH'] == '2G','CELLID'].tolist()

    coln2g = ['site','nodo','cellnamex','cellname','neid','cellid','trxname','trxid','powtunit','egbtspowt','powl','egbtspowt esperado']
    df2gexp = pd.DataFrame(columns=coln2g)

    df2gexp.site = sites2g
    df2gexp.nodo = nodes2g
    df2gexp.cellnamex = cellnamex2g
    df2gexp.cellname = cellname2g
    df2gexp.neid = controller2g
    df2gexp.cellid = cellid2g
    if len(dfgtrxdev) > 0:
        for el,cnel in enumerate(cellname2g):
            cellnames2gdf = dfgtrxdev['TRX Name'].tolist()
            L = len(cnel)
            for eldf,cndf in enumerate(cellnames2gdf):
                if len(cndf)>L:
                    cellnames2gdf[eldf] = cndf[0:L]
            igtrxdev = (np.array(cellnames2gdf) == cnel)
            if any(igtrxdev):
                df2gexp.loc[el,'trxname'] = dfgtrxdev.loc[igtrxdev,'TRX Name'].tolist()[0]
                df2gexp.loc[el,'trxid'] = dfgtrxdev.loc[igtrxdev,'TRX ID'].tolist()[0]
                df2gexp.loc[el,'powtunit'] = dfgtrxdev.loc[igtrxdev,'POWT'].tolist()[0]
                df2gexp.loc[el,'egbtspowt'] = dfgtrxdev.loc[igtrxdev,'eGBTS Power Type(0.1dBm)'].tolist()[0]
                df2gexp.loc[el,'powl'] = dfgtrxdev.loc[igtrxdev,'Power Level'].tolist()[0]

        df2gexp.loc[:,'egbtspowt esperado'] = 'OK'
    

    #-------------------------------3G------------------------------------

    sites3g = ct.loc[ct.loc[:,'TECH'] == '3G','SITE'].tolist()
    nodes3g = ct.loc[ct.loc[:,'TECH'] == '3G','NODE'].tolist()
    cellnamex3g = ct.loc[ct.loc[:,'TECH'] == '3G','CELLNAMEX'].tolist()
    cellname3g = ct.loc[ct.loc[:,'TECH'] == '3G','CELLNAME'].tolist()
    controller3g = ct.loc[ct.loc[:,'TECH'] == '3G','CONTROLLER'].tolist()
    cellid3g = ct.loc[ct.loc[:,'TECH'] == '3G','CELLID'].tolist()
    
    coln3g = ['site','nodo','cellnamex','neid','cellname','cellid','maxpcpichpower','pcpichpower','pcpichpower esperado','maxtxpower','maxtxpower esperado','maxtxpw nodo','Check Diseño','Comentarios']
    df3gexp = pd.DataFrame(columns=coln3g)
    df3gexp.site = sites3g
    df3gexp.nodo = nodes3g
    df3gexp.cellnamex = cellnamex3g
    df3gexp.cellname = cellname3g
    df3gexp.cellid = cellid3g
    df3gexp.neid = controller3g
    modpow3g = ''
    if len(dfucell) > 0:
        for el,cidel in enumerate(cellid3g):
            iucell = (dfucell['Cell Name'] == cellname3g[el])|((dfucell['NE'] == controller3g[el])&(dfucell['Cell ID'] == cidel))
            iupcpich = (dfupcpich['Cell Name'] == cellname3g[el])|((dfupcpich['NE'] == controller3g[el])&(dfupcpich['Cell ID'] == cidel))
            if any(iucell):
                df3gexp.loc[el,'maxtxpower'] = dfucell.loc[iucell,'Max Transmit Power of Cell'].tolist()[0]
                cidelucell = dfucell.loc[iucell,'Cell ID'].tolist()[0]
                iulocellsectoreqm = (dfulocellsectoreqm['NE'] == nodes3g[el])&((dfulocellsectoreqm['Local Cell ID'] == str(cidelucell))|(dfulocellsectoreqm['Local Cell ID'] == int(cidelucell)))
            else:
                iulocellsectoreqm = iucell
            if any(iupcpich):
                df3gexp.loc[el,'pcpichpower'] = dfupcpich.loc[iupcpich,'PCPICH Transmit Power'].tolist()[0]
                df3gexp.loc[el,'maxpcpichpower'] = dfupcpich.loc[iupcpich,'Max Transmit Power of PCPICH'].tolist()[0]
            if any(iulocellsectoreqm):
                maxtxpwnodo = dfulocellsectoreqm.loc[iulocellsectoreqm,'Max Output Power(0.1dBm)'].tolist()[0]
                if (maxtxpwnodo == 65535) | (maxtxpwnodo == '65535'):
                    df3gexp.loc[el,'maxtxpw nodo'] = dfucell.loc[iucell,'Max Transmit Power of Cell'].tolist()[0]
                else:
                    df3gexp.loc[el,'maxtxpw nodo'] = maxtxpwnodo
            else:
                df3gexp.loc[el,'maxtxpw nodo'] = 0
        df3gexp,modpow3g,dfoptch = checkpower3g(df3gexp,sitess,dfoptch)
    #-------------------------------4G------------------------------------
   
    sites4g = ct.loc[ct.loc[:,'TECH'] == '4G','SITE'].tolist()
    nodes4g = ct.loc[ct.loc[:,'TECH'] == '4G','NODE'].tolist()
    cellname4g = ct.loc[ct.loc[:,'TECH'] == '4G','CELLNAME'].tolist()
    bandsec = np.array(get_techsec_from_cellnames(cellname4g))
    okzrs = np.array([ZRnode(x) for x in nodes4g])
    zone = np.array(['ZN' for x in range(len(okzrs))])
    zone[okzrs] = 'ZR'
    lcid = [localcellid(x,zone[el]) for el,x in enumerate(bandsec)]
    modpow4g = ''
    coln4g = ['site','nodo','neid','cellname','localcellid','dlbandwidth','antoutputpwr','referencesignalpwr','pb','papcoff','referencesignalpwr esperado','BANDA','ZONA','Check Diseño', 'CONFIGURACION']
    df4gexp = pd.DataFrame(columns=coln4g)
    df4gexp.site = sites4g
    df4gexp.nodo = nodes4g
    df4gexp.neid = nodes4g
    df4gexp.cellname = cellname4g
    df4gexp.localcellid = lcid
    df4gexp.ZONA = zone
    if len(dfpdschcfg) > 0:
        for el,lcidel in enumerate(lcid):
            df4gexp.loc[el,'BANDA'] = bandsec[el][0]
            ipdschcfg = (dfpdschcfg['Local cell ID'] == lcidel) & (dfpdschcfg['NE'] == nodes4g[el])
            if any(ipdschcfg):
                df4gexp.loc[el,'pb'] = dfpdschcfg.loc[ipdschcfg,'PB'].tolist()[0]
                df4gexp.loc[el,'referencesignalpwr'] = dfpdschcfg.loc[ipdschcfg,'Reference signal power(0.1dBm)'].tolist()[0]
            icell = (dfcell['Local Cell ID'] == lcidel) & (dfcell['NE'] == nodes4g[el])
            if any(icell):
                df4gexp.loc[el,'dlbandwidth'] = dfcell.loc[icell,'Downlink bandwidth'].tolist()[0]
                df4gexp.loc[el,'CONFIGURACION'] = dfcell.loc[icell,'Cell transmission and reception mode'].tolist()[0]        
            icelldlpcpdschpa = (dfcelldlpcpdschpa['Local cell ID'] == lcidel) & (dfcelldlpcpdschpa['NE'] == nodes4g[el])
            if any(icelldlpcpdschpa):
                df4gexp.loc[el,'papcoff'] = dfcelldlpcpdschpa.loc[icelldlpcpdschpa,'PA for even power distribution(dB)'].tolist()[0]  
            icellchpwrcfg = (dfcellchpwrcfg['Local cell ID'] == lcidel) & (dfcellchpwrcfg['NE'] == nodes4g[el])
            if any(icellchpwrcfg):
                df4gexp.loc[el,'antoutputpwr'] = dfcellchpwrcfg.loc[icellchpwrcfg,'Antenna Output Power(W)'].tolist()[0]
        df4gexp,modpow4g,dfoptch = checkpower4g(df4gexp,sitess,dfoptch)

    #-------------------------------5G------------------------------------

    sites5g = ct.loc[ct.loc[:,'TECH'] == '5G','SITE'].tolist()
    nodes5g = ct.loc[ct.loc[:,'TECH'] == '5G','NODE'].tolist()
    cellname5g = ct.loc[ct.loc[:,'TECH'] == '5G','CELLNAME'].tolist()

    coln5g = ['Site','Nodo','nrducellname','neid','nrducellid','dlbandwidth','txrxmode','maxtransmitpower','maxtransmitpower esperado','Check Diseño']
    df5gexp = pd.DataFrame(columns=coln5g)

    df5gexp.Site = sites5g
    df5gexp.Nodo = nodes5g
    df5gexp.nrducellname = cellname5g
    df5gexp.neid = nodes5g
    df5gexp['maxtransmitpower esperado'] = 'OK'
    df5gexp['Check Diseño'] = 'OK'
    if len(dfnrducelltrp) > 0:
        for el,cnel in enumerate(cellname5g):
            inrducell = dfnrducell['NR DU Cell Name'] == cnel
            if any(inrducell):
                nrducellid = dfnrducell.loc[inrducell,'NR DU Cell ID'].tolist()[0]
                df5gexp.loc[el,'nrducellid'] = nrducellid
                df5gexp.loc[el,'dlbandwidth'] = dfnrducell.loc[inrducell,'Downlink Bandwidth'].tolist()[0]
                if (type(nrducellid)==int)or((type(nrducellid) == str)and(str.isdigit(nrducellid))):
                    inrducelltrp = (((dfnrducelltrp['NR DU Cell ID']) == str(nrducellid))|((dfnrducelltrp['NR DU Cell ID']) == int(nrducellid))) & (dfnrducelltrp['NE'] == nodes5g[el])
                else:
                    inrducelltrp = [False]
                if any(inrducelltrp):
                    df5gexp.loc[el,'txrxmode'] = dfnrducelltrp.loc[inrducelltrp,'Transmit and Receive Mode'].tolist()[0]
                    df5gexp.loc[el,'maxtransmitpower'] = dfnrducelltrp.loc[inrducelltrp,'Max Transmit Power(0.1dBm)'].tolist()[0]

    fname = 'Response queries powers'
    create_folder(fname)
    create_folder(fname+'\\'+cname)
    with pd.ExcelWriter(fname+'\\'+cname+'\\PW_'+cname+'.xlsx') as writer:
        df2gexp.to_excel(writer,sheet_name='2G',index=False)
        df3gexp.to_excel(writer,sheet_name='3G',index=False)
        df4gexp.to_excel(writer,sheet_name='4G',index=False)
        df5gexp.to_excel(writer,sheet_name='5G',index=False)
    savetxt(modpow3g+modpow4g,fname+'\\'+cname+'\\changepower_'+cname)
    optchfname = 'OPTCH_PW_'
    optchsname = '121'
    if not os.path.exists(fname+'\\'+cname+'\\'+optchfname+".xlsx"):
        shutil.copy('TMP_POW\\'+optchfname+'.xlsx', fname+'\\'+cname+'\\'+optchfname+".xlsx")
        filename = fname+'\\'+cname+'\\'+optchfname+cname+".xlsx"
        os.rename(fname+'\\'+cname+'\\'+optchfname+".xlsx", filename)
        #Load file
        doctofill = openpyxl.load_workbook(filename)
        # Cluster cells sheet
        sheet = doctofill[optchsname]

        #=======FILL CONSTANTS===================
        posfill = np.array([2,4,6,7,10,11,18,19])
        now = datetime.datetime.now().strftime("%d/%m/%Y")
        valuesfill = ['QF-Huawei',now,now,'CERRADO','QF','Parametrización','PARAMETRIZACION',cname]
        for i in range(len(dfoptch)):
            for j,el in enumerate(posfill):
                sheet.cell(row=i+2, column=el, value=valuesfill[j])
        posfill = np.array([3,15,16,17])
        for i in range(len(dfoptch)):
            for j,el in enumerate(posfill):
                sheet.cell(row=i+2, column=el, value=dfoptch.iloc[i,j])

        doctofill.save(filename)


def pow2dbm(pow):
    dbm = 0
    if (type(pow) == int) | (type(pow) == float):
        dbm = int(np.round(20*np.log10(pow*1000)))
    return dbm

def dbm2pow(dbm):
    pow = 0
    if (type(dbm) == int) | (type(dbm) == float):
        pow = np.round((10**((dbm/10)/10))/100)/10
    return pow

def pow2powt(pow):
    if type(pow) == str:
        powt = -1
    else:
        powers = np.array([40,60,20,30,55,63,80,50,15,12,25,10,13.3,7.5,7,6,5.5,3.7,3.1,45,18,0.2,13,8.5,26,16,11,9.5,8,24,31,12.5,19,
        6.5,4.5,4,3,27,17,14,9,5,2.5,2,11.5,21,0.1,125,1.5,1,35,70,75,85,90,95,100,105,110,115,120,130,135,140,145,150,
        155,160])
        powt = np.argmin(np.abs(powers-pow))
    return powt

def powstr2powint(pow):
    for i in range(len(pow)):
        if pow[i][:-1].isdigit():
            pow[i] = int(pow[i][:-1])
    return pow

def powint2powstr(pow):
    for i in range(len(pow)):
        if (type(pow[i]) != str):
            pow[i] = str(int(pow[i]))+'W'
    return pow

def str2int(x):
    for i in range(len(x)):
        if x[i].isdigit():
            x[i] = int(x[i])
    return x

def read_process_powers(dffn):
    dfgtrxdev = pd.read_excel(dffn,sheet_name='gtrxdev')
    dfucell = pd.read_excel(dffn,sheet_name='ucell')
    dfulocell = pd.read_excel(dffn,sheet_name='ulocell')
    dfulocellsectoreqm = pd.read_excel(dffn,sheet_name='ulocellsectoreqm')
    dfulocellcontr = pd.read_excel(dffn,sheet_name='ulocellcontr')
    dfcell = pd.read_excel(dffn,sheet_name='cell')
    dfupcpich = pd.read_excel(dffn,sheet_name='upcpich')
    dfpdschcfg = pd.read_excel(dffn,sheet_name='pdschcfg')
    dfcellchpwrcfg = pd.read_excel(dffn,sheet_name='cellchpwrcfg')
    dfcelldlpcpdschpa = pd.read_excel(dffn,sheet_name='celldlpcpdschpa')
    dfnrducelltrp = pd.read_excel(dffn,sheet_name='nrducelltrp')
    dfnrcell = pd.read_excel(dffn,sheet_name='nrcell')
    dfnrducell = pd.read_excel(dffn,sheet_name='nrducell')
    return dfgtrxdev,dfucell,dfulocell,dfulocellsectoreqm,dfulocellcontr,dfcell,dfupcpich,dfpdschcfg,dfcellchpwrcfg,dfcelldlpcpdschpa,dfnrducelltrp,dfnrcell,dfnrducell
def Process_powers():
    if not 'clusters' in globals():
        global clusters
        clusters = extrae_info_clusters(clustersinput) 
    fname = 'Response queries powers'
    create_folder(fname)
    dffn = fname+'\\Powers_all.xlsx'
    if os.path.exists(dffn):
            dfgtrxdev,dfucell,dfulocell,dfulocellsectoreqm,dfulocellcontr,dfcell,dfupcpich,dfpdschcfg,dfcellchpwrcfg,dfcelldlpcpdschpa,dfnrducelltrp,dfnrcell,dfnrducell = read_process_powers(dffn)
    else:
        #----------2G--------------
        dfgtrxdev = get_GTRXDEVtext()
        dfgtrxdev['eGBTS Power Type(0.1dBm)'] = str2int(dfgtrxdev['eGBTS Power Type(0.1dBm)'].tolist())
        dfgtrxdev['Power Type(w)'] = powstr2powint(dfgtrxdev['GBTS Power Type(w)'].tolist())
        inoegnts = dfgtrxdev['eGBTS Power Type(0.1dBm)']=='<NULL>'
        inopowtype = dfgtrxdev['GBTS Power Type(w)']=='<NULL>'
        dfgtrxdev.loc[inoegnts,'eGBTS Power Type(0.1dBm)'] = [pow2dbm(x) for x in dfgtrxdev.loc[inoegnts,'Power Type(w)'].tolist()]
        dfgtrxdev.loc[inopowtype,'Power Type(w)'] = [dbm2pow(x) for x in dfgtrxdev.loc[inopowtype,'eGBTS Power Type(0.1dBm)'].tolist()]
        dfgtrxdev.loc[inopowtype,'GBTS Power Type(w)'] = powint2powstr(dfgtrxdev.loc[inopowtype,'Power Type(w)'].tolist())
        dfgtrxdev.loc[:,'POWT'] = [pow2powt(x) for x in dfgtrxdev.loc[:,'Power Type(w)'].tolist()]
        trxnamelist = dfgtrxdev['TRX Name'].tolist()
        for i,x in enumerate(trxnamelist):
            if x.find('_')!=-1:
                trxnamelist[i] = x[0:x.find('_')]
            if x.find('-')!=-1:
                trxnamelist[i] = x[0:x.find('-')]
            if x[-2].isdigit():
                trxnamelist[i] = x[0:-1]
        dfgtrxdev.loc[:,'CELLNAME'] = trxnamelist
        #----------3G--------------
        dfucell = get_UCELLtext()
        dfulocell = get_ULOCELLtext()
        dfulocellsectoreqm = get_ULOCELLSECTOREQMtext()
        dfulocellcontr = get_ULOCELLCONTRtext()
        dfupcpich = get_UPCPICHtext()
        #----------4G--------------
        dfcell = get_CELLtext()
        dfpdschcfg = get_PDSCHCFGtext()
        dfcelldlpcpdschpa = get_CELLDLPCPDSCHPAtext()
        dfcellchpwrcfg = get_CELLCHPWRCFGtext()
        #----------5G--------------
        dfnrducelltrp = get_NRDUCELLTRPtext()
        dfnrcell = get_NRCELLtext()
        dfnrducell = get_NRDUCELLtext()
        iok = [i for i,x in enumerate(dfnrcell['Cell Name'].tolist()) if type(x) == type('hello')]
        dfnrcell = dfnrcell.loc[iok,:]
        iok = [i for i,x in enumerate(dfnrducell['NR DU Cell Name'].tolist()) if type(x) == type('hello')]
        dfnrducell = dfnrducell.loc[iok,:]
        #---------generate df sheets------------
        with pd.ExcelWriter(fname+'\\'+'Powers_all'+'.xlsx') as writer:
            dfgtrxdev.to_excel(writer,sheet_name='gtrxdev',index=False)
            dfucell.to_excel(writer,sheet_name='ucell',index=False)
            dfulocell.to_excel(writer,sheet_name='ulocell',index=False)
            dfulocellsectoreqm.to_excel(writer,sheet_name='ulocellsectoreqm',index=False)
            dfulocellcontr.to_excel(writer,sheet_name='ulocellcontr',index=False)
            dfcell.to_excel(writer,sheet_name='cell',index=False)
            dfupcpich.to_excel(writer,sheet_name='upcpich',index=False)
            dfpdschcfg.to_excel(writer,sheet_name='pdschcfg',index=False)
            dfcellchpwrcfg.to_excel(writer,sheet_name='cellchpwrcfg',index=False)
            dfcelldlpcpdschpa.to_excel(writer,sheet_name='celldlpcpdschpa',index=False)
            dfnrducelltrp.to_excel(writer,sheet_name='nrducelltrp',index=False)
            dfnrcell.to_excel(writer,sheet_name='nrcell',index=False)
            dfnrducell.to_excel(writer,sheet_name='nrducell',index=False)
        dfgtrxdev,dfucell,dfulocell,dfulocellsectoreqm,dfulocellcontr,dfcell,dfupcpich,dfpdschcfg,dfcellchpwrcfg,dfcelldlpcpdschpa,dfnrducelltrp,dfnrcell,dfnrducell = read_process_powers(dffn)
    for i,c in enumerate(clusters):
        csitess = [x.name for x in c.sitess]
        if len(c.dfct)>0:
            process_power_ct(c.dfct,c.name,csitess,dfgtrxdev,dfucell,dfulocell,dfulocellsectoreqm,dfulocellcontr,dfupcpich,dfcell,dfpdschcfg,dfcelldlpcpdschpa,dfcellchpwrcfg,dfnrducelltrp,dfnrcell,dfnrducell)   
    dfclsit.to_excel(fname+'\\clusters_sites.xlsx')
    messagebox.showinfo(message="Process powers finished", title="Process powers")


#===============================================================


def create_window():
    nwd = tkinter.Toplevel(wd)
    nwd.title("Graphs")
    return nwd


def load_clusters():
    global clustersinput
    clustersinput = read_clusters_txt('Clusters.txt')
    if len(clustersinput) == 0:
        messagebox.showwarning(message="There is no clusters in Clusters.txt", title="Warning")
    else:
        messagebox.showinfo(message='Clusters loaded',title='Clusters.txt')
        global clusters
        clusters = extraeF_info_clusters(clustersinput)
    

global clustersinput
clustersinput = read_clusters_txt('Clusters.txt')
if len(clustersinput) == 0:
    messagebox.showwarning(message="There is no clusters in Clusters.txt", title="Warning")

#Vars

#clustersinput = ['5G_TDD_BAL1007','AX_451']

# Tkinter

wd = tkinter.Tk()
wd.title('Check PRB Cluster')
wd.geometry("900x700")
wd.config(background="White")

okPRBs = tkinter.IntVar()
okPRBs.set(0)

okHuella = tkinter.IntVar()
okHuella.set(0)

okHuellapost = tkinter.IntVar()
okHuellapost.set(0)

okiomtmp = tkinter.IntVar()
okiomtmp.set(0)

okqrets = tkinter.IntVar()
okqrets.set(0)

okexpret = tkinter.IntVar()
okexpret.set(0)

okqbal = tkinter.IntVar()
okqbal.set(0)

okqbalcosector = tkinter.IntVar()
okqbalcosector.set(0)

okqbalsaving = tkinter.IntVar()
okqbalsaving.set(1)

okpowersaving = tkinter.IntVar()
okpowersaving.set(1)

okgmr = tkinter.IntVar()
okgmr.set(0)

okpbal = tkinter.IntVar()
okpbal.set(0)

okcheckbal = tkinter.IntVar()
okcheckbal.set(0)
#Interface components

label_author = tkinter.Label(wd, 
    text="Developed by: José María Gómez Belmonte & Daniel López Alcalá")
label_title = tkinter.Label(wd,  
    text = "Huellers Tool", 
    width = 100, height = 4,  
    fg = "Black",
    font=("Arial", 20)) 

button_exit = tkinter.Button(wd,  
    text = "Exit", 
    command = exit) 

button_exportclusters = tkinter.Button(wd,  
    text = "Export clusters", 
    command = export_clusters)

button_loadretcsv = tkinter.Button(wd,  
    text = "Load RETs csv", 
    command = load_retcssv)

button_jointext = tkinter.Button(wd,  
    text = "Join text", 
    command = join_text)

button_loadbalcsv = tkinter.Button(wd,  
    text = "Load dfbal csv", 
    command = load_balcsv)

button_ExportRETs = tkinter.Button(wd,  
    text = "Export csv RETs", 
    command = export_csv_rets)

button_ExportBAL = tkinter.Button(wd,  
    text = "Export csv BAL", 
    command = process_bal)

button_lretheidis = tkinter.Button(wd,  
    text = "Load RETs HeidiSQL", 
    command = browsecsv_retheidis)

button_readftext = tkinter.Button(wd,  
    text = "Read text ATAE", 
    command = readtextfile)

button_relclusters = tkinter.Button(wd,  
    text = "Reload Clusters", 
    command = load_clusters)

button_cscoringprov = tkinter.Button(wd,  
    text = "Process Cell Scoring", 
    command = process_cellscoring_provincial)

label_datestart = tkinter.Label(wd, 
    text="Date start")
entry_datestart = tkinter.Entry(wd)
entry_datestart.insert(0,'2023-04-17')

label_dateend = tkinter.Label(wd, 
    text="Date end")
entry_dateend = tkinter.Entry(wd)
entry_dateend.insert(0,datetime.datetime.now().date().strftime("%Y-%m-%d"))

cbutton_PRBs = tkinter.Checkbutton(wd,
    text = "PRBs",
    variable = okPRBs,
    onvalue = 1,
    offvalue = 0)

cbutton_Huella = tkinter.Checkbutton(wd,
    text = "Footprint",
    variable = okHuella,
    onvalue = 1,
    offvalue = 0)
"""
cbutton_Huella_post = tkinter.Checkbutton(wd,
    text = "post",
    variable = okHuellapost,
    onvalue = 1,
    offvalue = 0)
"""
cbutton_iomtmp = tkinter.Checkbutton(wd,
    text = "IOM tmp",
    variable = okiomtmp,
    onvalue = 1,
    offvalue = 0)

cbutton_qrets = tkinter.Checkbutton(wd,
    text = "Query RETs",
    variable = okqrets,
    onvalue = 1,
    offvalue = 0)

cbutton_exprets = tkinter.Checkbutton(wd,
    text = "Export RETs",
    variable = okexpret,
    onvalue = 1,
    offvalue = 0)

cbutton_qbal = tkinter.Checkbutton(wd,
    text = "Query bal",
    variable = okqbal,
    onvalue = 1,
    offvalue = 0)

cbutton_qbalcosect = tkinter.Checkbutton(wd,
    text = "Query bal cosite",
    variable = okqbalcosector,
    onvalue = 1,
    offvalue = 0)

cbutton_pbalsaving = tkinter.Checkbutton(wd,
    text = "Save bal process time",
    variable = okqbalsaving,
    onvalue = 1,
    offvalue = 0)

cbutton_ppowsaving = tkinter.Checkbutton(wd,
    text = "Save power process time",
    variable = okpowersaving,
    onvalue = 1,
    offvalue = 0)

cbutton_pbal = tkinter.Checkbutton(wd,
    text = "Process bal",
    variable = okpbal,
    onvalue = 1,
    offvalue = 0)

cbutton_checkbal = tkinter.Checkbutton(wd,
    text = "Check bal",
    variable = okpbal,
    onvalue = 1,
    offvalue = 0)

cbutton_gmr = tkinter.Checkbutton(wd,
    text = "Generate MR",
    variable = okgmr,
    onvalue = 1,
    offvalue = 0)

button_qpowers = tkinter.Button(wd,  
    text = "Query Powers", 
    command = Query_powers)

button_ppowers = tkinter.Button(wd,  
    text = "Process Powers", 
    command = Process_powers)

button_cellscoronlyfootp = tkinter.Button(wd,  
    text = "Process Cell scoring footprint", 
    command = process_cell_scoring_only_footprint)



#Location of components

label_author.place(relx=0.25, rely=0.9, relwidth=0.5, relheight=0.1)
label_title.place(relx=0.25, rely=0, relwidth=0.5, relheight=0.1)
button_exit.place(relx=0.8, rely=0.9, relwidth=0.2, relheight=0.1)
button_relclusters.place(relx=0.8, rely=0.8, relwidth=0.2, relheight=0.1)
offright = 0.25
offup = 0.3
button_cscoringprov.place(relx=offright, rely=offup+0.3, relwidth=0.5/3, relheight=0.1)
button_cellscoronlyfootp.place(relx=offright, rely=offup+0.4, relwidth=0.5/3, relheight=0.05)
button_ExportRETs.place(relx=offright, rely=offup+0.2, relwidth=0.5/3, relheight=0.05)
button_exportclusters.place(relx=offright+0.5/3, rely=offup+0.5, relwidth=0.5/3, relheight=0.1)
button_qpowers.place(relx=offright+0.5/3, rely=offup+0.3, relwidth=0.5/3, relheight=0.05)
button_ppowers.place(relx=offright+0.5/3, rely=offup+0.35, relwidth=0.5/3, relheight=0.05)
button_ExportBAL.place(relx=offright+1/3, rely=offup+0.3, relwidth=0.5/3, relheight=0.1)
cbutton_pbalsaving.place(relx=offright+1/3, rely=offup+0.4, relwidth=0.5/3, relheight=0.05)
cbutton_ppowsaving.place(relx=offright+0.5/3, rely=offup+0.4, relwidth=0.5/3, relheight=0.05)
cbutton_checkbal.place(relx=offright+1/3, rely=offup+0.45, relwidth=0.5/3, relheight=0.05)
button_lretheidis.place(relx=offright, rely=offup, relwidth=0.5/3, relheight=0.05)
button_readftext.place(relx=offright, rely=offup+0.05, relwidth=0.5/3, relheight=0.05)
button_jointext.place(relx=offright, rely=offup+0.1, relwidth=0.5/3, relheight=0.05)
button_loadretcsv.place(relx=offright, rely=offup+0.15, relwidth=0.5/3, relheight=0.05)
#button_loadbalcsv.place(relx=offright, rely=offup+0.15, relwidth=0.5/3, relheight=0.05)
label_datestart.place(relx=offright+0.5/3, rely=offup, relwidth=0.5/3, relheight=0.05)
entry_datestart.place(relx=offright+0.5/3, rely=offup+0.05, relwidth=0.5/3, relheight=0.05)
label_dateend.place(relx=offright+1/3, rely=offup, relwidth=0.5/3, relheight=0.05)
entry_dateend.place(relx=offright+1/3, rely=offup+0.05, relwidth=0.5/3, relheight=0.05)
cbutton_PRBs.place(relx=offright+0.5/3, rely=offup+0.1, relwidth=0.5/3, relheight=0.05)
cbutton_Huella.place(relx=offright+0.5/3, rely=offup+.15, relwidth=0.5/3, relheight=0.05)
#cbutton_Huella_post.place(relx=offright+0.5/3, rely=offup+.2, relwidth=0.5/3, relheight=0.05)
cbutton_iomtmp.place(relx=offright+1/3, rely=offup+0.1, relwidth=0.5/3, relheight=0.05)
cbutton_qrets.place(relx=offright+1/3, rely=offup+0.15, relwidth=0.5/3, relheight=0.05)
cbutton_exprets.place(relx=offright+0.5/3, rely=offup+0.2, relwidth=0.5/3, relheight=0.05)
cbutton_qbal.place(relx=offright+1.5/3, rely=offup+0.2, relwidth=0.5/3, relheight=0.05)
cbutton_qbalcosect.place(relx=offright+1/3, rely=offup+0.2, relwidth=0.5/3, relheight=0.05)
cbutton_gmr.place(relx=offright+1.5/3, rely=offup+0.15, relwidth=0.5/3, relheight=0.05)
#cbutton_pbal.place(relx=offright+1.5/3, rely=offup+0.2, relwidth=0.5/3, relheight=0.05)

wd.mainloop()